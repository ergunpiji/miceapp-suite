"""
Haftalık Ödeme Listesi — Genel Müdür için
GET /payments/weekly                   — sayfa
POST /payments/weekly/decide            — onayla / reddet / ertele / yöntem değiştir
POST /payments/settings/weekday         — admin: ödeme günü ayarla
"""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Depends, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from auth import get_current_user, require_admin, require_module, get_company_id
from database import get_db
from models import (
    User, Invoice, Cheque, CreditCardStatement, CreditCard, CreditCardTxn,
    Employee, SalaryPayment, PayrollDecision, SystemSetting,
    BankAccount, BankMovement, CashEntry, ManualPaymentLine, Reference,
    PAYMENT_METHODS,
)
from templates_config import templates


router = APIRouter(prefix="/payments", tags=["payments"])

WEEKDAYS_TR = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
PAYMENT_METHOD_LABELS = dict(PAYMENT_METHODS)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_payment_weekday(db: Session) -> int:
    s = db.query(SystemSetting).filter(SystemSetting.key == "payment_weekday").first()
    if not s:
        return 2
    try:
        v = int(s.value)
        return v if 0 <= v <= 6 else 2
    except (TypeError, ValueError):
        return 2


def _set_payment_weekday(db: Session, weekday: int) -> None:
    s = db.query(SystemSetting).filter(SystemSetting.key == "payment_weekday").first()
    if s:
        s.value = str(weekday)
    else:
        db.add(SystemSetting(key="payment_weekday", value=str(weekday)))


def _next_payment_date(weekday: int, today: Optional[date] = None) -> date:
    if today is None:
        today = date.today()
    days_ahead = (weekday - today.weekday()) % 7
    return today + timedelta(days=days_ahead)


def _current_cycle_key() -> str:
    """ISO hafta anahtarı: 2026-W17 gibi."""
    iso = date.today().isocalendar()
    return f"{iso[0]}-W{iso[1]:02d}"


def _get_cycle_status(db: Session) -> str:
    """Haftalık ödeme döngüsü durumu: 'draft' (hazırlık) | 'submitted' (GM onayında).
    Yeni haftaya geçildiyse otomatik 'draft'a sıfırlanır."""
    s = db.query(SystemSetting).filter(SystemSetting.key == "weekly_cycle_status").first()
    if not s or not s.value:
        return "draft"
    try:
        week_key, status = s.value.split(":", 1)
    except ValueError:
        return "draft"
    if week_key != _current_cycle_key():
        return "draft"  # otomatik hafta resetleme
    return status if status in ("draft", "submitted") else "draft"


def _set_cycle_status(db: Session, status: str) -> None:
    """Mevcut haftaya status yaz."""
    if status not in ("draft", "submitted"):
        raise HTTPException(400, "Geçersiz cycle status")
    val = f"{_current_cycle_key()}:{status}"
    s = db.query(SystemSetting).filter(SystemSetting.key == "weekly_cycle_status").first()
    if s:
        s.value = val
    else:
        db.add(SystemSetting(key="weekly_cycle_status", value=val))


def _show_in_list(item, ref_date: date) -> bool:
    """Listede gösterilecek mi? Onaylananlar görünür kalır (yeşil ışık,
    fiili ödeme yapılınca status değişip filtreden düşer). Reddedilenler
    listeden kaldırılır. Ertelenenler ancak vade geldiğinde tekrar çıkar."""
    d = item.gm_decision
    if d is None:
        return True
    if d == "approved":
        return True
    if d == "rejected":
        return False
    if d == "postponed":
        return bool(item.gm_postpone_until and item.gm_postpone_until <= ref_date)
    return True


def _default_method_for(kalem_type: str, item) -> str:
    """Talimat oluştururken yöntem seçilmediyse kullanılacak default."""
    if kalem_type == "invoice":
        return getattr(item, "payment_method", None) or "banka"
    if kalem_type == "cheque":
        return "banka"  # verilen çek tahsilatı bankadan ödenir
    if kalem_type == "cc_statement":
        return "banka"
    if kalem_type == "manual":
        return getattr(item, "payment_method", None) or "banka"
    return "banka"  # payroll


def _is_actionable(item, ref_date: date) -> bool:
    """Yeni karar verilebilir mi? (Onayla/Reddet butonu + checkbox görünür)
    - Karar verilmemiş kalemler: actionable
    - Ertelenmiş + vade geldi: actionable (yeniden karar)
    - Kısmi onay + vade geldi: actionable (kalan için karar)
    - Tam onay veya henüz vadeye gelmemiş erteleme: not actionable
    """
    d = item.gm_decision
    if d is None:
        return True
    if d == "rejected":
        return False
    pp = getattr(item, "gm_postpone_until", None)
    if d == "postponed":
        return bool(pp and pp <= ref_date)
    if d == "approved":
        # Kısmi onaylı (gm_approved_amount set) ve vade geldi → tekrar karar zamanı
        approved_amt = getattr(item, "gm_approved_amount", None)
        if approved_amt and pp and pp <= ref_date:
            return True
        return False
    return False


def _cash_total(db, company_id: str = None) -> float:
    q_in  = db.query(func.sum(CashEntry.amount)).filter(CashEntry.entry_type == "giris")
    q_out = db.query(func.sum(CashEntry.amount)).filter(CashEntry.entry_type == "cikis")
    if company_id:
        q_in  = q_in.filter(CashEntry.company_id == company_id)
        q_out = q_out.filter(CashEntry.company_id == company_id)
    ins  = q_in.scalar() or 0
    outs = q_out.scalar() or 0
    return ins - outs


def _bank_total(db, company_id: str = None) -> float:
    q = db.query(BankAccount)
    if company_id:
        q = q.filter(BankAccount.company_id == company_id)
    accounts = q.all()
    total = 0.0
    for a in accounts:
        opening = a.opening_balance or 0
        ins = db.query(func.sum(BankMovement.amount)).filter(
            BankMovement.account_id == a.id, BankMovement.movement_type == "giris"
        ).scalar() or 0
        outs = db.query(func.sum(BankMovement.amount)).filter(
            BankMovement.account_id == a.id, BankMovement.movement_type == "cikis"
        ).scalar() or 0
        total += opening + ins - outs
    return total


def _cc_outstanding(db, card_id) -> float:
    unpaid = db.query(func.sum(CreditCardStatement.total_amount)).filter(
        CreditCardStatement.card_id == card_id,
        CreditCardStatement.status == "unpaid",
    ).scalar() or 0
    unassigned = db.query(func.sum(CreditCardTxn.amount)).filter(
        CreditCardTxn.card_id == card_id,
        CreditCardTxn.statement_id == None,  # noqa: E711
        CreditCardTxn.is_refund == False,  # noqa: E712
    ).scalar() or 0
    return unpaid + unassigned


def _payroll_due(db, period: str, company_id: str = None) -> dict:
    q = db.query(Employee).filter(Employee.active == True)  # noqa: E712
    if company_id:
        q = q.filter(Employee.company_id == company_id)
    employees = q.all()
    sp_q = db.query(SalaryPayment).filter(SalaryPayment.period == period)
    if company_id:
        sp_q = sp_q.join(Employee, SalaryPayment.employee_id == Employee.id).filter(
            Employee.company_id == company_id
        )
    paid_ids = {p.employee_id for p in sp_q.all()}
    unpaid = [e for e in employees if e.id not in paid_ids]
    total = sum(e.net_salary or 0 for e in unpaid)
    return {"unpaid_count": len(unpaid), "total": total, "employees": unpaid}


def _method_label(code: Optional[str]) -> str:
    if not code:
        return "—"
    return PAYMENT_METHOD_LABELS.get(code, code)


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@router.get("/weekly", response_class=HTMLResponse, name="weekly_payments")
async def weekly_payments_view(
    request: Request,
    current_user: User = Depends(require_module("payments_weekly")),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    weekday = _get_payment_weekday(db)
    next_date = _next_payment_date(weekday)
    next_payment_after = next_date + timedelta(days=7)  # bir sonraki ödeme günü
    today = date.today()
    period = today.strftime("%Y-%m")

    # Faturalar
    inv_q = db.query(Invoice).filter(
        Invoice.company_id == cid,
        Invoice.invoice_type == "gelen",
        Invoice.status.in_(["approved", "partial"]),
        Invoice.due_date != None,  # noqa: E711
        Invoice.due_date <= next_date,
        Invoice.deleted_at == None,  # noqa: E711
    ).order_by(Invoice.due_date.asc()).all()
    invoices = [i for i in inv_q if i.remaining > 0 and _show_in_list(i, next_date)]

    # Çekler
    chq_q = db.query(Cheque).filter(
        Cheque.company_id == cid,
        Cheque.cheque_type == "verilen",
        Cheque.status == "beklemede",
        Cheque.due_date <= next_date,
    ).order_by(Cheque.due_date.asc()).all()
    cheques = [c for c in chq_q if _show_in_list(c, next_date)]

    # KK Ekstreleri (Çarşamba kuralından bağımsız — kendi vadesi geçerli)
    cc_q = db.query(CreditCardStatement).filter(
        CreditCardStatement.company_id == cid,
        CreditCardStatement.status == "unpaid",
        CreditCardStatement.due_date <= next_date,
    ).order_by(CreditCardStatement.due_date.asc()).all()
    cc_stmts = [s for s in cc_q if _show_in_list(s, next_date)]

    # Maaş — bu ay ödenmemiş aktifler
    payroll_info = _payroll_due(db, period, cid)
    payroll_decision = db.query(PayrollDecision).filter(
        PayrollDecision.company_id == cid,
        PayrollDecision.period == period,
    ).first()
    payroll_show = (
        payroll_info["total"] > 0 and (
            payroll_decision is None
            or _show_in_list(payroll_decision, next_date)
        )
    )

    # Manuel kalemler — açık olanlar (paid/cancelled değil), _show_in_list filtresine göre
    manual_q = db.query(ManualPaymentLine).filter(
        ManualPaymentLine.company_id == cid,
        ManualPaymentLine.status == "open",
    ).order_by(ManualPaymentLine.due_date.asc().nullslast(), ManualPaymentLine.created_at.asc()).all()
    manuals = [m for m in manual_q if _show_in_list(m, next_date)]

    # Birleşik liste — template tek tablo render eder
    items = []
    for inv in invoices:
        items.append({
            "kalem_type": "invoice",
            "kalem_id": inv.id,
            "type_label": "Fatura",
            "type_color": "primary",
            "party": inv.vendor.name if inv.vendor else "—",
            "party_url": f"/vendors/{inv.vendor_id}" if inv.vendor_id else None,
            "ref_no": inv.reference.ref_no if inv.reference else "—",
            "ref_url": f"/references/{inv.ref_id}" if inv.ref_id else None,
            "detail_url": f"/invoices/{inv.id}",
            "due_date": inv.due_date,
            "amount": inv.remaining,
            "method_code": inv.gm_method_override or inv.payment_method,
            "method_label": _method_label(inv.gm_method_override or inv.payment_method),
            "method_override": inv.gm_method_override,
            "gm_decision": inv.gm_decision,
            "gm_postpone_until": inv.gm_postpone_until,
            "actionable": _is_actionable(inv, next_date),
            "gm_decision_note": inv.gm_decision_note,
            "gm_approved_amount": inv.gm_approved_amount,
            "preparer_note": inv.preparer_note,
        })

    for c in cheques:
        party = (c.vendor.name if c.vendor else None) or "—"
        items.append({
            "kalem_type": "cheque",
            "kalem_id": c.id,
            "type_label": "Çek",
            "type_color": "warning",
            "party": party,
            "ref_no": c.cheque_no or "—",
            "ref_url": None,
            "detail_url": "/cheques",
            "due_date": c.due_date,
            "amount": c.amount,
            "method_code": c.gm_method_override or "cek",
            "method_label": _method_label(c.gm_method_override or "cek"),
            "method_override": c.gm_method_override,
            "gm_decision": c.gm_decision,
            "gm_postpone_until": c.gm_postpone_until,
            "actionable": _is_actionable(c, next_date),
            "gm_decision_note": c.gm_decision_note,
            "gm_approved_amount": c.gm_approved_amount,
            "preparer_note": c.preparer_note,
        })

    for s in cc_stmts:
        items.append({
            "kalem_type": "cc_statement",
            "kalem_id": s.id,
            "type_label": "KK Ekstre",
            "type_color": "danger",
            "party": s.card.name if s.card else "—",
            "party_url": f"/credit-cards/{s.card_id}" if s.card_id else None,
            "ref_no": "—",
            "ref_url": None,
            "detail_url": f"/credit-cards/{s.card_id}",
            "due_date": s.due_date,
            "amount": s.total_amount,
            "method_code": s.gm_method_override or "kredi_karti",
            "method_label": _method_label(s.gm_method_override or "kredi_karti"),
            "method_override": s.gm_method_override,
            "gm_decision": s.gm_decision,
            "gm_postpone_until": s.gm_postpone_until,
            "actionable": _is_actionable(s, next_date),
            "gm_decision_note": s.gm_decision_note,
            "gm_approved_amount": s.gm_approved_amount,
            "preparer_note": s.preparer_note,
        })

    for m in manuals:
        items.append({
            "kalem_type": "manual",
            "kalem_id": m.id,
            "type_label": "Manuel",
            "type_color": "dark",
            "party": m.party or m.description or "—",
            "party_url": None,
            "ref_no": "—",
            "ref_url": None,
            "detail_url": None,
            "due_date": m.due_date,
            "amount": m.amount,
            "method_code": m.gm_method_override or m.payment_method,
            "method_label": _method_label(m.gm_method_override or m.payment_method),
            "method_override": m.gm_method_override,
            "gm_decision": m.gm_decision,
            "gm_postpone_until": m.gm_postpone_until,
            "actionable": _is_actionable(m, next_date),
            "gm_decision_note": m.gm_decision_note,
            "gm_approved_amount": m.gm_approved_amount,
            "preparer_note": m.preparer_note,
            "manual_description": m.description,
            "manual_can_delete": (m.gm_decision is None and m.status == "open"),
        })

    if payroll_show:
        method = (payroll_decision.gm_method_override if payroll_decision else None) or "banka"
        items.append({
            "kalem_type": "payroll",
            "kalem_id": period,
            "type_label": "Maaş",
            "type_color": "info",
            "party": f"Personel ({payroll_info['unpaid_count']} kişi)",
            "ref_no": period,
            "ref_url": None,
            "detail_url": "/employees",
            "due_date": None,
            "amount": payroll_info["total"],
            "method_code": method,
            "method_label": _method_label(method),
            "method_override": payroll_decision.gm_method_override if payroll_decision else None,
            "gm_decision": payroll_decision.gm_decision if payroll_decision else None,
            "gm_postpone_until": payroll_decision.gm_postpone_until if payroll_decision else None,
            "actionable": _is_actionable(payroll_decision, next_date) if payroll_decision else True,
            "gm_decision_note": payroll_decision.gm_decision_note if payroll_decision else None,
            "gm_approved_amount": payroll_decision.gm_approved_amount if payroll_decision else None,
            "preparer_note": payroll_decision.preparer_note if payroll_decision else None,
            "party_url": "/employees",
        })

    # Vade tarihine göre sırala (maaş sona)
    items.sort(key=lambda x: (x["due_date"] is None, x["due_date"] or date.max))

    grand_total = sum(it["amount"] for it in items)

    # Özet
    cash_balance = _cash_total(db, cid)
    bank_balance = _bank_total(db, cid)
    cards = db.query(CreditCard).filter(CreditCard.company_id == cid).order_by(CreditCard.name).all()
    card_summary = []
    for c in cards:
        used = _cc_outstanding(db, c.id)
        # Sonraki ödenmemiş ekstrenin vadesi; yoksa kartın statement_day + payment_offset
        # ayarlarından bir sonraki ödeme tarihini hesapla.
        next_stmt = db.query(CreditCardStatement).filter(
            CreditCardStatement.company_id == cid,
            CreditCardStatement.card_id == c.id,
            CreditCardStatement.status == "unpaid",
        ).order_by(CreditCardStatement.due_date.asc()).first()
        if next_stmt:
            next_due = next_stmt.due_date
        elif c.statement_day:
            sd = c.statement_day
            offset = c.payment_offset_days or 10
            t = date.today()
            stmt_month = t.month if t.day <= sd else (t.month % 12 + 1)
            stmt_year = t.year + (1 if (t.day > sd and t.month == 12) else 0)
            try:
                stmt_d = date(stmt_year, stmt_month, sd)
            except ValueError:
                import calendar as _cal
                stmt_d = date(stmt_year, stmt_month, _cal.monthrange(stmt_year, stmt_month)[1])
            next_due = stmt_d + timedelta(days=offset)
        else:
            next_due = None
        card_summary.append({
            "id": c.id, "name": c.name,
            "limit": c.credit_limit or 0,
            "used": used,
            "available": (c.credit_limit or 0) - used,
            "next_due": next_due,
        })
    cc_total_used = sum(c["used"] for c in card_summary)
    cc_total_available = sum(c["available"] for c in card_summary)

    return templates.TemplateResponse(
        "payments/weekly.html",
        {
            "request": request, "current_user": current_user,
            "page_title": "Haftalık Ödeme Listesi",
            "next_date": next_date,
            "next_payment_after": next_payment_after,
            "next_payment_after_iso": next_payment_after.isoformat(),
            "weekday": weekday,
            "weekday_name": WEEKDAYS_TR[weekday],
            "weekdays_tr": WEEKDAYS_TR,
            "items": items,
            "grand_total": grand_total,
            "cash_balance": cash_balance,
            "bank_balance": bank_balance,
            "card_summary": card_summary,
            "cc_total_used": cc_total_used,
            "cc_total_available": cc_total_available,
            "payment_methods": PAYMENT_METHODS,
            "cycle_status": _get_cycle_status(db),
            "active_references": db.query(Reference).filter(
                Reference.company_id == cid,
                Reference.status == "aktif",
            ).order_by(Reference.created_at.desc()).all(),
        },
    )


@router.post("/weekly/decide", name="weekly_payment_decide")
async def weekly_payment_decide(
    kalem_type: str = Form(...),
    kalem_id: str = Form(...),
    action: str = Form(...),
    postpone_date: str = Form(""),
    method_override: str = Form(""),
    note: str = Form(""),
    approved_amount: str = Form(""),
    current_user: User = Depends(require_module("payments_weekly")),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    if action not in ("approve", "reject", "postpone", "method"):
        raise HTTPException(400, "Geçersiz aksiyon")
    cs = _get_cycle_status(db)
    # Onay/Red sadece Genel Müdür yetkisinde
    if action in ("approve", "reject") and not current_user.is_approver:
        raise HTTPException(403, "Onay/Red için Genel Müdür yetkisi gereklidir.")
    # Onay/Red yalnızca liste GM onayına gönderildiğinde mümkün
    if action in ("approve", "reject") and cs != "submitted":
        raise HTTPException(
            409, "Önce listeyi GM onayına gönderin. (Liste şu an hazırlık aşamasında)"
        )
    # Submitted modunda ertele/yöntem değişikliği yalnızca GM yapabilir
    if action in ("postpone", "method") and cs == "submitted" and not current_user.is_approver:
        raise HTTPException(
            403, "Liste GM onayında; değişiklik için GM onayı veya 'Hazırlığa Geri Al' gerekir."
        )

    if kalem_type == "invoice":
        item = db.query(Invoice).get(int(kalem_id))
    elif kalem_type == "cheque":
        item = db.query(Cheque).get(int(kalem_id))
    elif kalem_type == "cc_statement":
        item = db.query(CreditCardStatement).get(int(kalem_id))
    elif kalem_type == "manual":
        item = db.query(ManualPaymentLine).get(int(kalem_id))
    elif kalem_type == "payroll":
        item = db.query(PayrollDecision).filter(
            PayrollDecision.company_id == cid,
            PayrollDecision.period == kalem_id,
        ).first()
        if not item:
            item = PayrollDecision(period=kalem_id, company_id=cid)
            db.add(item)
            db.flush()
    else:
        raise HTTPException(400, "Geçersiz kalem tipi")

    if not item:
        raise HTTPException(404, "Kalem bulunamadı")

    now = datetime.utcnow()
    note_clean = (note or "").strip() or None
    if action == "approve":
        # Kalemin tam tutarı (kısmi onayda kıyas için)
        if kalem_type == "invoice":
            full_amount = item.remaining
        elif kalem_type == "cheque":
            full_amount = item.amount
        elif kalem_type == "cc_statement":
            full_amount = item.total_amount
        elif kalem_type == "manual":
            full_amount = item.amount
        else:  # payroll
            full_amount = _payroll_due(db, item.period)["total"]

        # approved_amount opsiyonel: boş veya tam tutar → tam onay; daha az → kısmi
        try:
            req_amt = float(approved_amount) if (approved_amount or "").strip() else full_amount
        except (ValueError, TypeError) as exc:
            raise HTTPException(400, "Geçersiz onay tutarı") from exc
        if req_amt <= 0 or req_amt > round(full_amount + 0.01, 2):
            raise HTTPException(400, "Onay tutarı 0 ile tam tutar arasında olmalı")

        is_partial = req_amt < round(full_amount - 0.01, 2)
        item.gm_decision = "approved"
        item.gm_decision_at = now
        item.gm_decision_by = current_user.id
        item.gm_decision_note = note_clean
        if is_partial:
            # Kısmi onay: kalan kısmı için erteleme tarihi gerekli
            try:
                pd = date.fromisoformat(postpone_date)
            except (ValueError, TypeError) as exc:
                raise HTTPException(400, "Kısmi onay için erteleme tarihi gerekli") from exc
            if pd <= date.today():
                raise HTTPException(400, "Erteleme tarihi gelecekte olmalı")
            item.gm_approved_amount = round(req_amt, 2)
            item.gm_postpone_until = pd
        else:
            item.gm_approved_amount = None
            item.gm_postpone_until = None

        # Onaylanan tutar için PaymentInstruction yarat (operatör hedef hesabı sonra seçer)
        from routers.payment_instructions import create_instruction
        method_for_instr = item.gm_method_override or _default_method_for(kalem_type, item)
        source_id_or_period = item.period if kalem_type == "payroll" else item.id
        create_instruction(
            db,
            source_type=kalem_type,
            source_id_or_period=source_id_or_period,
            amount=round(req_amt, 2),
            payment_method=method_for_instr,
            note=note_clean or "",
            current_user=current_user,
            company_id=cid,
        )
    elif action == "reject":
        item.gm_decision = "rejected"
        item.gm_decision_at = now
        item.gm_decision_by = current_user.id
        item.gm_postpone_until = None
        item.gm_decision_note = note_clean
        # Pending instruction varsa iptal et
        from routers.payment_instructions import cancel_pending_for_source
        source_id_or_period = item.period if kalem_type == "payroll" else item.id
        cancel_pending_for_source(
            db, kalem_type, source_id_or_period, current_user.id,
            f"GM tarafından reddedildi" + (f": {note_clean}" if note_clean else ""),
        )
    elif action == "postpone":
        try:
            new_date = date.fromisoformat(postpone_date)
        except (ValueError, TypeError):
            raise HTTPException(400, "Geçersiz erteleme tarihi")
        if new_date <= date.today():
            raise HTTPException(400, "Erteleme tarihi gelecekte olmalı")
        item.gm_decision = "postponed"
        item.gm_decision_at = now
        item.gm_decision_by = current_user.id
        item.gm_postpone_until = new_date
    elif action == "method":
        valid = {m[0] for m in PAYMENT_METHODS}
        if method_override not in valid:
            raise HTTPException(400, "Geçersiz ödeme yöntemi")
        item.gm_method_override = method_override

    db.commit()
    return RedirectResponse(url="/payments/weekly", status_code=303)


@router.get("/weekly/export", name="weekly_payment_export")
async def weekly_payment_export(
    current_user: User = Depends(require_module("payments_weekly")),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    """Haftalık ödeme listesini Excel olarak indir."""
    weekday = _get_payment_weekday(db)
    next_date = _next_payment_date(weekday)
    today = date.today()
    period = today.strftime("%Y-%m")

    # View ile aynı filtreleme — yeniden hesapla
    invoices = db.query(Invoice).filter(
        Invoice.company_id == cid,
        Invoice.invoice_type == "gelen",
        Invoice.status.in_(["approved", "partial"]),
        Invoice.due_date.is_not(None),
        Invoice.due_date <= next_date,
        Invoice.deleted_at == None,  # noqa: E711
    ).order_by(Invoice.due_date.asc()).all()
    invoices = [i for i in invoices if i.remaining > 0 and _show_in_list(i, next_date)]

    cheques = db.query(Cheque).filter(
        Cheque.company_id == cid,
        Cheque.cheque_type == "verilen",
        Cheque.status == "beklemede",
        Cheque.due_date <= next_date,
    ).order_by(Cheque.due_date.asc()).all()
    cheques = [c for c in cheques if _show_in_list(c, next_date)]

    cc_stmts = db.query(CreditCardStatement).filter(
        CreditCardStatement.company_id == cid,
        CreditCardStatement.status == "unpaid",
        CreditCardStatement.due_date <= next_date,
    ).order_by(CreditCardStatement.due_date.asc()).all()
    cc_stmts = [s for s in cc_stmts if _show_in_list(s, next_date)]

    payroll_info = _payroll_due(db, period, cid)
    payroll_decision = db.query(PayrollDecision).filter(
        PayrollDecision.company_id == cid,
        PayrollDecision.period == period,
    ).first()
    payroll_show = (
        payroll_info["total"] > 0 and (
            payroll_decision is None or _show_in_list(payroll_decision, next_date)
        )
    )

    decision_label = {
        "approved": "Onaylandı",
        "rejected": "Reddedildi",
        "postponed": "Ertelendi",
    }

    def _decision_text(d):
        return decision_label.get(d, "Bekliyor")

    # ----- Stil sabitleri -----
    NAVY = "1A3A5C"     # Brand ana renk
    NAVY_2 = "1E5F8C"   # Brand sekonder
    GRAY = "F1F5F9"     # Açık gri
    LIGHT_BG = "F8FAFC" # Çok açık arka plan
    GREEN = "16A34A"
    RED = "DC2626"
    ORANGE = "F59E0B"

    from openpyxl.styles import Border, Side
    thin = Side(border_style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt = '#,##0.00 [$₺-tr-TR]'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Odeme Listesi"
    ws.sheet_view.showGridLines = False

    HEADERS = ["Tip", "Tedarikçi / İlgili", "Referans", "Vade",
               "Tutar", "Yöntem", "Durum", "Onaylanan", "Not"]
    NCOLS = len(HEADERS)
    last_letter = chr(ord('A') + NCOLS - 1)  # 'I'

    # ===== BAŞLIK BLOĞU =====
    # Şirket marka satırı
    ws.row_dimensions[1].height = 32
    ws.merge_cells(f"A1:{last_letter}1")
    c1 = ws["A1"]
    from templates_config import company
    c1.value = f"{company('short_name', 'PRİZMATİK')} — Finans Yönetim Programı"
    c1.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    c1.fill = PatternFill("solid", fgColor=NAVY)
    c1.alignment = Alignment(horizontal="center", vertical="center")

    # Liste başlığı
    ws.row_dimensions[2].height = 28
    ws.merge_cells(f"A2:{last_letter}2")
    c2 = ws["A2"]
    c2.value = f"Haftalık Ödeme Listesi  —  {next_date.strftime('%d.%m.%Y')}"
    c2.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    c2.fill = PatternFill("solid", fgColor=NAVY_2)
    c2.alignment = Alignment(horizontal="center", vertical="center")

    # Meta satırı (durum + tarih + toplam yer tutucu)
    ws.row_dimensions[3].height = 22
    ws.merge_cells(f"A3:D3")
    ws["A3"] = f"Hazırlandı: {today.strftime('%d.%m.%Y %H:%M')}    Hazırlayan: {current_user.name}"
    ws["A3"].font = Font(size=10, italic=True, color="64748B")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"E3:{last_letter}3")
    ws[f"E3"] = ""  # toplam birazdan basılacak
    ws[f"E3"].alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # Bölüm helper
    def _section_header(row_idx: int, label: str, count: int, total: float):
        ws.row_dimensions[row_idx].height = 22
        ws.merge_cells(f"A{row_idx}:{last_letter}{row_idx}")
        cell = ws.cell(row=row_idx, column=1)
        cell.value = f"  {label}  ·  {count} kalem  ·  Toplam {total:,.2f} ₺".replace(",", "X").replace(".", ",").replace("X", ".")
        cell.font = Font(size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY_2)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    def _write_table_header(row_idx: int):
        ws.row_dimensions[row_idx].height = 24
        for i, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=row_idx, column=i, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    def _write_data_row(row_idx: int, data: list, alt: bool, decision: Optional[str]):
        bg = PatternFill("solid", fgColor=LIGHT_BG) if alt else PatternFill(fill_type=None)
        for i, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=i, value=val)
            cell.font = Font(size=10)
            cell.border = border
            if not alt:
                cell.fill = PatternFill(fill_type=None)
            else:
                cell.fill = bg
            # Hizalama kuralları
            if i in (5, 8):  # Tutar, Onaylanan
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(val, (int, float)):
                    cell.number_format = money_fmt
            elif i == 4:  # Vade
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif i == 7:  # Durum
                cell.alignment = Alignment(horizontal="center", vertical="center")
                color_map = {
                    "Onaylandı": GREEN, "Reddedildi": RED,
                    "Ertelendi": "64748B", "Bekliyor": ORANGE,
                }
                if decision in (None, "approved", "rejected", "postponed") and val in color_map:
                    cell.font = Font(size=10, bold=True, color=color_map[val])
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

    # ===== BÖLÜMLER =====
    cur_row = 5  # başlık + meta + boşluk sonrası
    ws.row_dimensions[4].height = 8  # ince boşluk
    grand_total = 0.0

    sections = [
        ("FATURALAR", "Fatura", invoices, lambda inv: [
            "Fatura",
            inv.vendor.name if inv.vendor else "—",
            inv.reference.ref_no if inv.reference else "—",
            inv.due_date.strftime("%d.%m.%Y") if inv.due_date else "",
            float(inv.remaining or 0),
            _method_label(inv.gm_method_override or inv.payment_method),
            _decision_text(inv.gm_decision),
            float(inv.gm_approved_amount) if inv.gm_approved_amount else "",
            inv.gm_decision_note or "",
        ]),
        ("ÇEKLER", "Cek", cheques, lambda c: [
            "Çek",
            (c.vendor.name if c.vendor else "") or "—",
            c.cheque_no or "—",
            c.due_date.strftime("%d.%m.%Y") if c.due_date else "",
            float(c.amount or 0),
            _method_label(c.gm_method_override or "cek"),
            _decision_text(c.gm_decision),
            float(c.gm_approved_amount) if c.gm_approved_amount else "",
            c.gm_decision_note or "",
        ]),
        ("KREDİ KARTI EKSTRELERİ", "KK", cc_stmts, lambda s: [
            "KK Ekstre",
            s.card.name if s.card else "—",
            "—",
            s.due_date.strftime("%d.%m.%Y") if s.due_date else "",
            float(s.total_amount or 0),
            _method_label(s.gm_method_override or "kredi_karti"),
            _decision_text(s.gm_decision),
            float(s.gm_approved_amount) if s.gm_approved_amount else "",
            s.gm_decision_note or "",
        ]),
    ]

    payroll_rows = []
    if payroll_show:
        method_p = (payroll_decision.gm_method_override
                    if payroll_decision else None) or "banka"
        approved_amt_p = (payroll_decision.gm_approved_amount
                          if payroll_decision and payroll_decision.gm_approved_amount
                          else "")
        note_val_p = (payroll_decision.gm_decision_note
                      if payroll_decision else None) or ""
        payroll_rows.append([
            "Maaş",
            f"Personel ({payroll_info['unpaid_count']} kişi)",
            period, "",
            float(payroll_info["total"] or 0),
            _method_label(method_p),
            _decision_text(payroll_decision.gm_decision if payroll_decision else None),
            float(approved_amt_p) if approved_amt_p else "",
            note_val_p,
        ])

    for section_label, _key, items_list, builder in sections:
        if not items_list:
            continue
        section_total = 0.0
        rows_data = []
        for itm in items_list:
            data = builder(itm)
            section_total += data[4] or 0
            rows_data.append((data, itm))

        _section_header(cur_row, section_label, len(items_list), section_total)
        cur_row += 1
        _write_table_header(cur_row)
        cur_row += 1
        for idx, (data, itm) in enumerate(rows_data):
            decision = getattr(itm, "gm_decision", None)
            _write_data_row(cur_row, data, alt=(idx % 2 == 1), decision=decision)
            cur_row += 1
        # Bölüm alt-toplamı
        ws.row_dimensions[cur_row].height = 20
        ws.merge_cells(f"A{cur_row}:D{cur_row}")
        cell = ws.cell(row=cur_row, column=1, value=f"  {section_label} Ara Toplam")
        cell.font = Font(size=10, bold=True, color=NAVY)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.fill = PatternFill("solid", fgColor=GRAY)
        amt_cell = ws.cell(row=cur_row, column=5, value=section_total)
        amt_cell.font = Font(size=10, bold=True, color=NAVY)
        amt_cell.alignment = Alignment(horizontal="right", vertical="center")
        amt_cell.number_format = money_fmt
        amt_cell.fill = PatternFill("solid", fgColor=GRAY)
        for c_idx in range(6, NCOLS + 1):
            ws.cell(row=cur_row, column=c_idx).fill = PatternFill("solid", fgColor=GRAY)
        cur_row += 1
        # Bölümler arası boşluk
        ws.row_dimensions[cur_row].height = 6
        cur_row += 1

        grand_total += section_total

    # MAAŞ bölümü
    if payroll_rows:
        section_total = sum(r[4] or 0 for r in payroll_rows)
        _section_header(cur_row, "PERSONEL MAAŞI", len(payroll_rows), section_total)
        cur_row += 1
        _write_table_header(cur_row)
        cur_row += 1
        for idx, data in enumerate(payroll_rows):
            decision = payroll_decision.gm_decision if payroll_decision else None
            _write_data_row(cur_row, data, alt=(idx % 2 == 1), decision=decision)
            cur_row += 1
        grand_total += section_total

    # ===== GENEL TOPLAM =====
    ws.row_dimensions[cur_row].height = 6
    cur_row += 1
    ws.row_dimensions[cur_row].height = 30
    ws.merge_cells(f"A{cur_row}:D{cur_row}")
    total_label = ws.cell(row=cur_row, column=1, value="  GENEL TOPLAM")
    total_label.font = Font(size=13, bold=True, color="FFFFFF")
    total_label.fill = PatternFill("solid", fgColor=NAVY)
    total_label.alignment = Alignment(horizontal="right", vertical="center")
    total_amount = ws.cell(row=cur_row, column=5, value=grand_total)
    total_amount.font = Font(size=14, bold=True, color="FFFFFF")
    total_amount.fill = PatternFill("solid", fgColor=NAVY)
    total_amount.alignment = Alignment(horizontal="right", vertical="center")
    total_amount.number_format = money_fmt
    for c_idx in range(6, NCOLS + 1):
        c = ws.cell(row=cur_row, column=c_idx)
        c.fill = PatternFill("solid", fgColor=NAVY)
    # Üst satırdaki meta E3'e toplamı yaz
    ws[f"E3"] = f"Toplam: {grand_total:,.2f} ₺".replace(",", "X").replace(".", ",").replace("X", ".")
    ws[f"E3"].font = Font(size=11, bold=True, color="DC2626")

    # ===== Sütun genişlikleri =====
    widths = {"A": 14, "B": 36, "C": 18, "D": 13, "E": 18,
              "F": 18, "G": 14, "H": 16, "I": 36}
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w

    # Yazdırma ayarları
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"odeme-listesi-{next_date.isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/weekly/bulk-approve", name="weekly_payment_bulk_approve")
async def weekly_payment_bulk_approve(
    request: Request,
    current_user: User = Depends(require_module("payments_weekly")),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    if not current_user.is_approver:
        raise HTTPException(403, "Toplu onay için Genel Müdür yetkisi gereklidir.")
    if _get_cycle_status(db) != "submitted":
        raise HTTPException(409, "Önce listeyi GM onayına gönderin.")
    form = await request.form()
    entries = form.getlist("items")
    note_clean = (form.get("note", "") or "").strip() or None
    now = datetime.utcnow()
    count = 0
    for entry in entries:
        if ":" not in entry:
            continue
        kalem_type, kalem_id = entry.split(":", 1)
        if kalem_type == "invoice":
            item = db.query(Invoice).get(int(kalem_id))
        elif kalem_type == "cheque":
            item = db.query(Cheque).get(int(kalem_id))
        elif kalem_type == "cc_statement":
            item = db.query(CreditCardStatement).get(int(kalem_id))
        elif kalem_type == "manual":
            item = db.query(ManualPaymentLine).get(int(kalem_id))
        elif kalem_type == "payroll":
            item = db.query(PayrollDecision).filter(
                PayrollDecision.company_id == cid,
                PayrollDecision.period == kalem_id,
            ).first()
            if not item:
                item = PayrollDecision(period=kalem_id, company_id=cid)
                db.add(item)
                db.flush()
        else:
            continue
        if not item:
            continue
        item.gm_decision = "approved"
        item.gm_decision_at = now
        item.gm_decision_by = current_user.id
        item.gm_postpone_until = None
        if note_clean:
            item.gm_decision_note = note_clean

        # Tam onay tutarı için instruction yarat (kısmi onay decide endpoint'ine özel)
        from routers.payment_instructions import create_instruction, get_pending_for_source
        if kalem_type == "invoice":
            full_amount = item.remaining
        elif kalem_type == "cheque":
            full_amount = item.amount
        elif kalem_type == "cc_statement":
            full_amount = item.total_amount
        elif kalem_type == "manual":
            full_amount = item.amount
        else:
            full_amount = _payroll_due(db, item.period, cid)["total"]
        if full_amount and full_amount > 0:
            source_id_or_period = item.period if kalem_type == "payroll" else item.id
            if not get_pending_for_source(db, kalem_type, source_id_or_period):
                method_for_instr = item.gm_method_override or _default_method_for(kalem_type, item)
                create_instruction(
                    db, source_type=kalem_type, source_id_or_period=source_id_or_period,
                    amount=round(full_amount, 2), payment_method=method_for_instr,
                    note=note_clean or "", current_user=current_user,
                    company_id=cid,
                )
        count += 1
    db.commit()
    return RedirectResponse(url="/payments/weekly", status_code=303)


@router.post("/weekly/preparer-note", name="weekly_payment_preparer_note")
async def weekly_payment_preparer_note(
    kalem_type: str = Form(...),
    kalem_id: str = Form(...),
    note: str = Form(""),
    current_user: User = Depends(require_module("payments_weekly")),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    """Listeyi hazırlayan kişinin GM'e yönelik notu — herkes kaydedebilir.
    GM onayına gönderildikten sonra yalnızca GM düzenleyebilir."""
    if _get_cycle_status(db) == "submitted" and not current_user.is_approver:
        raise HTTPException(
            403, "Liste GM onayında; not değişikliği için GM yetkisi veya 'Hazırlığa Geri Al' gerekir."
        )
    note_clean = (note or "").strip() or None
    if kalem_type == "invoice":
        item = db.query(Invoice).get(int(kalem_id))
    elif kalem_type == "cheque":
        item = db.query(Cheque).get(int(kalem_id))
    elif kalem_type == "cc_statement":
        item = db.query(CreditCardStatement).get(int(kalem_id))
    elif kalem_type == "manual":
        item = db.query(ManualPaymentLine).get(int(kalem_id))
    elif kalem_type == "payroll":
        item = db.query(PayrollDecision).filter(
            PayrollDecision.company_id == cid,
            PayrollDecision.period == kalem_id,
        ).first()
        if not item:
            item = PayrollDecision(period=kalem_id, company_id=cid)
            db.add(item)
            db.flush()
    else:
        raise HTTPException(400, "Geçersiz kalem tipi")
    if not item:
        raise HTTPException(404, "Kalem bulunamadı")
    item.preparer_note = note_clean
    db.commit()
    return RedirectResponse(url="/payments/weekly", status_code=303)


# ---------------------------------------------------------------------------
# Manuel ödeme kalemi (ekle / sil)
# ---------------------------------------------------------------------------

@router.post("/weekly/manual/new", name="weekly_payment_manual_new")
async def weekly_manual_new(
    description: str = Form(...),
    amount: float = Form(...),
    payment_method: str = Form("banka"),
    party: str = Form(""),
    due_date: str = Form(""),
    ref_id: str = Form(""),
    notes: str = Form(""),
    current_user: User = Depends(require_module("payments_weekly")),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    """Liste hazırlık aşamasındayken kullanıcı manuel ödeme kalemi ekler.
    Referans seçilirse o referansa atanır; seçilmezse ödeme yapıldığında
    genel giderlere düşer (apply_manual_payment içinde GeneralExpense yaratılır)."""
    cs = _get_cycle_status(db)
    if cs == "submitted" and not current_user.is_approver:
        raise HTTPException(
            409, "Liste GM onayında; manuel kalem eklemek için GM yetkisi gerekir."
        )
    if not description.strip():
        raise HTTPException(400, "Açıklama zorunlu")
    if amount <= 0:
        raise HTTPException(400, "Tutar 0'dan büyük olmalı")
    valid = {m[0] for m in PAYMENT_METHODS}
    if payment_method not in valid:
        raise HTTPException(400, "Geçersiz ödeme yöntemi")
    dd = None
    if due_date.strip():
        try:
            dd = date.fromisoformat(due_date.strip())
        except ValueError:
            raise HTTPException(400, "Geçersiz vade tarihi")
    rid = None
    if (ref_id or "").strip():
        try:
            rid = int(ref_id)
            if not db.query(Reference).filter(Reference.company_id == cid, Reference.id == rid).first():
                rid = None
        except (ValueError, TypeError):
            rid = None
    line = ManualPaymentLine(
        description=description.strip(),
        party=party.strip() or None,
        amount=round(amount, 2),
        payment_method=payment_method,
        due_date=dd,
        ref_id=rid,
        notes=notes.strip() or None,
        status="open",
        created_by=current_user.id,
        company_id=cid,
    )
    db.add(line)
    db.commit()
    return RedirectResponse(url="/payments/weekly", status_code=303)


@router.post("/weekly/manual/{line_id}/delete", name="weekly_payment_manual_delete")
async def weekly_manual_delete(
    line_id: str,
    current_user: User = Depends(require_module("payments_weekly")),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    """Manuel kalemi sil — operatör hazırlık'ta, GM submitted'ta da silebilir."""
    cs = _get_cycle_status(db)
    if cs == "submitted" and not current_user.is_approver:
        raise HTTPException(409, "Liste GM onayında; silme için GM yetkisi gerekir")
    line = db.query(ManualPaymentLine).filter(
        ManualPaymentLine.company_id == cid,
        ManualPaymentLine.id == line_id,
    ).first()
    if not line:
        raise HTTPException(404, "Manuel kalem bulunamadı")
    if line.gm_decision is not None or line.status != "open":
        raise HTTPException(400, "Karar verilmiş veya işlenmiş kalem silinemez")
    db.delete(line)
    db.commit()
    return RedirectResponse(url="/payments/weekly", status_code=303)


@router.post("/weekly/submit", name="weekly_payment_submit")
async def weekly_payment_submit(
    current_user: User = Depends(require_module("payments_weekly")),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    """Listeyi GM onayına gönder (draft → submitted). GM kendisi gönderemez.
    Tüm aktif Genel Müdür'lere bilgilendirme e-postası gönderilir."""
    import os
    from email_helper import send_email

    if current_user.is_approver and not current_user.is_admin:
        raise HTTPException(403, "Listeyi GM kendisine gönderemez; operatör hazırlar.")
    if _get_cycle_status(db) != "draft":
        raise HTTPException(409, "Liste zaten GM onayında")
    _set_cycle_status(db, "submitted")
    db.commit()

    # GM(ler)e e-posta gönder — özet bilgilerle
    try:
        gms = db.query(User).filter(
            User.is_approver == True,  # noqa: E712
            User.active == True,  # noqa: E712
        ).all()
        gm_emails = [u.email for u in gms if u.email]
        if gm_emails:
            weekday = _get_payment_weekday(db)
            next_date = _next_payment_date(weekday)
            today = date.today()
            period = today.strftime("%Y-%m")

            # Özet hesapla — view ile aynı filtreleri kullan
            inv_q = db.query(Invoice).filter(
                Invoice.company_id == cid,
                Invoice.invoice_type == "gelen",
                Invoice.status.in_(["approved", "partial"]),
                Invoice.due_date.is_not(None),
                Invoice.due_date <= next_date,
                Invoice.deleted_at == None,  # noqa: E711
            ).all()
            invoices = [i for i in inv_q if i.remaining > 0 and _show_in_list(i, next_date)]
            chq_q = db.query(Cheque).filter(
                Cheque.company_id == cid,
                Cheque.cheque_type == "verilen",
                Cheque.status == "beklemede",
                Cheque.due_date <= next_date,
            ).all()
            cheques = [c for c in chq_q if _show_in_list(c, next_date)]
            cc_q = db.query(CreditCardStatement).filter(
                CreditCardStatement.company_id == cid,
                CreditCardStatement.status == "unpaid",
                CreditCardStatement.due_date <= next_date,
            ).all()
            cc_stmts = [s for s in cc_q if _show_in_list(s, next_date)]
            payroll_info = _payroll_due(db, period, cid)
            pd = db.query(PayrollDecision).filter(
                PayrollDecision.company_id == cid,
                PayrollDecision.period == period,
            ).first()
            payroll_show = payroll_info["total"] > 0 and (pd is None or _show_in_list(pd, next_date))
            manuals = [m for m in db.query(ManualPaymentLine).filter(
                ManualPaymentLine.company_id == cid,
                ManualPaymentLine.status == "open",
            ).all() if _show_in_list(m, next_date)]

            inv_t = sum(i.remaining for i in invoices)
            chq_t = sum(c.amount for c in cheques)
            cc_t = sum(s.total_amount for s in cc_stmts)
            mn_t = sum(m.amount for m in manuals)
            pr_t = payroll_info["total"] if payroll_show else 0
            grand = inv_t + chq_t + cc_t + mn_t + pr_t

            def _fmt(n):
                return f"₺{n:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

            app_url = (os.environ.get("APP_URL") or "").rstrip("/")
            link = f"{app_url}/payments/weekly" if app_url else "/payments/weekly"

            from templates_config import company as _company
            company_name = _company('short_name') or _company('name') or "Prizma Finans"
            brand_color = _company('brand_color', '#1A3A5C')

            subject = f"{company_name} — Haftalık Ödeme Listesi onayınız bekliyor ({next_date.strftime('%d.%m.%Y')})"
            html = f"""
<div style="font-family:'Segoe UI',Tahoma,Arial,sans-serif;max-width:640px;margin:auto;background:#f1f5f9;padding:24px;">
  <div style="background:{brand_color};color:#fff;padding:18px 22px;border-radius:6px 6px 0 0;">
    <div style="font-size:18px;font-weight:bold;">{company_name} — Finans Yönetim Programı</div>
  </div>
  <div style="background:#1E5F8C;color:#fff;padding:14px 22px;">
    <div style="font-size:15px;font-weight:bold;">Haftalık Ödeme Listesi Onayınız Bekliyor</div>
    <div style="font-size:13px;opacity:.9;margin-top:4px;">Sonraki ödeme günü: {next_date.strftime('%d.%m.%Y')}</div>
  </div>
  <div style="background:#fff;padding:22px;">
    <p style="margin:0 0 14px;color:#1e293b;">Sayın Genel Müdür,</p>
    <p style="margin:0 0 14px;color:#334155;">
      Operatör ({current_user.name}) bu haftaki ödeme listesini hazırladı ve onayınıza gönderdi.
    </p>
    <table style="width:100%;border-collapse:collapse;margin:16px 0;">
      <thead>
        <tr style="background:#f0f4f8;">
          <th style="text-align:left;padding:8px 12px;font-size:12px;color:#64748b;">Kalem Türü</th>
          <th style="text-align:right;padding:8px 12px;font-size:12px;color:#64748b;">Adet</th>
          <th style="text-align:right;padding:8px 12px;font-size:12px;color:#64748b;">Tutar</th>
        </tr>
      </thead>
      <tbody>
        <tr><td style="padding:8px 12px;border-top:1px solid #e2e8f0;">Faturalar</td>
            <td style="padding:8px 12px;text-align:right;border-top:1px solid #e2e8f0;">{len(invoices)}</td>
            <td style="padding:8px 12px;text-align:right;border-top:1px solid #e2e8f0;">{_fmt(inv_t)}</td></tr>
        <tr><td style="padding:8px 12px;border-top:1px solid #e2e8f0;">Çekler</td>
            <td style="padding:8px 12px;text-align:right;border-top:1px solid #e2e8f0;">{len(cheques)}</td>
            <td style="padding:8px 12px;text-align:right;border-top:1px solid #e2e8f0;">{_fmt(chq_t)}</td></tr>
        <tr><td style="padding:8px 12px;border-top:1px solid #e2e8f0;">KK Ekstreleri</td>
            <td style="padding:8px 12px;text-align:right;border-top:1px solid #e2e8f0;">{len(cc_stmts)}</td>
            <td style="padding:8px 12px;text-align:right;border-top:1px solid #e2e8f0;">{_fmt(cc_t)}</td></tr>
        <tr><td style="padding:8px 12px;border-top:1px solid #e2e8f0;">Manuel</td>
            <td style="padding:8px 12px;text-align:right;border-top:1px solid #e2e8f0;">{len(manuals)}</td>
            <td style="padding:8px 12px;text-align:right;border-top:1px solid #e2e8f0;">{_fmt(mn_t)}</td></tr>
        <tr><td style="padding:8px 12px;border-top:1px solid #e2e8f0;">Personel Maaşı</td>
            <td style="padding:8px 12px;text-align:right;border-top:1px solid #e2e8f0;">{'1' if payroll_show else '0'}</td>
            <td style="padding:8px 12px;text-align:right;border-top:1px solid #e2e8f0;">{_fmt(pr_t)}</td></tr>
        <tr style="background:{brand_color};color:#fff;font-weight:bold;">
          <td style="padding:10px 12px;">GENEL TOPLAM</td>
          <td></td>
          <td style="padding:10px 12px;text-align:right;">{_fmt(grand)}</td>
        </tr>
      </tbody>
    </table>
    <div style="text-align:center;margin:22px 0 8px;">
      <a href="{link}" style="display:inline-block;background:#16a34a;color:#fff;padding:12px 26px;border-radius:6px;text-decoration:none;font-weight:bold;">
        Listeyi Aç ve Onayla
      </a>
    </div>
    <p style="margin:14px 0 0;color:#64748b;font-size:12px;">
      Bu otomatik bir bildirimdir. Liste, sistemde "GM Onayında" durumuna geçirildi;
      onay/red kararlarınızı verebilirsiniz.
    </p>
  </div>
</div>
"""
            text_body = (
                f"Sayın Genel Müdür,\n\n"
                f"Operatör ({current_user.name}) {next_date.strftime('%d.%m.%Y')} ödeme günü için "
                f"hazırlanan listeyi onayınıza gönderdi.\n\n"
                f"Toplam: {_fmt(grand)} ({len(invoices)} fatura, {len(cheques)} çek, "
                f"{len(cc_stmts)} KK ekstresi, {len(manuals)} manuel"
                f"{', maaş ödemesi' if payroll_show else ''})\n\n"
                f"Listeyi açmak için: {link}\n"
            )
            send_email(gm_emails, subject, html, text_body)
    except Exception as exc:  # noqa: BLE001
        # Email hatası submit'i bozmasın
        print(f"[submit-email] hata: {exc}", flush=True)

    return RedirectResponse(url="/payments/weekly", status_code=303)


@router.post("/weekly/unsubmit", name="weekly_payment_unsubmit")
async def weekly_payment_unsubmit(
    current_user: User = Depends(require_module("payments_weekly")),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    """Onaya gönderilmiş listeyi geri çek (sadece GM/Admin)."""
    if not (current_user.is_admin or current_user.is_approver):
        raise HTTPException(403, "Sadece GM/Admin geri çekebilir")
    _set_cycle_status(db, "draft")
    db.commit()
    return RedirectResponse(url="/payments/weekly", status_code=303)


@router.post("/settings/weekday", name="weekly_payment_set_weekday")
async def set_payment_weekday(
    weekday: int = Form(...),
    current_user: User = Depends(require_module("payments_weekly", edit=True)),
    db: Session = Depends(get_db),
):
    if not (0 <= weekday <= 6):
        raise HTTPException(400, "Geçersiz gün (0-6 olmalı)")
    _set_payment_weekday(db, weekday)
    db.commit()
    return RedirectResponse(url="/payments/weekly", status_code=303)
