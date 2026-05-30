"""
E-Defter — yevmiye + büyük defter simülasyonu.
Veri hazırlığı + görüntüleme + Excel export. Mali mühür gerekmez.
GİB'e resmi gönderim mali mühür sonrası açılır (şu an disabled).
"""
import io
from datetime import date, datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi import APIRouter, Depends, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy import extract
from sqlalchemy.orm import Session

from auth import require_admin, require_mudur, get_company_id
from database import get_db
from models import (
    User, Invoice, GeneralExpense, BankMovement, CashEntry, Cheque,
    SystemSetting,
)
from templates_config import templates


router = APIRouter(prefix="/edefter", tags=["edefter"])


def _require_module_active(db: Session):
    s = db.query(SystemSetting).filter(SystemSetting.key == "module_edefter_enabled").first()
    if not (s and s.value == "1"):
        from fastapi import HTTPException
        raise HTTPException(404, "E-Defter modülü aktif değil")


# ---------------------------------------------------------------------------
# Hesap planı (basitleştirilmiş — TDHP'nin webapp ile uyumlu kısmı)
# ---------------------------------------------------------------------------

ACCOUNTS = {
    "100": "Kasa",
    "102": "Bankalar",
    "120": "Alıcılar (Müşteri)",
    "153": "Ticari Mallar",
    "320": "Satıcılar (Tedarikçi)",
    "335": "Personele Borçlar",
    "391": "Hesaplanan KDV",
    "191": "İndirilecek KDV",
    "600": "Yurtiçi Satışlar",
    "740": "Hizmet Üretim Maliyeti",
    "770": "Genel Yönetim Giderleri",
    "780": "Finansman Giderleri",
}


# ---------------------------------------------------------------------------
# Yevmiye girişleri (otomatik üretim — webapp verisinden)
# ---------------------------------------------------------------------------

def _build_journal(db: Session, year: int, month: int, cid: int) -> list:
    """Belirli ay için yevmiye satırlarını üret.
    Her invoice / expense / cash / bank movement için 2-3 satırlık yevmiye fişi.
    Format: [{date, doc_no, description, account_code, account_name, debit, credit, ref_type, ref_id}]"""
    entries = []
    seq = 1

    # Faturalar
    invoices = db.query(Invoice).filter(
        Invoice.company_id == cid,
        Invoice.status.in_(["approved", "partial", "paid"]),
        extract("year", Invoice.invoice_date) == year,
        extract("month", Invoice.invoice_date) == month,
    ).all()
    for inv in invoices:
        if not inv.invoice_date:
            continue
        total = inv.amount * (1 + inv.vat_rate)
        vat_amt = inv.amount * inv.vat_rate
        is_income = inv.invoice_type in ("kesilen", "komisyon")
        doc_no = f"YV-{year}{month:02d}-{seq:04d}"
        seq += 1
        desc = f"{'Satış' if is_income else 'Alış'} faturası — {inv.invoice_no or inv.id}"
        if is_income:
            # 120 Alıcılar / 600 Satışlar / 391 Hesaplanan KDV
            entries.append({
                "date": inv.invoice_date, "doc_no": doc_no, "description": desc,
                "account_code": "120", "account_name": ACCOUNTS["120"],
                "debit": round(total, 2), "credit": 0,
                "ref_type": "invoice", "ref_id": inv.id,
            })
            entries.append({
                "date": inv.invoice_date, "doc_no": doc_no, "description": desc,
                "account_code": "600", "account_name": ACCOUNTS["600"],
                "debit": 0, "credit": round(inv.amount, 2),
                "ref_type": "invoice", "ref_id": inv.id,
            })
            entries.append({
                "date": inv.invoice_date, "doc_no": doc_no, "description": desc,
                "account_code": "391", "account_name": ACCOUNTS["391"],
                "debit": 0, "credit": round(vat_amt, 2),
                "ref_type": "invoice", "ref_id": inv.id,
            })
        else:
            # 740/770 Maliyet / 191 İndirilecek KDV / 320 Satıcılar
            cost_account = "770" if inv.amount > 0 else "740"
            entries.append({
                "date": inv.invoice_date, "doc_no": doc_no, "description": desc,
                "account_code": cost_account, "account_name": ACCOUNTS[cost_account],
                "debit": round(inv.amount, 2), "credit": 0,
                "ref_type": "invoice", "ref_id": inv.id,
            })
            entries.append({
                "date": inv.invoice_date, "doc_no": doc_no, "description": desc,
                "account_code": "191", "account_name": ACCOUNTS["191"],
                "debit": round(vat_amt, 2), "credit": 0,
                "ref_type": "invoice", "ref_id": inv.id,
            })
            entries.append({
                "date": inv.invoice_date, "doc_no": doc_no, "description": desc,
                "account_code": "320", "account_name": ACCOUNTS["320"],
                "debit": 0, "credit": round(total, 2),
                "ref_type": "invoice", "ref_id": inv.id,
            })

    # Genel giderler
    expenses = db.query(GeneralExpense).filter(
        GeneralExpense.company_id == cid,
        extract("year", GeneralExpense.expense_date) == year,
        extract("month", GeneralExpense.expense_date) == month,
    ).all()
    for e in expenses:
        if not e.expense_date:
            continue
        doc_no = f"YV-{year}{month:02d}-{seq:04d}"
        seq += 1
        desc = f"Genel gider — {e.description or 'açıklama yok'}"
        # 770 Genel Yönetim / 100 Kasa veya 102 Banka (tahmini — webapp kaynağı belirsiz)
        entries.append({
            "date": e.expense_date, "doc_no": doc_no, "description": desc,
            "account_code": "770", "account_name": ACCOUNTS["770"],
            "debit": round(e.amount, 2), "credit": 0,
            "ref_type": "expense", "ref_id": e.id,
        })
        entries.append({
            "date": e.expense_date, "doc_no": doc_no, "description": desc,
            "account_code": "100", "account_name": ACCOUNTS["100"],
            "debit": 0, "credit": round(e.amount, 2),
            "ref_type": "expense", "ref_id": e.id,
        })

    # Tarihe göre sırala
    entries.sort(key=lambda x: (x["date"], x["doc_no"]))
    return entries


def _build_ledger(journal: list) -> dict:
    """Yevmiye girişlerinden büyük defter (her hesap için ayrı muavin)."""
    ledger = defaultdict(lambda: {"name": "", "entries": [], "total_debit": 0, "total_credit": 0})
    for e in journal:
        code = e["account_code"]
        ledger[code]["name"] = e["account_name"]
        ledger[code]["entries"].append(e)
        ledger[code]["total_debit"] += e["debit"]
        ledger[code]["total_credit"] += e["credit"]
    # Bakiye hesapla
    for code, acc in ledger.items():
        acc["balance"] = acc["total_debit"] - acc["total_credit"]
    return dict(sorted(ledger.items()))


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@router.get("", response_class=HTMLResponse, name="edefter_index")
async def edefter_index(
    request: Request,
    year: int = None,
    month: int = None,
    view: str = "journal",
    current_user: User = Depends(require_mudur),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    _require_module_active(db)
    today = date.today()
    if not year:
        year = today.year
    if not month:
        month = today.month

    journal = _build_journal(db, year, month, cid)
    ledger = _build_ledger(journal)

    total_debit = sum(e["debit"] for e in journal)
    total_credit = sum(e["credit"] for e in journal)

    return templates.TemplateResponse(
        "edefter/index.html",
        {
            "request": request,
            "current_user": current_user,
            "page_title": "E-Defter (Simülasyon)",
            "year": year, "month": month,
            "view": view,
            "journal": journal,
            "ledger": ledger,
            "total_debit": total_debit,
            "total_credit": total_credit,
            "balance_check": abs(total_debit - total_credit) < 0.01,
            "month_names": ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
                            "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"],
        },
    )


@router.get("/export", name="edefter_export")
async def edefter_export(
    year: int = None,
    month: int = None,
    current_user: User = Depends(require_mudur),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    _require_module_active(db)
    today = date.today()
    if not year:
        year = today.year
    if not month:
        month = today.month

    journal = _build_journal(db, year, month, cid)
    ledger = _build_ledger(journal)

    NAVY = "1A3A5C"
    NAVY_2 = "1E5F8C"
    GRAY = "F1F5F9"
    money_fmt = '#,##0.00'
    thin = Side(border_style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    month_names = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
                   "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]

    wb = openpyxl.Workbook()

    # Sheet 1 — Yevmiye Defteri
    ws = wb.active
    ws.title = "Yevmiye"
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"Yevmiye Defteri — {month_names[month-1]} {year}"
    c.font = Font(size=14, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Tarih", "Yev. No", "Açıklama", "Hesap Kodu", "Hesap Adı", "Borç", "Alacak"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY_2)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for r_idx, e in enumerate(journal, start=4):
        ws.cell(row=r_idx, column=1, value=e["date"]).number_format = "DD.MM.YYYY"
        ws.cell(row=r_idx, column=2, value=e["doc_no"])
        ws.cell(row=r_idx, column=3, value=e["description"])
        ws.cell(row=r_idx, column=4, value=e["account_code"]).alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=5, value=e["account_name"])
        cell = ws.cell(row=r_idx, column=6, value=e["debit"] if e["debit"] else None)
        cell.number_format = money_fmt
        cell.alignment = Alignment(horizontal="right")
        cell = ws.cell(row=r_idx, column=7, value=e["credit"] if e["credit"] else None)
        cell.number_format = money_fmt
        cell.alignment = Alignment(horizontal="right")

    # Toplam satırı
    total_row = 4 + len(journal) + 1
    ws.cell(row=total_row, column=5, value="TOPLAM").font = Font(bold=True)
    ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="right")
    cell = ws.cell(row=total_row, column=6, value=sum(e["debit"] for e in journal))
    cell.font = Font(bold=True, color=NAVY)
    cell.number_format = money_fmt
    cell.fill = PatternFill("solid", fgColor=GRAY)
    cell = ws.cell(row=total_row, column=7, value=sum(e["credit"] for e in journal))
    cell.font = Font(bold=True, color=NAVY)
    cell.number_format = money_fmt
    cell.fill = PatternFill("solid", fgColor=GRAY)

    for letter, w in zip("ABCDEFG", [12, 16, 40, 12, 28, 14, 14]):
        ws.column_dimensions[letter].width = w

    # Sheet 2 — Büyük Defter (her hesap ayrı bölüm)
    ws2 = wb.create_sheet("Büyük Defter")
    ws2.sheet_view.showGridLines = False
    ws2.row_dimensions[1].height = 30
    ws2.merge_cells("A1:F1")
    c = ws2["A1"]
    c.value = f"Büyük Defter — {month_names[month-1]} {year}"
    c.font = Font(size=14, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    cur_row = 3
    for code, acc in ledger.items():
        # Hesap başlığı
        ws2.merge_cells(f"A{cur_row}:F{cur_row}")
        cell = ws2.cell(row=cur_row, column=1,
                        value=f"  {code} — {acc['name']}  (Borç: {acc['total_debit']:,.2f} / Alacak: {acc['total_credit']:,.2f} / Bakiye: {acc['balance']:,.2f})")
        cell.font = Font(size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY_2)
        cur_row += 1

        # Sub-headers
        for i, h in enumerate(["Tarih", "Yev. No", "Açıklama", "Borç", "Alacak", "Bakiye"], 1):
            cell = ws2.cell(row=cur_row, column=i, value=h)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill("solid", fgColor=GRAY)
            cell.alignment = Alignment(horizontal="center")
        cur_row += 1

        running = 0
        for e in acc["entries"]:
            running += e["debit"] - e["credit"]
            ws2.cell(row=cur_row, column=1, value=e["date"]).number_format = "DD.MM.YYYY"
            ws2.cell(row=cur_row, column=2, value=e["doc_no"])
            ws2.cell(row=cur_row, column=3, value=e["description"])
            cell = ws2.cell(row=cur_row, column=4, value=e["debit"] if e["debit"] else None)
            cell.number_format = money_fmt
            cell.alignment = Alignment(horizontal="right")
            cell = ws2.cell(row=cur_row, column=5, value=e["credit"] if e["credit"] else None)
            cell.number_format = money_fmt
            cell.alignment = Alignment(horizontal="right")
            cell = ws2.cell(row=cur_row, column=6, value=running)
            cell.number_format = money_fmt
            cell.alignment = Alignment(horizontal="right")
            cell.font = Font(bold=True, color="16A34A" if running >= 0 else "DC2626")
            cur_row += 1
        cur_row += 1  # boşluk

    for letter, w in zip("ABCDEF", [12, 16, 40, 14, 14, 14]):
        ws2.column_dimensions[letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"edefter-{year}-{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
