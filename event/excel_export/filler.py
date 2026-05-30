"""
Müşteri Excel template'ini cell_map ile doldurur.

cell_map yapısı (customer.excel_config_json içinde):
{
  "vat_mode": "exclusive" | "inclusive",
  "header": {
      "B4": "event_name",
      "C5": "check_in",
      ...
  },
  "data_block": {
      "start_row": 11,
      "end_anchor_text": "ARA TOPLAM",
      "sheet": "Sheet1",
      "columns": {
          "B": "service_name",
          "C": "notes",
          "D": "nights",
          "E": "qty",
          "F": "sale_price_eur",
          "H": "sale_price"
      }
  }
}

Formül kolonları (G, I, J vb.) template'in ilk dolu satırından otomatik tespit edilir.
"""
from __future__ import annotations

import io
import os
import re

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string, get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

from .builder import SECTION_LABELS, SECTIONS_ORDER


# Bölüm ara toplam etiketleri
SECTION_SUBTOTAL_LABELS: dict[str, str] = {
    "accommodation": "Konaklama Ara Toplam",
    "meeting":       "Toplantı / Salon Ara Toplam",
    "fb":            "Yeme & İçme Ara Toplam",
    "teknik":        "Teknik Ekipman Ara Toplam",
    "dekor":         "Dekor / Süsleme Ara Toplam",
    "transfer":      "Transfer / Ulaşım Ara Toplam",
    "tasarim":       "Tasarım & Baskı Ara Toplam",
    "other":         "Diğer Hizmetler Ara Toplam",
}


# ── Stil sabitleri ─────────────────────────────────────────────────────────────
def _HDR_FONT(): return Font(bold=True, color="FFFFFF", size=10)
def _HDR_FILL(): return PatternFill(fill_type="solid", fgColor="1E5F8C")
def _SUB_FONT(): return Font(bold=True, color="1A3A5C", size=10)
def _SUB_FILL(): return PatternFill(fill_type="solid", fgColor="D0E8F5")
def _TOT_FONT(): return Font(bold=True, color="FFFFFF", size=11)
def _TOT_FILL(): return PatternFill(fill_type="solid", fgColor="1A3A5C")
def _DAT_FONT(): return Font(color="000000", size=10)


def _primary_contact(customer) -> str:
    """Müşterinin ilk yetkili kişisinin adını döner."""
    if not customer:
        return ""
    try:
        import json as _json
        contacts = _json.loads(getattr(customer, "contacts_json", None) or "[]")
        if contacts:
            c = contacts[0]
            name  = (c.get("name") or "").strip()
            title = (c.get("title") or "").strip()
            return f"{name} ({title})" if title else name
    except Exception:
        pass
    return ""


def _budget_totals(budget) -> dict:
    """Bütçenin offer_currency cinsinden genel toplamlarını hesaplar."""
    currency   = (getattr(budget, "offer_currency", None) or "TRY").upper()
    offer_rate = budget.rate_to_try(currency) or 1.0

    total_sale_excl = total_sale_incl = 0.0
    total_cost_excl = total_cost_incl = 0.0

    for r in (budget.rows or []):
        qty    = float(r.get("qty",    1) or 1)
        nights = float(r.get("nights", 1) or 1)
        sale   = float(r.get("sale_price", 0) or 0)
        cost   = float(r.get("cost_price", 0) or 0)
        vat    = float(r.get("vat_rate",   0) or 0)
        row_cur  = (r.get("currency") or "TRY").upper()
        row_rate = budget.rate_to_try(row_cur) or 1.0
        conv     = row_rate / offer_rate

        s = sale * conv * qty * nights
        c = cost * conv * qty * nights
        total_sale_excl += s
        total_sale_incl += s * (1 + vat / 100)
        total_cost_excl += c
        total_cost_incl += c * (1 + vat / 100)

    return {
        "venue_name":      budget.venue_name or "",
        "currency":        currency,
        "total_sale_excl": round(total_sale_excl, 2),
        "total_sale_incl": round(total_sale_incl, 2),
        "total_cost_excl": round(total_cost_excl, 2),
        "total_cost_incl": round(total_cost_incl, 2),
    }


def _sf_sale(budget, currency: str) -> float:
    """Hizmet bedeli tutarını offer_currency cinsinden hesaplar."""
    pct = float(budget.service_fee_pct or 0)
    if not pct:
        return 0.0
    offer_rate = budget.rate_to_try(currency) or 1.0
    base = 0.0
    for r in budget.rows:
        if r.get("is_service_fee") or r.get("is_accommodation_tax"):
            continue
        sale      = float(r.get("sale_price", 0) or 0)
        qty       = float(r.get("qty",    1) or 1)
        nights    = float(r.get("nights", 1) or 1)
        row_cur   = (r.get("currency") or "TRY").upper()
        row_rate  = budget.rate_to_try(row_cur) or 1.0
        conv      = row_rate / offer_rate
        base     += sale * qty * nights * conv
    return round(base * pct / 100, 2)


# ── Header alan çözücüleri ─────────────────────────────────────────────────────
def _header_resolvers(budget, request, customer, creator) -> dict:
    req = request

    def _date(d):
        if not d:
            return ""
        if isinstance(d, str):
            try:
                from datetime import date as _date_cls
                d = _date_cls.fromisoformat(d[:10])
            except Exception:
                return d
        return d.strftime("%d.%m.%Y")

    cities = ""
    if req:
        cities = (", ".join(req.cities)
                  if getattr(req, "cities", None)
                  else getattr(req, "city", ""))

    currency = (budget.offer_currency or "TRY").upper()
    sf        = _sf_sale(budget, currency)

    totals = _budget_totals(budget)

    return {
        "event_name":     (getattr(req, "event_name", None)  or budget.venue_name or ""),
        "ref_no":         (getattr(req, "request_no", None)  or ""),
        "check_in":       _date(getattr(req, "check_in", None)),
        "check_out":      _date(getattr(req, "check_out", None)),
        "venue_name":     (budget.venue_name or ""),
        "customer_name":  (getattr(customer, "name", None)
                           or getattr(req, "client_name", None)
                           or ""),
        "creator_name":   (f"{getattr(creator, 'name', '')} "
                           f"{getattr(creator, 'surname', '')}".strip()
                           if creator else ""),
        "eur_rate":       budget.rate_to_try("EUR"),
        "usd_rate":       budget.rate_to_try("USD"),
        "attendee_count": (getattr(req, "attendee_count", None) or ""),
        "city":           cities,
        # Servis bedeli
        "sf_pct":         float(budget.service_fee_pct or 0),
        "sf_sale":        sf,
        "sf_vat":         round(sf * 0.20, 2),
        "sf_total":       round(sf * 1.20, 2),
        # Müşteri yetkili kişi (ilk kontak)
        "contact_name":  _primary_contact(customer),
        # Teklif son tarihi
        "quote_deadline": _date(getattr(req, "quote_deadline", None)),
        # Bütçe toplamları (özet sayfa için)
        "total_sale_excl": totals["total_sale_excl"],
        "total_sale_incl": totals["total_sale_incl"],
        "total_cost_excl": totals["total_cost_excl"],
        "total_cost_incl": totals["total_cost_incl"],
    }


# ── Satır alan çözücüleri ──────────────────────────────────────────────────────
def _row_value(field: str, row: dict, budget, currency: str) -> float | str:
    qty    = float(row.get("qty",    1) or 1)
    nights = float(row.get("nights", 1) or 1)
    sale0  = float(row.get("sale_price", 0) or 0)   # ham (row_cur cinsinden)
    vat    = float(row.get("vat_rate",   0) or 0)
    sf_pct = float(budget.service_fee_pct or 0)

    row_cur    = (row.get("currency") or "TRY").upper()
    row_rate   = budget.rate_to_try(row_cur) or 1.0   # 1 row_cur = ? TRY
    offer_rate = budget.rate_to_try(currency) or 1.0  # 1 offer_cur = ? TRY

    # offer_currency cinsinden birim fiyat
    sale = sale0 * row_rate / offer_rate if row_cur != currency else sale0

    # TRY cinsinden yardımcı değerler
    sale_try = sale0 * row_rate           # birim fiyat TRY
    sf_mul   = 1 + sf_pct / 100          # servis bedeli çarpanı

    def _to_cur(target: str) -> float:
        t_rate = budget.rate_to_try(target) or 1.0
        return (sale0 * row_rate) / t_rate

    match field:
        case "service_name":         return row.get("service_name", "")
        case "notes":                return row.get("notes", "")
        case "unit":                 return row.get("unit", "Adet")
        case "qty":                  return qty
        case "nights":               return nights
        case "vat_rate":             return vat
        case "vat_pct":              return vat / 100
        # Servis bedeli yüzdesi (is_service_fee satırında sf_percent, diğerlerinde 0)
        case "sf_pct":               return float(row.get("sf_percent", 0) or budget.service_fee_pct or 0)
        # Offer currency
        case "sale_price":           return round(sale, 2)
        case "sale_price_inc":       return round(sale * (1 + vat / 100), 2)
        case "total_excl":           return round(sale * qty * nights, 2)
        case "total_incl":           return round(sale * (1 + vat / 100) * qty * nights, 2)
        # EUR / USD
        case "sale_price_eur":       return round(_to_cur("EUR"), 2)
        case "sale_price_usd":       return round(_to_cur("USD"), 2)
        case "total_eur":            return round(_to_cur("EUR") * qty * nights, 2)
        case "total_usd":            return round(_to_cur("USD") * qty * nights, 2)
        # Kur (satırın para biriminin TRY değeri)
        case "kur":                  return round(row_rate, 4)
        # TRY bazlı (servis bedeli hariç)
        case "sale_price_try":       return round(sale_try, 2)
        case "total_excl_try":       return round(sale_try * qty * nights, 2)
        case "total_incl_try":       return round(sale_try * (1 + vat / 100) * qty * nights, 2)
        # TRY bazlı (servis bedeli dahil, KDV hariç)
        case "sale_price_sf":        return round(sale * sf_mul, 2)
        case "sale_price_sf_try":    return round(sale_try * sf_mul, 2)
        case "total_incl_sf_try":    return round(sale_try * sf_mul * qty * nights, 2)
        # TRY bazlı (servis bedeli dahil, KDV dahil)
        case "total_incl_sf_vat_try":return round(sale_try * sf_mul * (1 + vat / 100) * qty * nights, 2)
        case _:                      return ""


# ── Formül şablonu çıkarıcı ────────────────────────────────────────────────────
def _extract_formula_templates(ws, start_row: int, written_cols: set, max_col: int) -> dict:
    """
    Template'deki veri satırlarından formül şablonlarını çıkarır.
    Yazan kolonlar (col_defs) hariç, formül içeren kolonları tespit eder.
    Satır numarasını {row} ile değiştirir.
    """
    formula_cols: dict[str, str] = {}
    for r_idx in range(start_row, min(start_row + 30, ws.max_row + 1)):
        found = False
        for c_idx in range(1, max_col + 1):
            col_letter = get_column_letter(c_idx)
            if col_letter.upper() in written_cols:
                continue
            if col_letter in formula_cols:
                continue
            try:
                cell = ws.cell(row=r_idx, column=c_idx)
            except Exception:
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                # Satır numarasını {row} ile değiştir
                pattern = r'(?<=[A-Za-z\$])(' + str(r_idx) + r')(?=[^0-9]|$)'
                tpl = re.sub(pattern, '{row}', cell.value)
                formula_cols[col_letter] = tpl
                found = True
        if found:
            break
    return formula_cols


# ── Güvenli hücre yazma ────────────────────────────────────────────────────────
def _safe_set(ws, row: int, col_letter: str, value,
              font=None, fill=None) -> None:
    try:
        ci   = column_index_from_string(col_letter.upper())
        cell = ws.cell(row=row, column=ci)
        cell.value = value
        if font is not None:
            cell.font = font
        if fill is not None:
            cell.fill = fill
    except (AttributeError, Exception):
        pass


# ── Ana fill fonksiyonu (in-place) ────────────────────────────────────────────
def _fill_ws(ws, cell_map: dict, budget, request, customer, creator) -> None:
    """Worksheet'i cell_map ile doldurur: header, bölüm başlıkları, ara toplamlar, genel toplam."""
    currency = (budget.offer_currency or "TRY").upper()

    # 1. Header hücrelerini doldur
    header_vals = _header_resolvers(budget, request, customer, creator)
    for cell_addr, field_name in (cell_map.get("header") or {}).items():
        val = header_vals.get(field_name)
        if val is not None:
            try:
                ws[cell_addr.upper()] = val
                ws[cell_addr.upper()].font = Font(color="000000")
            except (AttributeError, Exception):
                pass

    data_block = cell_map.get("data_block")
    if not data_block:
        return

    start_row  = int(data_block.get("start_row", 1))
    col_defs   = data_block.get("columns", {})           # {letter: field_name}
    end_anchor = data_block.get("end_anchor_text")
    label_col  = "B"                                      # bölüm etiketleri için varsayılan kolon

    # 2. Anchor satırını bul
    anchor_row = None
    if end_anchor:
        for r_idx in range(start_row, ws.max_row + 1):
            for c_idx in range(1, ws.max_column + 1):
                v = ws.cell(row=r_idx, column=c_idx).value
                if v and isinstance(v, str) and end_anchor.lower() in v.lower():
                    anchor_row = r_idx
                    break
            if anchor_row:
                break

    # 3. Formül şablonlarını temizlemeden önce çıkar
    #    Önce cell_map'ten kullanıcı tanımlı formüller al, sonra template'den otomatik bul
    user_formula_cols: dict[str, str] = {
        c.upper(): tpl
        for c, tpl in (data_block.get("formula_columns") or {}).items()
    }
    written_cols = {c.upper() for c in col_defs} | set(user_formula_cols.keys())
    auto_formula_cols = _extract_formula_templates(
        ws, start_row, written_cols, ws.max_column
    )
    # Kullanıcı tanımlı formüller template'den çıkarılanlara göre önceliklidir
    formula_cols = {**auto_formula_cols, **user_formula_cols}
    print(f"[FILLER] start_row={start_row} anchor={anchor_row} written={sorted(written_cols)} formula_cols={dict(formula_cols)}", flush=True)

    # 4. Veri alanını temizle (start_row → anchor+10 arası)
    clear_end = (anchor_row + 10) if anchor_row else (start_row + 100)
    for r_idx in range(start_row, min(clear_end + 1, ws.max_row + 1)):
        for c_idx in range(1, ws.max_column + 1):
            try:
                ws.cell(row=r_idx, column=c_idx).value = None
            except AttributeError:
                pass

    # 5. Bütçe satırlarını bölümlere ayır
    rows_by_sec: dict[str, list] = {}
    service_fee = None
    for r in budget.rows:
        if r.get("is_service_fee"):
            service_fee = r
            continue
        sec = r.get("section", "other")
        rows_by_sec.setdefault(sec, []).append(r)

    # col_defs'teki sayısal toplam alanları — servis bedeli için SUM formülü yazılacak
    _NUMERIC_TOTAL_FIELDS = {
        "sale_price", "sale_price_inc", "total_excl", "total_incl",
        "sale_price_try", "total_excl_try", "total_incl_try",
        "sale_price_sf", "sale_price_sf_try", "total_incl_sf_try", "total_incl_sf_vat_try",
        "sale_price_eur", "sale_price_usd", "total_eur", "total_usd",
    }

    # 6. Bölümleri yaz
    current_row = start_row
    subtotal_rows: dict[str, list[int]] = {}   # col_letter -> [subtotal satır numaraları]
    # col_defs sayısal kolonları için veri aralıklarını tut (servis bedeli hesabı için)
    col_defs_ranges: dict[str, list[str]] = {}  # col_letter -> ["B5:B8", ...]

    for sec in SECTIONS_ORDER:
        sec_rows = rows_by_sec.get(sec, [])
        if not sec_rows:
            continue

        sec_label = SECTION_LABELS.get(sec, sec)

        # Bölüm başlık satırı
        _safe_set(ws, current_row, label_col, sec_label,
                  font=_HDR_FONT(), fill=_HDR_FILL())
        current_row += 1

        sec_data_start = current_row

        # Veri satırları
        for row_data in sec_rows:
            for col_letter, field_name in col_defs.items():
                val = _row_value(field_name, row_data, budget, currency)
                _safe_set(ws, current_row, col_letter, val, font=_DAT_FONT())
            # Formül kolonları
            for col_letter, tpl in formula_cols.items():
                formula = tpl.replace("{row}", str(current_row))
                _safe_set(ws, current_row, col_letter, formula, font=_DAT_FONT())
            current_row += 1

        sec_data_end = current_row - 1
        print(f"[FILLER] sec={sec} hdr={sec_data_start-1} data={sec_data_start}..{sec_data_end} subtotal={current_row}", flush=True)

        # col_defs sayısal kolonları için bu bölümün aralığını kaydet
        for col_letter, field_name in col_defs.items():
            if field_name in _NUMERIC_TOTAL_FIELDS:
                col_defs_ranges.setdefault(col_letter, []).append(
                    f"{col_letter}{sec_data_start}:{col_letter}{sec_data_end}"
                )

        # Ara toplam satırı
        sub_label = SECTION_SUBTOTAL_LABELS.get(sec, f"{sec_label} Ara Toplam")
        _safe_set(ws, current_row, label_col, sub_label,
                  font=_SUB_FONT(), fill=_SUB_FILL())
        for col_letter in formula_cols:
            formula = f"=SUM({col_letter}{sec_data_start}:{col_letter}{sec_data_end})"
            _safe_set(ws, current_row, col_letter, formula,
                      font=_SUB_FONT(), fill=_SUB_FILL())
            subtotal_rows.setdefault(col_letter, []).append(current_row)
        current_row += 1

    # Hizmet bedeli satırı
    if service_fee:
        sf_label = service_fee.get("service_name") or "Hizmet Bedeli"
        sf_pct   = float(service_fee.get("sf_percent", 0) or budget.service_fee_pct or 0)
        _safe_set(ws, current_row, label_col, sf_label, font=_DAT_FONT())

        # col_defs kolonları:
        # - sayısal toplam alanları → =SUM(tüm veri aralıkları)*pct/100
        # - diğerleri (service_name, unit, qty, sf_pct vb.) → _row_value ile yaz
        for col_letter, field_name in col_defs.items():
            if field_name in _NUMERIC_TOTAL_FIELDS and sf_pct:
                ranges = col_defs_ranges.get(col_letter, [])
                if ranges:
                    range_args = ",".join(ranges)
                    formula = f"=SUM({range_args})*{sf_pct}/100"
                    _safe_set(ws, current_row, col_letter, formula, font=_DAT_FONT())
                # ranges yoksa hiçbir şey yazma
            else:
                val = _row_value(field_name, service_fee, budget, currency)
                _safe_set(ws, current_row, col_letter, val, font=_DAT_FONT())

        # formula_cols: normal satır formülü UYGULANMAZ.
        # Bunun yerine ara toplam hücrelerinin yüzdesi alınır: =(sub1+sub2+...)*pct/100
        for col_letter in formula_cols:
            subs = subtotal_rows.get(col_letter, [])
            if subs and sf_pct:
                refs    = "+".join(f"{col_letter}{r}" for r in subs)
                formula = f"=({refs})*{sf_pct}/100"
            else:
                formula = ""
            _safe_set(ws, current_row, col_letter, formula, font=_DAT_FONT())
            subtotal_rows.setdefault(col_letter, []).append(current_row)
        current_row += 1

    # Genel toplam satırı
    if subtotal_rows:
        _safe_set(ws, current_row, label_col, "GENEL TOPLAM (KDV Hariç)",
                  font=_TOT_FONT(), fill=_TOT_FILL())
        for col_letter, rows in subtotal_rows.items():
            refs = "+".join(f"{col_letter}{r}" for r in rows)
            _safe_set(ws, current_row, col_letter, f"={refs}",
                      font=_TOT_FONT(), fill=_TOT_FILL())
        current_row += 1

    # 7. Kullanılmayan eski satırları sil
    if clear_end >= current_row:
        try:
            ws.delete_rows(current_row, clear_end - current_row + 1)
        except Exception:
            pass


# ── Özet sayfa oluşturucular ──────────────────────────────────────────────────
def _add_summary_sheet_auto(wb, entries: list, sheet_name: str = "Teklif Özeti") -> None:
    """Programatik özet sayfası oluşturur ve workbook'a başa ekler."""
    if not entries:
        return

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name, index=0)

    first = entries[0]
    req = first.get("request")
    cus = first.get("customer")
    bud = first.get("budget")

    def _date(d):
        if not d:
            return ""
        if isinstance(d, str):
            try:
                from datetime import date as _dc
                d = _dc.fromisoformat(d[:10])
            except Exception:
                return d
        return d.strftime("%d.%m.%Y")

    cities_str = ""
    if req:
        cities_str = (", ".join(req.cities)
                      if getattr(req, "cities", None)
                      else getattr(req, "city", "") or "")

    # Satır 1 — Başlık
    ws.merge_cells("A1:G1")
    c = ws.cell(1, 1, sheet_name.upper())
    c.font      = Font(bold=True, size=14, color="FFFFFF")
    c.fill      = PatternFill(fill_type="solid", fgColor="1A3A5C")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Satır 2+ — Etkinlik bilgileri
    info_items = [
        ("Etkinlik Adı",       getattr(req, "event_name", None) or (bud.venue_name if bud else "")),
        ("Referans No",        getattr(req, "request_no",  None) or ""),
        ("Şehir",              cities_str),
        ("Etkinlik Tarihleri", f"{_date(getattr(req,'check_in',None))} – "
                               f"{_date(getattr(req,'check_out',None))}" if req else ""),
        ("Katılımcı Sayısı",   str(getattr(req, "attendee_count", "") or "")),
        ("Müşteri / Firma",    getattr(cus, "name", None)
                               or getattr(req, "client_name", None) or ""),
        ("Teklif Son Tarihi",  _date(getattr(req, "quote_deadline", None))),
    ]

    for i, (lbl, val) in enumerate(info_items):
        r = i + 2
        lc = ws.cell(r, 1, lbl)
        lc.font      = Font(bold=True, size=10, color="1A3A5C")
        lc.fill      = PatternFill(fill_type="solid", fgColor="EFF6FF")
        lc.alignment = Alignment(vertical="center", indent=1)
        vc = ws.cell(r, 2, val)
        vc.font      = Font(size=10, color="1E293B")
        vc.fill      = PatternFill(fill_type="solid", fgColor="FFFFFF")
        ws.merge_cells(f"B{r}:G{r}")
        ws.row_dimensions[r].height = 16

    # Karşılaştırma tablosu
    tbl_row = len(info_items) + 3   # 1 boş satır boşluk
    hdr_labels = [
        "Mekan / Tedarikçi", "Para Birimi",
        "KDV Hariç Toplam", "KDV Dahil Toplam",
        "Maliyet (KDV Hariç)", "Maliyet (KDV Dahil)", "Kar Marjı",
    ]
    for j, h in enumerate(hdr_labels):
        cell = ws.cell(tbl_row, j + 1, h)
        cell.font      = _HDR_FONT()
        cell.fill      = _HDR_FILL()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[tbl_row].height = 32

    for i, entry in enumerate(entries):
        totals = _budget_totals(entry["budget"])
        dr = tbl_row + 1 + i

        margin = ""
        if totals["total_sale_excl"] and totals["total_cost_excl"]:
            m = (totals["total_sale_excl"] - totals["total_cost_excl"]) / totals["total_sale_excl"] * 100
            margin = f"%{round(m, 1)}"

        row_vals = [
            totals["venue_name"], totals["currency"],
            totals["total_sale_excl"], totals["total_sale_incl"],
            totals["total_cost_excl"], totals["total_cost_incl"],
            margin,
        ]
        bg   = "FFFFFF" if i % 2 == 0 else "F8FAFC"
        fill = PatternFill(fill_type="solid", fgColor=bg)
        for j, val in enumerate(row_vals):
            cell = ws.cell(dr, j + 1, val)
            cell.font      = _DAT_FONT()
            cell.fill      = fill
            cell.alignment = Alignment(vertical="center")
            if 2 <= j <= 5:
                cell.number_format = "#,##0.00"
        ws.row_dimensions[dr].height = 18

    # Çok bütçe: genel toplam satırı
    if len(entries) > 1:
        gt_row    = tbl_row + len(entries) + 1
        first_dr  = tbl_row + 1
        last_dr   = tbl_row + len(entries)
        ws.cell(gt_row, 1, "GENEL TOPLAM").font      = _TOT_FONT()
        ws.cell(gt_row, 1).fill      = _TOT_FILL()
        ws.cell(gt_row, 1).alignment = Alignment(vertical="center", indent=1)
        for col_idx in range(3, 7):
            cl   = get_column_letter(col_idx)
            cell = ws.cell(gt_row, col_idx, f"=SUM({cl}{first_dr}:{cl}{last_dr})")
            cell.font          = _TOT_FONT()
            cell.fill          = _TOT_FILL()
            cell.number_format = "#,##0.00"
        ws.row_dimensions[gt_row].height = 20

    # Sütun genişlikleri
    for j, w in enumerate([32, 14, 20, 20, 20, 20, 14]):
        ws.column_dimensions[get_column_letter(j + 1)].width = w


def _fill_summary_template(wb, entries: list, summary_cfg: dict) -> None:
    """
    Template modunda özet sayfasını doldurur.
    Kaynak sayfa (source_sheet) yoksa otomatiğe düşer.
    """
    source_sheet = summary_cfg.get("source_sheet")
    sheet_name   = summary_cfg.get("sheet_name") or "Teklif Özeti"
    header_map   = summary_cfg.get("header") or {}
    data_block   = summary_cfg.get("data_block") or {}

    if not source_sheet or source_sheet not in wb.sheetnames:
        _add_summary_sheet_auto(wb, entries, sheet_name)
        return

    ws = wb[source_sheet]
    if source_sheet != sheet_name:
        ws.title = sheet_name

    if not entries:
        return

    first    = entries[0]
    hdr_vals = _header_resolvers(
        first["budget"], first.get("request"),
        first.get("customer"), first.get("creator"),
    )

    # Header hücreleri
    for cell_addr, field_name in header_map.items():
        val = hdr_vals.get(field_name)
        if val is not None:
            try:
                ws[cell_addr.upper()] = val
                ws[cell_addr.upper()].font = Font(color="000000")
            except Exception:
                pass

    # Veri satırları (her bütçe → bir satır)
    if data_block:
        start_row    = int(data_block.get("start_row") or 1)
        col_defs     = data_block.get("columns") or {}
        formula_cols = data_block.get("formula_columns") or {}

        for i, entry in enumerate(entries):
            totals = _budget_totals(entry["budget"])
            if totals["total_sale_excl"] and totals["total_cost_excl"]:
                m = (totals["total_sale_excl"] - totals["total_cost_excl"]) \
                    / totals["total_sale_excl"] * 100
                totals["margin_pct"] = round(m, 1)
            else:
                totals["margin_pct"] = 0.0
            totals["budget_index"] = i + 1

            row_num = start_row + i
            for col_letter, field_name in col_defs.items():
                val = totals.get(field_name, "")
                _safe_set(ws, row_num, col_letter, val, font=_DAT_FONT())
            for col_letter, tpl in formula_cols.items():
                formula = tpl.replace("{row}", str(row_num))
                _safe_set(ws, row_num, col_letter, formula, font=_DAT_FONT())


# ── Tek bütçe export ──────────────────────────────────────────────────────────
def fill_customer_template(
    template_path: str,
    cell_map: dict,
    budget,
    request,
    customer,
    creator,
) -> io.BytesIO:
    """Müşteri template'ini tek bütçe ile doldurur."""
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl kurulu değil")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template bulunamadı: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    sheet_name = (cell_map.get("data_block") or {}).get("sheet")
    ws = (wb[sheet_name]
          if sheet_name and sheet_name in wb.sheetnames
          else wb.active)

    _fill_ws(ws, cell_map, budget, request, customer, creator)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── Çok bütçe export (her bütçe ayrı sheet) ──────────────────────────────────
def fill_customer_template_multi(
    template_path: str,
    cell_map: dict,
    entries: list,
) -> io.BytesIO:
    """
    Birden fazla bütçeyi tek dosyada, her biri template kopyası olarak doldurur.
    entries: [{"budget": ..., "request": ..., "customer": ..., "creator": ...}]
    """
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl kurulu değil")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template bulunamadı: {template_path}")

    sheet_name = (cell_map.get("data_block") or {}).get("sheet")
    used_titles: list[str] = []

    wb = openpyxl.load_workbook(template_path)
    src_ws = (wb[sheet_name]
              if sheet_name and sheet_name in wb.sheetnames
              else wb.active)

    for i, entry in enumerate(entries):
        b   = entry["budget"]
        req = entry.get("request")
        cus = entry.get("customer")
        cre = entry.get("creator")

        ws = src_ws if i == 0 else wb.copy_worksheet(src_ws)

        # Sheet adı = mekan/otel adı
        raw_title = (b.venue_name or f"Bütçe {i+1}")[:28].strip()
        title = raw_title
        suffix = 2
        while title in used_titles:
            title = f"{raw_title[:25]} {suffix}"
            suffix += 1
        used_titles.append(title)
        ws.title = title

        _fill_ws(ws, cell_map, b, req, cus, cre)

    # ── Özet sayfa (Teklif Özeti) ────────────────────────────────────────
    summary_cfg = cell_map.get("summary") or {}
    if summary_cfg.get("enabled"):
        sum_sheet = summary_cfg.get("sheet_name") or "Teklif Özeti"
        if (summary_cfg.get("mode") or "auto") == "template":
            _fill_summary_template(wb, entries, summary_cfg)
        else:
            _add_summary_sheet_auto(wb, entries, sum_sheet)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
