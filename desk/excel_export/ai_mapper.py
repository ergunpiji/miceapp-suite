"""
Claude API ile Excel template analizi ve cell_map öğrenmesi.

Kullanım:
    result = await analyze_template("/path/to/template.xlsx", api_key="sk-ant-...")
    if result["error"]:
        print(result["error"])
    else:
        cell_map = result["cell_map"]
        # → customer.excel_config_json'a kaydet
"""
from __future__ import annotations

import json
import os

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ── Sistem promptu ─────────────────────────────────────────────────────────────
_SYSTEM_PROMPT = """Sen bir Excel template analiz uzmanısın.
Sana bir Excel dosyasının içeriği JSON formatında verilecek (satır × sütun matrisi).
Bu template bir etkinlik organizasyon şirketinin müşterisine ait RFQ / fiyat teklifi formatıdır.

Görevin: Template'deki hücreleri E-dem bütçe sisteminin alanlarıyla eşleştirerek cell_map JSON döndür.

──────────────────────────────────────
E-DEM ALANLARI
──────────────────────────────────────
HEADER alanları (tek bir hücreye yazılır):
  event_name      → etkinlik adı
  ref_no          → referans numarası (TOP-ABC-2504-001 gibi)
  check_in        → etkinlik başlangıç tarihi
  check_out       → etkinlik bitiş tarihi
  venue_name      → mekan adı
  customer_name   → müşteri / firma adı
  creator_name    → teklifi hazırlayan kişi adı
  eur_rate        → 1 EUR = ? TL kur değeri (sayı)
  usd_rate        → 1 USD = ? TL kur değeri (sayı)
  attendee_count  → katılımcı sayısı
  city            → şehir(ler)

SATIR alanları (her bütçe kalemi için tekrarlanan):
  service_name    → hizmet / kalem adı
  notes           → not / açıklama
  unit            → birim (Gece, Kişi, Adet, Gün...)
  qty             → miktar / kişi sayısı
  nights          → gece / gün sayısı
  sale_price      → birim satış fiyatı (KDV hariç, teklif para biriminde)
  sale_price_inc  → birim satış fiyatı (KDV dahil)
  vat_rate        → KDV yüzdesi (20, 10, 0 gibi)
  vat_pct         → KDV oranı (0.20, 0.10 gibi)
  total_excl      → toplam (KDV hariç) = sale_price × qty × nights
  total_incl      → toplam (KDV dahil)
  sale_price_eur  → birim satış fiyatı Euro cinsinden
  total_eur       → toplam Euro cinsinden
  sale_price_usd  → birim satış fiyatı USD cinsinden
  total_usd       → toplam USD cinsinden

──────────────────────────────────────
ÇIKTI FORMATI (sadece JSON, başka metin YAZMA)
──────────────────────────────────────
{
  "vat_mode": "exclusive",
  "header": {
    "B3": "event_name",
    "C4": "ref_no",
    "F4": "eur_rate"
  },
  "data_block": {
    "start_row": 9,
    "end_anchor_text": "ARA TOPLAM",
    "sheet": null,
    "columns": {
      "B": "service_name",
      "C": "notes",
      "D": "nights",
      "E": "qty",
      "F": "sale_price_eur",
      "H": "sale_price",
      "J": "total_incl"
    },
    "section_header_col": "B"
  }
}

KURALLAR:
1. data_block.start_row: İlk gerçek VERİ satırı numarası (başlık satırı değil)
2. end_anchor_text: Veri bloğunun bittiğini gösteren hücre içeriği (varsa, yoksa null)
3. section_header_col: Kategoriler ayrı satır olarak yazılıyorsa hangi sütun (yoksa null)
4. vat_mode: Template'de KDV hariç fiyat görünüyorsa "exclusive", KDV dahil ise "inclusive"
5. Emin olmadığın header alanlarını ekleme — boş {} daha iyi
6. SADECE geçerli bir JSON döndür — başka hiçbir metin yazma"""


def parse_template_structure(template_path: str, max_rows: int = 30) -> list[list]:
    """
    Excel template'ini JSON-serializable matrise dönüştürür.
    AI'ya göndermek için kullanılır.
    """
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl kurulu değil")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    result = []
    for row in ws.iter_rows(min_row=1,
                            max_row=min(max_rows, ws.max_row),
                            values_only=True):
        serialized = []
        for cell in row:
            if cell is None:
                serialized.append(None)
            elif hasattr(cell, "isoformat"):
                serialized.append(str(cell))
            else:
                serialized.append(cell)
        if any(c is not None for c in serialized):
            result.append(serialized)

    return result


def _extract_json(text: str) -> str:
    """Yanıt metninden JSON bloğunu çıkarır."""
    text = text.strip()
    # ```json ... ``` bloğu varsa içini al
    if "```" in text:
        parts = text.split("```")
        for part in parts:
            part = part.strip()
            if part.startswith("json"):
                part = part[4:].strip()
            if part.startswith("{"):
                return part
    # Doğrudan JSON ise
    if text.startswith("{"):
        return text
    # İçinde JSON var mı?
    start = text.find("{")
    end   = text.rfind("}")
    if start != -1 and end != -1:
        return text[start:end + 1]
    return text


async def analyze_template(
    template_path: str,
    api_key: str | None = None,
    model: str | None = None,
) -> dict:
    """
    AI ile template analizi.
    Önce GEMINI_API_KEY, yoksa ANTHROPIC_API_KEY kullanır.

    Returns:
        {"cell_map": dict, "raw_response": str, "error": str | None}
    """
    gemini_key    = os.environ.get("GEMINI_API_KEY", "")
    anthropic_key = api_key or os.environ.get("ANTHROPIC_API_KEY", "")

    if gemini_key:
        return await _analyze_with_gemini(template_path, gemini_key, model or "auto")
    elif anthropic_key:
        return await _analyze_with_claude(template_path, anthropic_key, model or "claude-haiku-4-5-20251001")
    else:
        return {
            "cell_map": {},
            "raw_response": "",
            "error": "API anahtarı bulunamadı. GEMINI_API_KEY veya ANTHROPIC_API_KEY set edin.",
        }


def _gemini_list_models(api_key: str) -> list[str]:
    """generateContent destekleyen modelleri döner (flash önce)."""
    import urllib.request
    url = f"https://generativelanguage.googleapis.com/v1beta/models?key={api_key}&pageSize=50"
    try:
        with urllib.request.urlopen(url, timeout=10) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        all_models = [
            m["name"].replace("models/", "")
            for m in data.get("models", [])
            if "generateContent" in m.get("supportedGenerationMethods", [])
        ]
        # flash modelleri öne al
        flash = [m for m in all_models if "flash" in m]
        rest  = [m for m in all_models if "flash" not in m]
        return flash + rest
    except Exception:
        return []


def _gemini_call(api_key: str, model: str, payload: bytes) -> tuple[str, str | None]:
    """
    Tek model denemesi. (raw_text, error) döner.
    404 hatası → (None, "404") ile ayrıştırılabilir.
    """
    import urllib.request, urllib.error
    url = (f"https://generativelanguage.googleapis.com/v1beta/models/"
           f"{model}:generateContent?key={api_key}")
    req = urllib.request.Request(
        url, data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=90) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        text = data["candidates"][0]["content"]["parts"][0]["text"]
        return text, None
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        return "", f"HTTP{exc.code}:{body[:300]}"
    except Exception as exc:
        return "", str(exc)


async def _analyze_with_gemini(template_path: str, api_key: str, model: str) -> dict:
    """
    Gemini REST API ile template analizi.
    Çalışan modeli otomatik bulur — ListModels'dan gelen sırayla dener,
    404/not-available olanları atlar.
    """
    try:
        structure = parse_template_structure(template_path, max_rows=20)
    except Exception as exc:
        return {"cell_map": {}, "raw_response": "", "error": str(exc)}

    prompt = (
        f"{_SYSTEM_PROMPT}\n\n"
        "Excel template yapısı:\n\n"
        f"```json\n{json.dumps(structure, ensure_ascii=False)}\n```\n\n"
        "Bu template için E-dem cell_map JSON'ını döndür."
    )
    payload = json.dumps({
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.1, "maxOutputTokens": 8192},
    }).encode("utf-8")

    # Model listesini al, her birini dene
    candidates = _gemini_list_models(api_key)
    if not candidates:
        candidates = [
            "gemini-2.5-flash-preview-04-17",
            "gemini-2.5-pro-preview-03-25",
            "gemini-1.5-flash-latest",
            "gemini-1.5-flash-001",
            "gemini-1.5-pro-latest",
        ]

    last_error = "Uygun model bulunamadı"
    for m in candidates:
        raw, err = _gemini_call(api_key, m, payload)
        if err is None:
            try:
                cell_map = json.loads(_extract_json(raw))
                return {"cell_map": cell_map, "raw_response": raw, "error": None}
            except json.JSONDecodeError as jex:
                return {"cell_map": {}, "raw_response": raw,
                        "error": f"JSON parse hatası ({m}): {jex}"}
        # 404 / not-available / timeout → sonrakini dene
        skip_keywords = ("404", "not found", "no longer available", "timed out", "time out")
        if any(kw in err.lower() for kw in skip_keywords):
            last_error = err
            continue
        # Başka hata (auth, quota vb.) → dur
        return {"cell_map": {}, "raw_response": "", "error": err}

    return {"cell_map": {}, "raw_response": "", "error": last_error}


async def _analyze_with_claude(template_path: str, api_key: str, model: str) -> dict:
    """Claude (Anthropic) API ile template analizi."""
    try:
        import anthropic
    except ImportError:
        return {
            "cell_map": {},
            "raw_response": "",
            "error": "anthropic paketi kurulu değil. pip install anthropic",
        }

    try:
        structure = parse_template_structure(template_path, max_rows=20)
    except Exception as exc:
        return {"cell_map": {}, "raw_response": "", "error": str(exc)}

    user_msg = (
        "Excel template yapısı (satır listesi, her satır hücre değerlerini içerir):\n\n"
        f"```json\n{json.dumps(structure, ensure_ascii=False)}\n```\n\n"
        "Bu template için E-dem cell_map JSON'ını döndür."
    )

    raw = ""
    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model=model,
            max_tokens=4096,
            system=_SYSTEM_PROMPT,
            messages=[{"role": "user", "content": user_msg}],
        )
        raw = response.content[0].text
        cell_map = json.loads(_extract_json(raw))
        return {"cell_map": cell_map, "raw_response": raw, "error": None}
    except json.JSONDecodeError as exc:
        return {
            "cell_map": {},
            "raw_response": raw,
            "error": f"Claude yanıtı JSON parse hatası: {exc}",
        }
    except Exception as exc:
        return {"cell_map": {}, "raw_response": raw, "error": str(exc)}
