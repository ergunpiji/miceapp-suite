"""Fon havuzu Excel raporu — özet + T cetveli + alt referans özeti."""
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# Renk paleti
_FILL_HEADER  = PatternFill("solid", fgColor="1E3A5F")    # koyu lacivert
_FILL_SUB     = PatternFill("solid", fgColor="F1F5F9")    # açık gri
_FILL_GREEN   = PatternFill("solid", fgColor="DCFCE7")    # açıklama
_FILL_BLUE    = PatternFill("solid", fgColor="DBEAFE")
_FILL_YELLOW  = PatternFill("solid", fgColor="FEF3C7")
_FILL_BALANCE = PatternFill("solid", fgColor="F0FDF4")

_FONT_HEADER  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_FONT_SUB     = Font(name="Calibri", size=10, bold=True, color="1E293B")
_FONT_BODY    = Font(name="Calibri", size=10, color="334155")
_FONT_BOLD    = Font(name="Calibri", size=10, bold=True, color="1E293B")
_FONT_TITLE   = Font(name="Calibri", size=14, bold=True, color="1E3A5F")

_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
_ALIGN_LEFT  = Alignment(horizontal="left",  vertical="center", wrap_text=True)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
_ALIGN_CENT  = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _fmt_amount(currency: str) -> str:
    """Para birimine göre Excel hücre formatı."""
    sym = {"TRY": "₺", "USD": "$", "EUR": "€"}.get(currency, currency)
    # 1.234,56 ₺ formatı
    return f'#,##0.00 "{sym}"'


def build_fund_pool_excel(fund_req, db) -> io.BytesIO:
    """Fon havuzu için 3 sayfalı Excel raporu üretir.

    Sayfalar:
      1) Özet — müşteri, fon, bakiye dağılımı
      2) Hesap Hareketleri — T cetveli (tarih, açıklama, borç, alacak, bakiye)
      3) Alt Referans Özeti — referans bazlı toplamlar
    """
    from models import FundTransfer, Invoice, Request as ReqModel, Customer
    from utils.funds import get_fund_balance

    wb = openpyxl.Workbook()

    cur = fund_req.fund_currency or "TRY"
    money_fmt = _fmt_amount(cur)
    customer = (db.query(Customer).filter(Customer.id == fund_req.customer_id).first()
                if fund_req.customer_id else None)
    balance = get_fund_balance(fund_req, db)

    initial_invoice = (db.query(Invoice)
                         .filter(Invoice.request_id == fund_req.id,
                                 Invoice.invoice_type == "kesilen")
                         .order_by(Invoice.created_at.asc())
                         .first())

    transfers = (db.query(FundTransfer)
                   .filter(FundTransfer.fund_request_id == fund_req.id)
                   .order_by(FundTransfer.transfer_date.asc(),
                             FundTransfer.created_at.asc())
                   .all())

    # Alt referans haritası — adlar için
    alt_ids = {t.related_request_id for t in transfers}
    alt_map = {r.id: r for r in db.query(ReqModel).filter(ReqModel.id.in_(alt_ids)).all()} if alt_ids else {}

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 1: Özet
    # ─────────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Özet"

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40
    ws.row_dimensions[1].height = 28

    ws["A1"] = "FON HAVUZU RAPORU"
    ws["A1"].font = _FONT_TITLE
    ws.merge_cells("A1:B1")
    ws["A1"].alignment = _ALIGN_CENT

    rows = [
        ("Müşteri",            customer.name if customer else fund_req.client_name or "—"),
        ("Müşteri Kodu",       (customer.code or "").upper() if customer else "—"),
        ("Fon Adı",            fund_req.event_name),
        ("Referans No",        fund_req.request_no),
        ("Yıl",                fund_req.check_in[:4] if fund_req.check_in else "—"),
        ("Para Birimi",        cur),
        ("KDV Oranı",          f"%{int(balance['vat_rate'])}"),
        ("",                   ""),  # boşluk
    ]
    r = 3
    for label, val in rows:
        ws.cell(row=r, column=1, value=label).font = _FONT_SUB
        ws.cell(row=r, column=1).fill = _FILL_SUB
        ws.cell(row=r, column=2, value=val).font = _FONT_BODY
        ws.cell(row=r, column=1).border = _BORDER
        ws.cell(row=r, column=2).border = _BORDER
        r += 1

    # Bakiye dağılımı
    r += 1
    ws.cell(row=r, column=1, value="BAKİYE DAĞILIMI").font = _FONT_HEADER
    ws.cell(row=r, column=1).fill = _FILL_HEADER
    ws.cell(row=r, column=1).alignment = _ALIGN_CENT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    _FILL_NONE = PatternFill(fill_type=None)
    bal_rows = [
        ("Başlangıç Tutarı (KDV dahil)",  balance["initial"]),
        ("KDV Hariç Karşılık",            balance["initial_excl"]),
        ("Toplam Dağıtım (Out)",          balance["out_total"]),
        ("Toplam İade (In)",              balance["in_total"]),
        ("Kalan Bakiye",                  balance["remaining"]),
    ]
    for label, val in bal_rows:
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = _FONT_SUB if label == "Kalan Bakiye" else _FONT_BODY
        c1.fill = _FILL_BALANCE if label == "Kalan Bakiye" else _FILL_SUB
        c1.border = _BORDER
        c2 = ws.cell(row=r, column=2, value=val)
        c2.font = _FONT_BOLD if label == "Kalan Bakiye" else _FONT_BODY
        c2.fill = _FILL_BALANCE if label == "Kalan Bakiye" else _FILL_NONE
        c2.alignment = _ALIGN_RIGHT
        c2.number_format = money_fmt
        c2.border = _BORDER
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Rapor Tarihi:").font = _FONT_BODY
    ws.cell(row=r, column=2, value=datetime.now().strftime("%d.%m.%Y %H:%M")).font = _FONT_BODY
    ws.cell(row=r, column=2).alignment = _ALIGN_RIGHT

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 2: Hesap Hareketleri (T cetveli)
    # ─────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Hesap Hareketleri")
    headers = ["Tarih", "Açıklama", "Alt Referans / Belge",
               f"Borç (Çıkış · {cur})", f"Alacak (Giriş · {cur})",
               f"Bakiye ({cur})"]
    widths = [12, 40, 35, 18, 18, 18]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Başlık satırı
    for col, hd in enumerate(headers, 1):
        c = ws2.cell(row=1, column=col, value=hd)
        c.font = _FONT_HEADER
        c.fill = _FILL_HEADER
        c.alignment = _ALIGN_CENT
        c.border = _BORDER
    ws2.row_dimensions[1].height = 22

    # Satırlar — kronolojik (eski → yeni)
    running_balance = 0.0
    row_idx = 2

    # 1. Açılış faturası
    if initial_invoice:
        opening_amount = float(initial_invoice.total_amount or 0)
        running_balance += opening_amount
        desc = initial_invoice.description or "Havuz açılış faturası"
        ref_text = f"Fatura No: {initial_invoice.invoice_no}" if initial_invoice.invoice_no else "—"
        if initial_invoice.document_name:
            ref_text += f"\n📎 {initial_invoice.document_name}"
        cells = [
            (initial_invoice.invoice_date or fund_req.check_in or "", _ALIGN_LEFT, None),
            (desc,                                                    _ALIGN_LEFT, None),
            (ref_text,                                                _ALIGN_LEFT, None),
            (None,                                                    _ALIGN_RIGHT, money_fmt),
            (opening_amount,                                          _ALIGN_RIGHT, money_fmt),
            (running_balance,                                         _ALIGN_RIGHT, money_fmt),
        ]
        for col, (val, align, fmt) in enumerate(cells, 1):
            c = ws2.cell(row=row_idx, column=col, value=val)
            c.font = _FONT_BOLD
            c.fill = _FILL_GREEN
            c.alignment = align
            c.border = _BORDER
            if fmt:
                c.number_format = fmt
        row_idx += 1

    # 2. Transferler — tarih sırasıyla
    for t in transfers:
        amount = float(t.amount or 0)
        if t.direction == "out":
            running_balance -= amount
            borc = amount
            alacak = None
            row_fill = _FILL_YELLOW
        else:  # in
            running_balance += amount
            borc = None
            alacak = amount
            row_fill = _FILL_BLUE

        alt = alt_map.get(t.related_request_id)
        if alt:
            ref_text = f"{alt.request_no}\n{alt.event_name}"
        else:
            ref_text = "—"

        dir_label = "Dağıtım" if t.direction == "out" else "İade"
        desc = t.description or f"Fon transferi ({dir_label})"

        cells = [
            (t.transfer_date,  _ALIGN_LEFT,  None),
            (desc,             _ALIGN_LEFT,  None),
            (ref_text,         _ALIGN_LEFT,  None),
            (borc,             _ALIGN_RIGHT, money_fmt),
            (alacak,           _ALIGN_RIGHT, money_fmt),
            (running_balance,  _ALIGN_RIGHT, money_fmt),
        ]
        for col, (val, align, fmt) in enumerate(cells, 1):
            c = ws2.cell(row=row_idx, column=col, value=val)
            c.font = _FONT_BODY
            c.fill = row_fill
            c.alignment = align
            c.border = _BORDER
            if fmt:
                c.number_format = fmt
        row_idx += 1

    # Toplam satırı
    total_borc   = balance["out_total"]
    total_alacak = balance["initial"] + balance["in_total"]
    final_row_data = [
        ("",                "",          "TOPLAM",
         total_borc, total_alacak, balance["remaining"])
    ]
    for col, val in enumerate(final_row_data[0], 1):
        c = ws2.cell(row=row_idx, column=col, value=val)
        c.font = _FONT_HEADER
        c.fill = _FILL_HEADER
        c.alignment = _ALIGN_RIGHT if col >= 4 else _ALIGN_CENT
        c.border = _BORDER
        if col >= 4:
            c.number_format = money_fmt

    # Donmuş başlık satırı
    ws2.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────────────
    # Sheet 3: Alt Referans Özeti
    # ─────────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Alt Referans Özeti")
    headers3 = ["Referans No", "Etkinlik Adı", "Etkinlik Tarihi",
                f"Toplam Dağıtım ({cur})", f"Toplam İade ({cur})", f"Net Kullanım ({cur})",
                "İşlem Sayısı"]
    widths3 = [18, 35, 14, 20, 18, 20, 14]
    for i, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    for col, hd in enumerate(headers3, 1):
        c = ws3.cell(row=1, column=col, value=hd)
        c.font = _FONT_HEADER
        c.fill = _FILL_HEADER
        c.alignment = _ALIGN_CENT
        c.border = _BORDER
    ws3.row_dimensions[1].height = 22

    # Toplama
    by_ref: dict = {}
    for t in transfers:
        d = by_ref.setdefault(t.related_request_id, {"out": 0.0, "in": 0.0, "count": 0})
        d["count"] += 1
        if t.direction == "out":
            d["out"] += float(t.amount or 0)
        else:
            d["in"] += float(t.amount or 0)

    row_idx = 2
    for rid, agg in sorted(by_ref.items(), key=lambda x: -(x[1]["out"] - x[1]["in"])):
        alt = alt_map.get(rid)
        net = agg["out"] - agg["in"]
        cells = [
            (alt.request_no if alt else "(silinmiş)", _ALIGN_LEFT, None),
            (alt.event_name if alt else "—",          _ALIGN_LEFT, None),
            (alt.check_in if alt else "—",            _ALIGN_LEFT, None),
            (agg["out"],                              _ALIGN_RIGHT, money_fmt),
            (agg["in"],                               _ALIGN_RIGHT, money_fmt),
            (net,                                     _ALIGN_RIGHT, money_fmt),
            (agg["count"],                            _ALIGN_CENT, "0"),
        ]
        for col, (val, align, fmt) in enumerate(cells, 1):
            c = ws3.cell(row=row_idx, column=col, value=val)
            c.font = _FONT_BOLD if col == 6 else _FONT_BODY
            c.alignment = align
            c.border = _BORDER
            if fmt:
                c.number_format = fmt
        row_idx += 1

    # Toplam satırı
    total_row = (
        ("", "", "TOPLAM",
         balance["out_total"], balance["in_total"],
         balance["out_total"] - balance["in_total"],
         len(transfers))
    )
    for col, val in enumerate(total_row, 1):
        c = ws3.cell(row=row_idx, column=col, value=val)
        c.font = _FONT_HEADER
        c.fill = _FILL_HEADER
        c.alignment = _ALIGN_RIGHT if col >= 4 else _ALIGN_CENT
        c.border = _BORDER
        if 4 <= col <= 6:
            c.number_format = money_fmt

    ws3.freeze_panes = "A2"

    # Çıktıyı bytes'a yaz
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
