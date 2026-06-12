"""
Raporlar
"""

from datetime import date
from collections import defaultdict
from fastapi import APIRouter, Depends, Request
from fastapi.responses import HTMLResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, extract, or_

from fastapi import Form, HTTPException
from fastapi.responses import RedirectResponse
from auth import get_current_user, require_admin, require_gm, require_mudur, require_module, get_company_id
from database import get_db
from models import (
    Invoice, GeneralExpense, CashEntry, BankMovement,
    CreditCardStatement, CreditCardTxn, Cheque, Customer, Vendor, VendorPrepayment,
    Employee, SalaryPayment, EmployeeBenefit, User,
    AnnualBudget, BudgetLine, FixedExpense, GeneralExpenseCategory,
    PayrollRecord, PayrollSettings,
)
from templates_config import templates

router = APIRouter(prefix="/reports", tags=["reports"])


@router.get("/pl", response_class=HTMLResponse, name="report_pl")
async def report_pl(
    request: Request,
    year: int = None,
    current_user: User = Depends(require_mudur),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    if not year:
        year = date.today().year

    invoices = db.query(Invoice).filter(
        Invoice.company_id == cid,
        Invoice.status.in_(["approved", "partial", "paid"]),
        extract("year", Invoice.invoice_date) == year,
        Invoice.deleted_at == None,  # noqa: E711
        or_(Invoice.approval_status.is_(None), Invoice.approval_status != "onay_bekliyor"),
    ).all()

    # Gelir tarafı: kesilen + komisyon + iade_kesilen (vendor'a iade → maliyet azalır → gelir)
    kesilen     = sum(i.amount for i in invoices if i.invoice_type in ("kesilen", "komisyon"))
    iade_kesilen = sum(i.amount for i in invoices if i.invoice_type == "iade_kesilen")
    net_gelir   = kesilen + iade_kesilen

    # Maliyet tarafı: gelen + iade_gelen (müşteri iadesi → gelir azalır → maliyet)
    gelen      = sum(i.amount for i in invoices if i.invoice_type == "gelen")
    iade_gelen = sum(i.amount for i in invoices if i.invoice_type == "iade_gelen")
    net_maliyet = gelen + iade_gelen

    expenses = db.query(GeneralExpense).filter(
        GeneralExpense.company_id == cid,
        extract("year", GeneralExpense.expense_date) == year,
    ).all()
    personel_gider = sum(e.amount for e in expenses if e.source in ("salary", "benefit"))
    diger_gider = sum(e.amount for e in expenses if e.source not in ("salary", "benefit"))

    gross_profit = net_gelir - net_maliyet
    net_profit = gross_profit - personel_gider - diger_gider

    return templates.TemplateResponse(
        "reports/pl.html",
        {
            "request": request, "current_user": current_user,
            "year": year,
            "kesilen": kesilen, "iade_kesilen": iade_kesilen,
            "net_gelir": net_gelir,
            "gelen": gelen, "iade_gelen": iade_gelen,
            "net_maliyet": net_maliyet,
            "gross_profit": gross_profit,
            "personel_gider": personel_gider,
            "diger_gider": diger_gider,
            "net_profit": net_profit,
            "page_title": f"P&L — {year}",
        },
    )


def _cc_due_date(card, txn_date):
    """KK txn için ödeme son tarihini kart ayarlarından hesaplar."""
    import calendar as _cal
    from datetime import timedelta as _td
    sd = card.statement_day or 1
    offset = card.payment_offset_days or 10
    if txn_date.day < sd:
        try:
            close = txn_date.replace(day=sd)
        except ValueError:
            close = txn_date.replace(day=_cal.monthrange(txn_date.year, txn_date.month)[1])
    else:
        m, y = txn_date.month + 1, txn_date.year
        if m > 12:
            m, y = 1, y + 1
        try:
            close = txn_date.replace(year=y, month=m, day=sd)
        except ValueError:
            close = date(y, m, _cal.monthrange(y, m)[1])
    return close + _td(days=offset)


def _parse_payment_term_days(s, default: int = 30) -> int:
    """Customer.payment_term serbest metin → gün. 'peşin'→0, '30 gün'→30, boş→default."""
    import re
    if not s:
        return default
    t = str(s).strip().lower()
    if "peşin" in t or "pesin" in t or "peş" in t:
        return 0
    m = re.search(r"\d+", t)
    return int(m.group()) if m else default


@router.get("/cash-flow", response_class=HTMLResponse, name="report_cash_flow")
async def report_cash_flow(
    request: Request,
    weeks: int = 8,
    forecast: int = 1,
    current_user: User = Depends(require_module("reports_financial")),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    from datetime import timedelta
    today = date.today()
    week_start = today - timedelta(days=today.weekday())  # bu haftanın Pazartesi'si
    weeks_end = week_start + timedelta(weeks=weeks)

    # KK txn'larını önceden hesapla (statement'lı ve statement'sız)
    cc_txns_due = []  # list of (txn, due_date)
    for txn in (
        db.query(CreditCardTxn)
        .join(CreditCardStatement, CreditCardTxn.statement_id == CreditCardStatement.id)
        .filter(
            CreditCardTxn.company_id == cid,
            CreditCardTxn.is_refund == False,
            CreditCardStatement.status == "unpaid",
            CreditCardStatement.due_date >= week_start,
            CreditCardStatement.due_date <= weeks_end,
        ).all()
    ):
        cc_txns_due.append((txn, txn.statement.due_date))

    for txn in (
        db.query(CreditCardTxn)
        .filter(
            CreditCardTxn.company_id == cid,
            CreditCardTxn.is_refund == False,
            CreditCardTxn.statement_id == None,  # noqa: E711
        ).all()
    ):
        if txn.card:
            due = _cc_due_date(txn.card, txn.txn_date)
            if week_start <= due <= weeks_end:
                cc_txns_due.append((txn, due))

    from collections import defaultdict as _dd

    # KK harcamalarını KART + son ödeme günü bazında KÜMÜLATİF topla → tek "ekstre ödeme"
    # satırı (satır satır yazmak yerine).
    _cc_agg = _dd(lambda: {"amount": 0.0, "card_name": "Kredi Kartı"})
    for txn, due in cc_txns_due:
        key = ((txn.card_id if txn.card else None) or "?", due)
        _cc_agg[key]["amount"] += (txn.amount or 0)
        if txn.card:
            _cc_agg[key]["card_name"] = txn.card.name
    cc_statements = [
        {"card_name": v["card_name"], "due_date": k[1], "amount": round(v["amount"], 2)}
        for k, v in _cc_agg.items()
    ]

    # --- Sabit Gider projeksiyonlarını önceden hesapla ---
    import calendar as _cal

    # 8 haftanın kapsadığı ay-yıl kombinasyonları
    covered_months: set = set()
    for i in range(weeks):
        d = week_start + timedelta(weeks=i)
        covered_months.add((d.year, d.month))
        covered_months.add(((d + timedelta(days=6)).year, (d + timedelta(days=6)).month))

    # Aktif sabit giderler
    active_fixed = db.query(FixedExpense).filter(
        FixedExpense.company_id == cid,
        FixedExpense.active == True,  # noqa: E712
    ).all()

    # date → list[dict]
    proj_outflows: dict = _dd(list)

    for yr, mo in covered_months:
        # Sabit gider projeksiyonu
        for fe in active_fixed:
            if mo not in _fixed_expense_months(fe, yr):
                continue
            try:
                fe_date = date(yr, mo, fe.start_date.day)
            except ValueError:
                fe_date = date(yr, mo, _cal.monthrange(yr, mo)[1])
            proj_outflows[fe_date].append({
                "type": "fixed",
                "label": fe.label,
                "sub": "Sabit Gider",
                "date": fe_date,
                "amount": fe.amount,
            })
    # ---------------------------------------------------------------

    # --- Gelen faturalar: TEDARİKÇİ ÖN ÖDEMESİ DÜŞÜLEREK önceden hesapla ---
    # Tedarikçiye yapılmış ön ödeme bir kez nakit çıkışı olarak görünür (ödeme tarihinde);
    # fatura geldiğinde tüm borç DEĞİL, KALAN (borç − ön ödeme) vadesinde eklenir.
    from sqlalchemy import func as _sa_func
    _prepay_bal = _dd(float)
    for vp in db.query(VendorPrepayment).filter(
        VendorPrepayment.company_id == cid,
        VendorPrepayment.payment_type == "prepayment",
    ).all():
        if vp.vendor_id:
            _prepay_bal[vp.vendor_id] += (vp.amount or 0)

    gelen_outflows = []
    _eff_inv_date = _sa_func.coalesce(Invoice.gm_postpone_until, Invoice.due_date)
    for inv in db.query(Invoice).filter(
        Invoice.company_id == cid,
        Invoice.invoice_type == "gelen",
        Invoice.status == "approved",
        Invoice.deleted_at == None,  # noqa: E711
        _eff_inv_date >= week_start,
        _eff_inv_date <= weeks_end,
    ).order_by(_eff_inv_date).all():
        eff_date = inv.gm_postpone_until or inv.due_date
        # Ödenecek tutar KDV DAHİL (gerçek borç) — gelir tarafı da total_with_vat kullanır.
        # GM kısmi onay verdiyse o tutar (zaten KDV dahil) geçerli.
        amt = inv.gm_approved_amount or inv.total_with_vat or inv.amount or 0
        applied = 0.0
        if inv.vendor_id and _prepay_bal.get(inv.vendor_id, 0) > 0:
            applied = min(amt, _prepay_bal[inv.vendor_id])
            _prepay_bal[inv.vendor_id] -= applied
        net = round(amt - applied, 2)
        if net <= 0:
            continue  # tamamı ön ödemeden karşılandı
        suffix = (" · GM ileri vade" if inv.gm_postpone_until else "") + \
                 (" · ön ödeme düşüldü" if applied > 0 else "")
        gelen_outflows.append({
            "type": "invoice",
            "label": (inv.vendor.name if inv.vendor else (inv.invoice_no or f"Fatura #{inv.id}")) + suffix,
            "sub": inv.reference.ref_no if inv.reference else "",
            "date": eff_date,
            "amount": net,
        })

    # ── Tahmini tahsilat (event onaylı bütçelerinden) — sıfır veri girişi ──
    # Beklenen = onaylı bütçe satışı (KDV dahil) − o referansın kesilen faturaları.
    # Tarih = etkinlik bitiş + müşteri vadesi. Kesilen fatura geldikçe otomatik kapanır.
    forecast_items_all = []
    if forecast:
        from models import DeskRequest, DeskBudget
        fc_reqs = db.query(DeskRequest).filter(
            DeskRequest.company_id == cid,
            DeskRequest.confirmed_budget_id.isnot(None),
            DeskRequest.status.notin_(["cancelled", "closed"]),
        ).all()
        for r in fc_reqs:
            bgt = db.query(DeskBudget).filter(DeskBudget.id == r.confirmed_budget_id).first()
            if not bgt:
                continue
            expected = bgt.grand_sale or 0.0
            if expected <= 0:
                continue
            invoiced = sum(
                (inv.total_with_vat or inv.amount or 0.0)
                for inv in db.query(Invoice).filter(
                    Invoice.request_id == r.id,
                    Invoice.invoice_type == "kesilen",
                    Invoice.deleted_at == None,  # noqa: E711
                ).all()
            )
            draft = round(expected - invoiced, 2)
            if draft <= 0:
                continue
            base = r.check_out or r.check_in
            if not base:
                continue
            cust = (
                db.query(Customer).filter(Customer.id == r.customer_id).first()
                if r.customer_id else None
            )
            term = _parse_payment_term_days(cust.payment_term if cust else None)
            try:
                eff = date.fromisoformat(base) + timedelta(days=term)
            except Exception:
                continue
            forecast_items_all.append({
                "type":       "forecast",
                "date":       eff,
                "amount":     draft,
                "label":      (cust.name if cust else (r.client_name or "—")),
                "sub":        f"Onaylı bütçe · {r.request_no or ''}",
                "request_no": r.request_no or "",
                "event_name": r.event_name or "",
            })

    # ── Tahmini gider (tedarikçi ödeme taahhütleri) — Faz 2 ──
    # Taahhüt − bağlı gelen faturalar = beklenen ödeme; tarih = taahhüdün beklenen tarihi.
    forecast_out_all = []
    if forecast:
        from models import DeskSupplierCommitment, DeskRequest
        from collections import defaultdict as _dd2
        commits = db.query(DeskSupplierCommitment).filter(
            DeskSupplierCommitment.company_id == cid,
            DeskSupplierCommitment.status == "open",
        ).all()
        _groups = _dd2(list)
        for cm in commits:
            _groups[(cm.request_id, cm.vendor_id)].append(cm)
        for (rid, vid), cms in _groups.items():
            q = db.query(Invoice).filter(
                Invoice.company_id == cid,
                Invoice.request_id == rid,
                Invoice.invoice_type == "gelen",
                Invoice.deleted_at == None,  # noqa: E711
            )
            if vid:
                q = q.filter(Invoice.vendor_id == vid)
            invoiced = sum((inv.total_with_vat or inv.amount or 0.0) for inv in q.all())
            _req = db.query(DeskRequest).filter(DeskRequest.id == rid).first()
            for cm in sorted(cms, key=lambda c: c.expected_payment_date or "9999"):
                amt = cm.amount or 0.0
                alloc = min(amt, invoiced)
                invoiced -= alloc
                remaining = round(amt - alloc, 2)
                if remaining <= 0:
                    continue
                try:
                    eff = date.fromisoformat(cm.expected_payment_date)
                except Exception:
                    continue
                forecast_out_all.append({
                    "date":         eff,
                    "amount":       remaining,
                    "vendor_name":  cm.vendor_name or "—",
                    "section":      cm.section,
                    "payment_type": cm.payment_type,
                    "request_no":   (_req.request_no if _req else ""),
                })

    weeks_data = []
    for i in range(weeks):
        wstart = week_start + timedelta(weeks=i)
        wend = wstart + timedelta(days=6)
        label = f"H{i + 1}" if i > 0 else "Bu Hafta"
        forecast_in = [f for f in forecast_items_all if wstart <= f["date"] <= wend]
        forecast_out = [f for f in forecast_out_all if wstart <= f["date"] <= wend]

        incoming = []  # gelir: tahsilat beklenen kesilen faturalar + kasa/banka girişleri
        outgoing = []  # gider: ödeme bekleyen gelen faturalar, çekler, KK ekstreler, kasa/banka çıkışları

        # Kesilen faturalar → collection_date (veya due_date) baz alınarak beklenen tahsilat
        from sqlalchemy import or_, and_, case as sa_case
        for inv in db.query(Invoice).filter(
            Invoice.company_id == cid,
            Invoice.invoice_type.in_(["kesilen", "komisyon"]),
            Invoice.status == "approved",
            Invoice.deleted_at == None,  # noqa: E711
            or_(
                and_(Invoice.collection_date != None, Invoice.collection_date >= wstart, Invoice.collection_date <= wend),  # noqa: E711
                and_(Invoice.collection_date == None, Invoice.due_date >= wstart, Invoice.due_date <= wend),  # noqa: E711
            ),
        ).all():
            eff_date = inv.collection_date or inv.due_date
            incoming.append({
                "type": "invoice",
                "label": inv.reference.ref_no if inv.reference else (inv.invoice_no or f"Fatura #{inv.id}"),
                "sub": inv.customer.name if inv.customer else (inv.vendor.name if inv.vendor else ""),
                "date": eff_date,
                "amount": inv.total_with_vat,
                "invoice_id": inv.id,
            })

        # Kasa girişleri
        for e in db.query(CashEntry).filter(
            CashEntry.company_id == cid,
            CashEntry.entry_type == "giris",
            CashEntry.entry_date >= wstart,
            CashEntry.entry_date <= wend,
        ).all():
            incoming.append({
                "type": "cash",
                "label": e.description or "Kasa Girişi",
                "sub": "Kasa",
                "date": e.entry_date,
                "amount": e.amount,
            })

        # Banka girişleri
        for m in db.query(BankMovement).filter(
            BankMovement.company_id == cid,
            BankMovement.movement_type == "giris",
            BankMovement.movement_date >= wstart,
            BankMovement.movement_date <= wend,
        ).all():
            incoming.append({
                "type": "bank",
                "label": m.description or "Banka Girişi",
                "sub": m.account.name if m.account else "Banka",
                "date": m.movement_date,
                "amount": m.amount,
            })

        # Gelen faturalar (ön ödeme düşülmüş, önceden hesaplanmış)
        for go in gelen_outflows:
            if wstart <= go["date"] <= wend:
                outgoing.append(go)

        # Verilen çekler → GM ileri vade onayı varsa gm_postpone_until, yoksa due_date
        from sqlalchemy import func as _sa_func
        eff_cheque_date = _sa_func.coalesce(Cheque.gm_postpone_until, Cheque.due_date)
        for c in db.query(Cheque).filter(
            Cheque.company_id == cid,
            Cheque.cheque_type == "verilen",
            Cheque.status == "beklemede",
            eff_cheque_date >= wstart,
            eff_cheque_date <= wend,
        ).all():
            eff_date = c.gm_postpone_until or c.due_date
            postponed = bool(c.gm_postpone_until)
            label_suffix = " · GM ileri vade" if postponed else ""
            outgoing.append({
                "type": "cheque",
                "label": f"Çek — {c.cheque_no or c.id}{label_suffix}",
                "sub": c.vendor.name if c.vendor else "",
                "date": eff_date,
                "amount": c.gm_approved_amount or c.amount,
            })

        # KK ekstreleri → kart bazında kümülatif tek satır (önceden hesaplanmış)
        for st in cc_statements:
            if wstart <= st["due_date"] <= wend:
                outgoing.append({
                    "type": "cc_statement",
                    "label": f"{st['card_name']} kredi kartı ekstre ödeme",
                    "sub": "Kredi Kartı Ekstre",
                    "date": st["due_date"],
                    "amount": st["amount"],
                })

        # Kasa çıkışları
        for e in db.query(CashEntry).filter(
            CashEntry.company_id == cid,
            CashEntry.entry_type == "cikis",
            CashEntry.entry_date >= wstart,
            CashEntry.entry_date <= wend,
        ).all():
            outgoing.append({
                "type": "cash",
                "label": e.description or "Kasa Çıkışı",
                "sub": "Kasa",
                "date": e.entry_date,
                "amount": e.amount,
            })

        # Banka çıkışları
        for m in db.query(BankMovement).filter(
            BankMovement.company_id == cid,
            BankMovement.movement_type == "cikis",
            BankMovement.movement_date >= wstart,
            BankMovement.movement_date <= wend,
        ).all():
            outgoing.append({
                "type": "bank",
                "label": m.description or "Banka Çıkışı",
                "sub": m.account.name if m.account else "Banka",
                "date": m.movement_date,
                "amount": m.amount,
            })

        # Maaş + Sabit Gider projeksiyonları
        for proj_date, items in proj_outflows.items():
            if wstart <= proj_date <= wend:
                outgoing.extend(items)

        total_in = sum(x["amount"] for x in incoming)
        total_out = sum(x["amount"] for x in outgoing)

        weeks_data.append({
            "label": label,
            "start": wstart.strftime("%d.%m"),
            "end": wend.strftime("%d.%m"),
            "total_in": total_in,
            "total_out": total_out,
            "incoming": sorted(incoming, key=lambda x: x["date"]),
            "outgoing": sorted(outgoing, key=lambda x: x["date"]),
            "forecast_in": sorted(forecast_in, key=lambda x: x["date"]),
            "total_fc_in": round(sum(f["amount"] for f in forecast_in), 2),
            "forecast_out": sorted(forecast_out, key=lambda x: x["date"]),
            "total_fc_out": round(sum(f["amount"] for f in forecast_out), 2),
        })

    # Vadesi geçmiş ödenmemiş faturalar (kesilen)
    overdue = db.query(Invoice).filter(
        Invoice.company_id == cid,
        Invoice.invoice_type.in_(["kesilen", "komisyon"]),
        Invoice.status == "approved",
        Invoice.deleted_at == None,  # noqa: E711
        Invoice.due_date < today,
        Invoice.due_date.isnot(None),
    ).all()
    total_overdue = sum(i.amount for i in overdue)

    return templates.TemplateResponse(
        "reports/cash_flow.html",
        {
            "request": request, "current_user": current_user,
            "weeks_data": weeks_data, "weeks": weeks,
            "overdue": overdue, "total_overdue": total_overdue,
            "show_forecast": bool(forecast),
            "total_forecast": round(sum(f["amount"] for f in forecast_items_all), 2),
            "total_forecast_out": round(sum(f["amount"] for f in forecast_out_all), 2),
            "page_title": "Nakit Akışı",
        },
    )


@router.get("/ledger/customer/{customer_id}", response_class=HTMLResponse, name="report_customer_ledger")
async def report_customer_ledger(
    customer_id: str,
    request: Request,
    current_user: User = Depends(require_mudur),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    from models import Reference
    customer = db.query(Customer).filter(
        Customer.company_id == cid, Customer.id == customer_id
    ).first()
    if not customer:
        from fastapi import HTTPException
        raise HTTPException(status_code=404)
    refs = db.query(Reference).filter(
        Reference.company_id == cid, Reference.customer_id == customer_id
    ).all()
    ref_ids = [r.id for r in refs]
    invoices = db.query(Invoice).filter(
        Invoice.company_id == cid,
        Invoice.ref_id.in_(ref_ids),
        Invoice.status.in_(["approved", "paid"]),
        Invoice.deleted_at == None,  # noqa: E711
    ).order_by(Invoice.invoice_date).all()

    total_kesilen = sum(i.amount for i in invoices if i.invoice_type == "kesilen")
    total_paid = sum(i.amount for i in invoices if i.status == "paid" and i.invoice_type == "kesilen")
    balance = total_kesilen - total_paid

    return templates.TemplateResponse(
        "reports/ledger.html",
        {
            "request": request, "current_user": current_user,
            "entity": customer, "entity_type": "customer",
            "invoices": invoices,
            "total_kesilen": total_kesilen, "total_paid": total_paid,
            "balance": balance,
            "page_title": f"Müşteri Cari — {customer.name}",
        },
    )


@router.get("/ledger/vendor/{vendor_id}", response_class=HTMLResponse, name="report_vendor_ledger")
async def report_vendor_ledger(
    vendor_id: str,
    request: Request,
    current_user: User = Depends(require_mudur),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    vendor = db.query(Vendor).filter(
        Vendor.company_id == cid, Vendor.id == vendor_id
    ).first()
    if not vendor:
        from fastapi import HTTPException
        raise HTTPException(status_code=404)
    invoices = db.query(Invoice).filter(
        Invoice.company_id == cid,
        Invoice.vendor_id == vendor_id,
        Invoice.status.in_(["approved", "paid"]),
        Invoice.deleted_at == None,  # noqa: E711
    ).order_by(Invoice.invoice_date).all()

    total_gelen = sum(i.amount for i in invoices if i.invoice_type == "gelen")
    total_paid = sum(i.amount for i in invoices if i.status == "paid" and i.invoice_type == "gelen")
    balance = total_gelen - total_paid

    return templates.TemplateResponse(
        "reports/ledger.html",
        {
            "request": request, "current_user": current_user,
            "entity": vendor, "entity_type": "vendor",
            "invoices": invoices,
            "total_gelen": total_gelen, "total_paid": total_paid,
            "balance": balance,
            "page_title": f"Tedarikçi Cari — {vendor.name}",
        },
    )


@router.get("/payroll", response_class=HTMLResponse, name="report_payroll")
async def report_payroll(
    request: Request,
    period: str = None,
    current_user: User = Depends(require_mudur),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    if not period:
        period = date.today().strftime("%Y-%m")

    # Yeni bordro modülünden onaylı + ödenmiş kayıtlar
    records = (
        db.query(PayrollRecord)
        .filter(
            PayrollRecord.company_id == cid,
            PayrollRecord.period == period,
            PayrollRecord.status.in_(["onaylandi", "odendi"]),
        )
        .join(Employee, PayrollRecord.employee_id == Employee.id)
        .order_by(Employee.name)
        .all()
    )

    # Tüm dönemleri seçim listesi için çek
    all_periods = [
        r[0] for r in db.query(PayrollRecord.period)
        .filter(
            PayrollRecord.company_id == cid,
            PayrollRecord.status.in_(["onaylandi", "odendi"]),
        )
        .distinct()
        .order_by(PayrollRecord.period.desc())
        .all()
    ]

    totals = {
        "gross":        round(sum(r.total_gross       or 0 for r in records), 2),
        "sgk_emp":      round(sum((r.sgk_employee     or 0) + (r.unemployment_emp  or 0) for r in records), 2),
        "sgk_empl":     round(sum((r.sgk_employer     or 0) + (r.unemployment_empl or 0) for r in records), 2),
        "income_tax":   round(sum(r.income_tax        or 0 for r in records), 2),
        "stamp_tax":    round(sum(r.stamp_tax         or 0 for r in records), 2),
        "net":          round(sum(r.ele_gecen         or 0 for r in records), 2),
        "employer_cost":round(sum(r.employer_cost     or 0 for r in records), 2),
    }

    return templates.TemplateResponse(
        "reports/payroll.html",
        {
            "request": request, "current_user": current_user,
            "period": period,
            "records": records,
            "all_periods": all_periods,
            "totals": totals,
            "page_title": f"Bordro Raporu — {period}",
        },
    )


# ---------------------------------------------------------------------------
# Faaliyet Raporu
# ---------------------------------------------------------------------------

def _fixed_expense_months(fe: FixedExpense, year: int) -> list[int]:
    """Verilen yılda hangi aylarda bu sabit gider gerçekleşir, liste döner."""
    months = []
    for m in range(1, 13):
        month_start = date(year, m, 1)
        # Bitiş tarihi kontrolü
        if fe.end_date and month_start > fe.end_date:
            continue
        # Başlangıç tarihi kontrolü (ay bazında)
        if date(year, m, 1) < date(fe.start_date.year, fe.start_date.month, 1):
            continue
        if fe.recurrence == "monthly":
            months.append(m)
        elif fe.recurrence == "quarterly":
            # start_date'in ayından itibaren her 3 ayda bir
            start_month = fe.start_date.month
            if (m - start_month) % 3 == 0:
                months.append(m)
        elif fe.recurrence == "yearly":
            if m == fe.start_date.month:
                months.append(m)
        elif fe.recurrence == "once":
            if year == fe.start_date.year and m == fe.start_date.month:
                months.append(m)
    return months


@router.get("/activity", response_class=HTMLResponse, name="report_activity")
async def report_activity(
    request: Request,
    year: int = None,
    current_user: User = Depends(require_mudur),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    if not year:
        year = date.today().year
    today = date.today()
    year_prev = year - 1

    def is_past(m: int) -> bool:
        return year < today.year or (year == today.year and m <= today.month)

    # Cari yıl: GeneralExpense kategori × ay
    curr_by_cat_month: dict[int, dict[int, float]] = defaultdict(lambda: defaultdict(float))
    for e in db.query(GeneralExpense).filter(
        GeneralExpense.company_id == cid,
        extract("year", GeneralExpense.expense_date) == year,
    ).all():
        curr_by_cat_month[e.category_id or 0][e.expense_date.month] += e.amount or 0

    # Önceki yıl: kategori toplamları
    prev_by_cat: dict[int, float] = defaultdict(float)
    for e in db.query(GeneralExpense).filter(
        GeneralExpense.company_id == cid,
        extract("year", GeneralExpense.expense_date) == year_prev,
    ).all():
        prev_by_cat[e.category_id or 0] += e.amount or 0

    # Bütçe satırları — hangi kategoriler seçili?
    budget = db.query(AnnualBudget).filter(
        AnnualBudget.company_id == cid, AnnualBudget.year == year
    ).first()
    budget_by_cat: dict[int, list[float]] = {}
    cats_with_bl: set[int] = set()
    if budget:
        for bl in budget.lines:
            if bl.category_id:
                budget_by_cat[bl.category_id] = [
                    getattr(bl, f"month_{m}", 0.0) for m in range(1, 13)
                ]
                cats_with_bl.add(bl.category_id)

    # Aktif kategori = BudgetLine olan + gerçek verisi olan
    cats_with_curr = {cid for cid, mm in curr_by_cat_month.items()
                      if cid != 0 and any(v > 0 for v in mm.values())}
    cats_with_prev = {cid for cid, tot in prev_by_cat.items() if cid != 0 and tot > 0}
    active_cat_id_set = cats_with_bl | cats_with_curr | cats_with_prev

    # Sadece 2 ana başlık: Personel ve Genel Giderler
    _REPORT_CATS = ["Personel (*)", "Genel Giderler"]
    top_cats = db.query(GeneralExpenseCategory).filter(
        GeneralExpenseCategory.parent_id.is_(None),
        GeneralExpenseCategory.name.in_(_REPORT_CATS),
    ).order_by(GeneralExpenseCategory.sort_order).all()

    sections = []
    grand_prev = 0.0
    grand_curr_ytd = 0.0
    grand_monthly = [0.0] * 12
    grand_forecast = 0.0

    for top_cat in top_cats:
        children = sorted(top_cat.children, key=lambda c: c.sort_order) if top_cat.children else [top_cat]
        rows = []

        for child in children:
            active = child.id in active_cat_id_set
            has_actual = child.id in (cats_with_curr | cats_with_prev)
            prev_total = prev_by_cat.get(child.id, 0.0)
            monthly = [curr_by_cat_month[child.id].get(m, 0.0) for m in range(1, 13)]
            curr_ytd = sum(monthly[m - 1] for m in range(1, 13) if is_past(m))
            budget_months = budget_by_cat.get(child.id, [0.0] * 12)
            forecast = sum(
                monthly[m - 1] if is_past(m) else budget_months[m - 1]
                for m in range(1, 13)
            )
            rows.append({
                "cat_id": child.id,
                "label": child.name,
                "active": active,
                "has_actual": has_actual,
                "prev_total": prev_total,
                "curr_ytd": curr_ytd,
                "monthly": monthly,
                "budget": budget_months,
                "forecast": forecast,
            })

        # Grand totals: sadece aktif satırlar
        active_rows = [r for r in rows if r["active"]]
        sec_prev = sum(r["prev_total"] for r in active_rows)
        sec_ytd = sum(r["curr_ytd"] for r in active_rows)
        sec_monthly = [sum(r["monthly"][i] for r in active_rows) for i in range(12)]
        sec_budget = [sum(r["budget"][i] for r in active_rows) for i in range(12)]
        sec_forecast = sum(r["forecast"] for r in active_rows)

        grand_prev += sec_prev
        grand_curr_ytd += sec_ytd
        for i in range(12):
            grand_monthly[i] += sec_monthly[i]
        grand_forecast += sec_forecast

        _lbl = "Personel" if top_cat.name == "Personel (*)" else top_cat.name
        sections.append({
            "cat_id": top_cat.id,
            "label": _lbl,
            "rows": rows,
            "prev_total": sec_prev,
            "curr_ytd": sec_ytd,
            "monthly_total": sec_monthly,
            "budget_monthly": sec_budget,
            "forecast": sec_forecast,
            "has_active": any(r["active"] for r in rows),
        })

    # Sabit giderler öngörüsü
    fixed_expenses = db.query(FixedExpense).filter(
        FixedExpense.company_id == cid, FixedExpense.active == True  # noqa: E712
    ).all()
    fixed_by_month: dict[int, float] = defaultdict(float)
    for fe in fixed_expenses:
        for m in _fixed_expense_months(fe, year):
            fixed_by_month[m] += fe.amount or 0
    fixed_monthly = [fixed_by_month.get(m, 0.0) for m in range(1, 13)]

    return templates.TemplateResponse(
        "reports/activity.html",
        {
            "request": request, "current_user": current_user,
            "year": year,
            "year_prev": year_prev,
            "today": today,
            "months_short": ["Oca", "Şub", "Mar", "Nis", "May", "Haz",
                             "Tem", "Ağu", "Eyl", "Eki", "Kas", "Ara"],
            "budget": budget,
            "sections": sections,
            "fixed_expenses": fixed_expenses,
            "fixed_monthly": fixed_monthly,
            "grand_prev": grand_prev,
            "grand_curr_ytd": grand_curr_ytd,
            "grand_monthly": grand_monthly,
            "grand_forecast": grand_forecast,
            "page_title": f"Faaliyet Raporu — {year}",
        },
    )


@router.get("/activity/export", name="report_activity_export")
async def report_activity_export(
    year: int = None,
    current_user: User = Depends(require_mudur),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    if not year:
        year = date.today().year
    today = date.today()
    year_prev = year - 1

    def is_past(m: int) -> bool:
        return year < today.year or (year == today.year and m <= today.month)

    MONTHS = ["Oca", "Şub", "Mar", "Nis", "May", "Haz",
              "Tem", "Ağu", "Eyl", "Eki", "Kas", "Ara"]

    # --- Veri hesaplama (report_activity ile aynı mantık) ---
    curr_by_cat_month: dict = defaultdict(lambda: defaultdict(float))
    for e in db.query(GeneralExpense).filter(
        GeneralExpense.company_id == cid,
        extract("year", GeneralExpense.expense_date) == year,
    ).all():
        curr_by_cat_month[e.category_id or 0][e.expense_date.month] += e.amount or 0

    prev_by_cat: dict = defaultdict(float)
    for e in db.query(GeneralExpense).filter(
        GeneralExpense.company_id == cid,
        extract("year", GeneralExpense.expense_date) == year_prev,
    ).all():
        prev_by_cat[e.category_id or 0] += e.amount or 0

    budget = db.query(AnnualBudget).filter(
        AnnualBudget.company_id == cid, AnnualBudget.year == year
    ).first()
    budget_by_cat: dict = {}
    cats_with_bl: set = set()
    if budget:
        for bl in budget.lines:
            if bl.category_id:
                budget_by_cat[bl.category_id] = [
                    getattr(bl, f"month_{m}", 0.0) for m in range(1, 13)
                ]
                cats_with_bl.add(bl.category_id)

    cats_with_curr = {cid for cid, mm in curr_by_cat_month.items()
                      if cid != 0 and any(v > 0 for v in mm.values())}
    cats_with_prev = {cid for cid, tot in prev_by_cat.items() if cid != 0 and tot > 0}
    active_cat_id_set = cats_with_bl | cats_with_curr | cats_with_prev

    _REPORT_CATS = ["Personel (*)", "Genel Giderler"]
    top_cats = db.query(GeneralExpenseCategory).filter(
        GeneralExpenseCategory.parent_id.is_(None),
        GeneralExpenseCategory.name.in_(_REPORT_CATS),
    ).order_by(GeneralExpenseCategory.sort_order).all()

    sections = []
    grand_prev = 0.0
    grand_curr_ytd = 0.0
    grand_monthly = [0.0] * 12
    grand_forecast = 0.0

    for top_cat in top_cats:
        children = sorted(top_cat.children, key=lambda c: c.sort_order) if top_cat.children else [top_cat]
        rows = []
        for child in children:
            active = child.id in active_cat_id_set
            prev_total = prev_by_cat.get(child.id, 0.0)
            monthly = [curr_by_cat_month[child.id].get(m, 0.0) for m in range(1, 13)]
            curr_ytd = sum(monthly[m - 1] for m in range(1, 13) if is_past(m))
            budget_months = budget_by_cat.get(child.id, [0.0] * 12)
            forecast = sum(
                monthly[m - 1] if is_past(m) else budget_months[m - 1]
                for m in range(1, 13)
            )
            rows.append({
                "label": child.name, "active": active,
                "prev_total": prev_total, "curr_ytd": curr_ytd,
                "monthly": monthly, "budget": budget_months, "forecast": forecast,
            })

        active_rows = [r for r in rows if r["active"]]
        sec_prev = sum(r["prev_total"] for r in active_rows)
        sec_ytd = sum(r["curr_ytd"] for r in active_rows)
        sec_monthly = [sum(r["monthly"][i] for r in active_rows) for i in range(12)]
        sec_forecast = sum(r["forecast"] for r in active_rows)
        grand_prev += sec_prev
        grand_curr_ytd += sec_ytd
        for i in range(12):
            grand_monthly[i] += sec_monthly[i]
        grand_forecast += sec_forecast

        _lbl = "Personel" if top_cat.name == "Personel (*)" else top_cat.name
        sections.append({
            "label": _lbl,
            "rows": rows,
            "prev_total": sec_prev, "curr_ytd": sec_ytd,
            "monthly_total": sec_monthly, "forecast": sec_forecast,
        })

    # --- Excel oluştur ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Faaliyet {year}"

    NUM_FMT = '#,##0.00'

    def _side():
        return Side(style="thin", color="BBBBBB")

    def _border():
        s = _side()
        return Border(left=s, right=s, top=s, bottom=s)

    # Stil sabitleri
    HDR_FILL   = PatternFill("solid", fgColor="1E293B")
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
    SEC_FILL   = PatternFill("solid", fgColor="E2E8F0")
    SEC_FONT   = Font(bold=True, color="1E293B", size=10)
    GRAND_FILL = PatternFill("solid", fgColor="1E293B")
    GRAND_FONT = Font(bold=True, color="FCA5A5", size=11)
    PAST_FONT  = Font(bold=True, color="1E293B", size=10)
    BGET_FONT  = Font(italic=True, color="64748B", size=9)
    CENTER     = Alignment(horizontal="center", vertical="center")
    RIGHT      = Alignment(horizontal="right",  vertical="center")
    LEFT       = Alignment(horizontal="left",   vertical="center")

    # Sütun başlıkları: GİDER KALEMİ | ÖNCEKİ YIL | YTD | Oca..Ara | TAHMİN
    # Her ay için 2 satır: Gerçekleşen + Bütçe
    # Basit yapı: tek satır — gerçekleşen/bütçe alt alta değil, yan yana
    headers = ["Gider Kalemi", f"{year_prev} Gerç.", f"{year} YTD",
               *MONTHS, "Tahmin"]
    bget_headers = ["", "", "", *[f"{m} (Bütçe)" for m in MONTHS], ""]

    # Satır 1: başlık
    ws.append(headers)
    hdr_row = ws.max_row
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(hdr_row, col)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER if col > 1 else LEFT
        cell.border = _border()

    # Satır 2: bütçe başlıkları
    ws.append(bget_headers)
    bg_row = ws.max_row
    for col, val in enumerate(bget_headers, 1):
        cell = ws.cell(bg_row, col)
        if val:
            cell.fill = PatternFill("solid", fgColor="F8FAFC")
            cell.font = BGET_FONT
            cell.alignment = CENTER
            cell.border = _border()

    def _write_row(label, prev, ytd, monthly, budget_months, forecast, is_section=False, is_grand=False):
        row_data = [label, prev or None, ytd or None, *[v or None for v in monthly], forecast or None]
        ws.append(row_data)
        r = ws.max_row
        for col in range(1, len(row_data) + 1):
            cell = ws.cell(r, col)
            cell.border = _border()
            if is_grand:
                cell.fill = GRAND_FILL
                cell.font = GRAND_FONT
                cell.alignment = RIGHT if col > 1 else LEFT
            elif is_section:
                cell.fill = SEC_FILL
                cell.font = SEC_FONT
                cell.alignment = RIGHT if col > 1 else LEFT
            else:
                cell.font = PAST_FONT
                cell.alignment = RIGHT if col > 1 else LEFT
            if col > 1 and cell.value is not None:
                cell.number_format = NUM_FMT

        # Bütçe satırı
        brow_data = ["", None, None, *[v or None for v in budget_months], None]
        ws.append(brow_data)
        br = ws.max_row
        for col in range(1, len(brow_data) + 1):
            cell = ws.cell(br, col)
            cell.border = _border()
            cell.font = BGET_FONT
            cell.alignment = RIGHT if col > 1 else LEFT
            if col > 1 and cell.value is not None:
                cell.number_format = NUM_FMT

    for section in sections:
        # Section header
        sec_budget = [sum(r["budget"][i] for r in section["rows"]) for i in range(12)]
        _write_row(
            section["label"].upper(),
            section["prev_total"], section["curr_ytd"],
            section["monthly_total"], sec_budget, section["forecast"],
            is_section=True,
        )
        for row in section["rows"]:
            if not row["active"]:
                continue
            _write_row(
                "  " + row["label"],
                row["prev_total"], row["curr_ytd"],
                row["monthly"], row["budget"], row["forecast"],
            )

    # Grand total
    _write_row(
        "TOPLAM GİDERLER",
        grand_prev, grand_curr_ytd,
        grand_monthly, [0.0] * 12, grand_forecast,
        is_grand=True,
    )

    # Sütun genişlikleri
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    for i in range(4, 4 + 12):
        ws.column_dimensions[get_column_letter(i)].width = 13
    ws.column_dimensions[get_column_letter(16)].width = 14

    ws.freeze_panes = "B3"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"faaliyet-raporu-{year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/activity/budget", name="report_activity_budget_save")
async def report_activity_budget_save(
    request: Request,
    current_user: User = Depends(require_mudur),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    import json as _json
    form = await request.form()
    year = int(form.get("year", date.today().year))

    budget = db.query(AnnualBudget).filter(
        AnnualBudget.company_id == cid, AnnualBudget.year == year
    ).first()
    if not budget:
        budget = AnnualBudget(year=year, created_by=current_user.id, company_id=cid)
        db.add(budget)
        db.flush()

    # Section sıralamasını güncelle (sort_order)
    section_order_json = form.get("section_order_json", "")
    if section_order_json:
        try:
            order = _json.loads(section_order_json)
            for idx, sec_id in enumerate(order):
                cat = db.query(GeneralExpenseCategory).get(int(sec_id))
                if cat and cat.parent_id is None:
                    cat.sort_order = idx * 10
        except Exception:
            pass

    # show_cat_N hidden input'larından hangi kategoriler seçili
    active_cat_ids: set[int] = set()
    for key in form.keys():
        if key.startswith("show_cat_"):
            try:
                active_cat_ids.add(int(key[9:]))
            except ValueError:
                pass

    # Mevcut satırları temizle, aktif olanları yeniden yaz
    db.query(BudgetLine).filter(BudgetLine.budget_id == budget.id).delete()

    for cat_id in active_cat_ids:
        cat = db.query(GeneralExpenseCategory).get(cat_id)
        if not cat:
            continue
        label = form.get(f"label_cat_{cat_id}", cat.name)
        bl = BudgetLine(
            budget_id=budget.id,
            line_type="gider",
            category_id=cat_id,
            label=label,
        )
        for m in range(1, 13):
            try:
                val = float(form.get(f"line_cat_{cat_id}_month_{m}", 0) or 0)
            except (ValueError, TypeError):
                val = 0.0
            setattr(bl, f"month_{m}", val)
        db.add(bl)

    db.commit()
    return RedirectResponse(url=f"/reports/activity?year={year}", status_code=303)


@router.post("/activity/fixed-expense/add", name="report_activity_fixed_add")
async def report_activity_fixed_add(
    label: str = Form(...),
    amount: float = Form(...),
    recurrence: str = Form("monthly"),
    start_date: str = Form(...),
    end_date: str = Form(""),
    category_id: str = Form(None),
    notes: str = Form(""),
    year: int = Form(None),
    current_user: User = Depends(require_mudur),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    from datetime import date as dt_date
    fe = FixedExpense(
        label=label.strip(),
        amount=amount,
        recurrence=recurrence,
        start_date=dt_date.fromisoformat(start_date),
        end_date=dt_date.fromisoformat(end_date) if end_date else None,
        category_id=category_id or None,
        notes=notes.strip(),
        active=True,
        created_by=current_user.id,
        company_id=cid,
    )
    db.add(fe)
    db.flush()

    redirect_year = year or dt_date.today().year

    # Kategori seçildiyse bütçeyi otomatik doldur
    if fe.category_id:
        cat = db.query(GeneralExpenseCategory).get(fe.category_id)
        bgt = db.query(AnnualBudget).filter(
            AnnualBudget.company_id == cid, AnnualBudget.year == redirect_year
        ).first()
        if not bgt:
            bgt = AnnualBudget(year=redirect_year, created_by=current_user.id, company_id=cid)
            db.add(bgt)
            db.flush()
        bl = db.query(BudgetLine).filter(
            BudgetLine.budget_id == bgt.id,
            BudgetLine.category_id == fe.category_id,
        ).first()
        if not bl:
            bl = BudgetLine(
                budget_id=bgt.id,
                line_type="gider",
                category_id=fe.category_id,
                label=cat.name if cat else fe.label,
            )
            db.add(bl)
            db.flush()
        for m in _fixed_expense_months(fe, redirect_year):
            current_val = getattr(bl, f"month_{m}", 0.0) or 0.0
            setattr(bl, f"month_{m}", current_val + fe.amount)

    db.commit()
    return RedirectResponse(url=f"/reports/activity?year={redirect_year}", status_code=303)


@router.post("/activity/fixed-expense/{fe_id}/delete", name="report_activity_fixed_delete")
async def report_activity_fixed_delete(
    fe_id: str,
    year: int = Form(None),
    current_user: User = Depends(require_mudur),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    fe = db.query(FixedExpense).filter(
        FixedExpense.company_id == cid, FixedExpense.id == fe_id
    ).first()
    if fe:
        db.delete(fe)
        db.commit()
    redirect_year = year or date.today().year
    return RedirectResponse(url=f"/reports/activity?year={redirect_year}", status_code=303)
