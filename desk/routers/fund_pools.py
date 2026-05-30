"""
Fon Havuzu yönetimi
"""

import io
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from fastapi import APIRouter, Depends, Form, HTTPException, Query, Request, status
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session

from auth import get_current_user, get_company_id, require_admin, require_module
from database import get_db
from models import Customer, FundPool, FundTransfer, Reference, User
from templates_config import templates

router = APIRouter(prefix="/fund-pools", tags=["fund-pools"])

CURRENCY_SYMBOLS = {"TRY": "₺", "USD": "$", "EUR": "€"}


def _pool_stats(pool: FundPool) -> dict:
    total_out = sum(t.amount for t in pool.transfers if t.direction == "out")
    total_in  = sum(t.amount for t in pool.transfers if t.direction == "in")
    balance   = pool.initial_amount - total_out + total_in
    return {
        "total_out": total_out,
        "total_in":  total_in,
        "balance":   balance,
    }


# ---------------------------------------------------------------------------
# Liste
# ---------------------------------------------------------------------------

@router.get("", response_class=HTMLResponse, name="fund_pools_list")
async def fund_pools_list(
    request: Request,
    current_user: User = Depends(require_module("fund_pools")),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    pools = db.query(FundPool).filter(FundPool.company_id == cid).order_by(FundPool.created_at.desc()).all()
    pools_data = [{"pool": p, **_pool_stats(p)} for p in pools]
    return templates.TemplateResponse(
        "fund_pools/list.html",
        {
            "request": request, "current_user": current_user,
            "pools_data": pools_data,
            "page_title": "Fon Havuzları",
            "CURRENCY_SYMBOLS": CURRENCY_SYMBOLS,
        },
    )


# ---------------------------------------------------------------------------
# Yeni Havuz
# ---------------------------------------------------------------------------

@router.get("/new", response_class=HTMLResponse, name="fund_pool_new_get")
async def fund_pool_new_get(
    request: Request,
    current_user: User = Depends(require_module("fund_pools", edit=True)),
    db: Session = Depends(get_db),
):
    customers = db.query(Customer).filter(Customer.active == True, Customer.company_id == current_user.company_id).order_by(Customer.name).all()  # noqa: E712
    return templates.TemplateResponse(
        "fund_pools/form.html",
        {
            "request": request, "current_user": current_user,
            "customers": customers,
            "today_str": date.today().isoformat(),
            "page_title": "Yeni Fon Havuzu",
        },
    )


@router.post("/new", name="fund_pool_new_post")
async def fund_pool_new_post(
    request: Request,
    name: str = Form(...),
    customer_id: str = Form(""),
    currency: str = Form("TRY"),
    initial_amount: float = Form(...),
    vat_rate: float = Form(0.20),
    invoice_date: str = Form(""),
    invoice_no: str = Form(""),
    year: str = Form(""),
    notes: str = Form(""),
    current_user: User = Depends(require_module("fund_pools", edit=True)),
    db: Session = Depends(get_db),
):
    pool = FundPool(
        company_id=current_user.company_id,
        name=name.strip(),
        customer_id=int(customer_id) if customer_id else None,
        currency=currency,
        initial_amount=initial_amount,
        vat_rate=vat_rate,
        invoice_date=date.fromisoformat(invoice_date) if invoice_date else None,
        invoice_no=invoice_no.strip() or None,
        year=int(year) if year else date.today().year,
        notes=notes.strip() or None,
        created_by=current_user.id,
    )
    db.add(pool)
    db.commit()
    return RedirectResponse(url=f"/fund-pools/{pool.id}", status_code=status.HTTP_302_FOUND)


# ---------------------------------------------------------------------------
# Detay
# ---------------------------------------------------------------------------

@router.get("/{pool_id}", response_class=HTMLResponse, name="fund_pool_detail")
async def fund_pool_detail(
    pool_id: str,
    request: Request,
    current_user: User = Depends(require_module("fund_pools")),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    pool = db.query(FundPool).filter(FundPool.id == pool_id, FundPool.company_id == cid).first()
    if not pool:
        raise HTTPException(status_code=404)

    stats = _pool_stats(pool)

    # Transfer'larda bağlı olmayan aktif referanslar (yeni transfer için dropdown)
    references = db.query(Reference).order_by(Reference.ref_no.desc()).all()

    return templates.TemplateResponse(
        "fund_pools/detail.html",
        {
            "request": request, "current_user": current_user,
            "pool": pool,
            "references": references,
            "today_str": date.today().isoformat(),
            "sym": CURRENCY_SYMBOLS.get(pool.currency, pool.currency),
            "page_title": f"Fon — {pool.name}",
            **stats,
        },
    )


# ---------------------------------------------------------------------------
# Transfer Ekle
# ---------------------------------------------------------------------------

@router.post("/{pool_id}/transfer", name="fund_transfer_add")
async def fund_transfer_add(
    pool_id: str,
    direction: str = Form(...),
    amount: float = Form(...),
    vat_rate: float = Form(0.20),
    exchange_rate: float = Form(1.0),
    transfer_date: str = Form(...),
    ref_id: str = Form(""),
    description: str = Form(""),
    current_user: User = Depends(require_module("fund_pools", edit=True)),
    db: Session = Depends(get_db),
):
    pool = db.query(FundPool).filter(FundPool.id == pool_id, FundPool.company_id == current_user.company_id).first()
    if not pool:
        raise HTTPException(status_code=404)

    if direction == "out":
        stats = _pool_stats(pool)
        if amount > stats["balance"] + 0.005:
            raise HTTPException(
                status_code=400,
                detail=f"Yetersiz bakiye. Kalan: {stats['balance']:,.2f} {pool.currency}",
            )

    t = FundTransfer(
        company_id=current_user.company_id,
        fund_pool_id=pool_id,
        direction=direction,
        amount=amount,
        vat_rate=vat_rate,
        exchange_rate=exchange_rate,
        transfer_date=date.fromisoformat(transfer_date),
        ref_id=int(ref_id) if ref_id else None,
        description=description.strip() or None,
        created_by=current_user.id,
    )
    db.add(t)
    db.commit()
    return RedirectResponse(url=f"/fund-pools/{pool_id}", status_code=status.HTTP_302_FOUND)


# ---------------------------------------------------------------------------
# Transfer Sil
# ---------------------------------------------------------------------------

@router.post("/{pool_id}/transfer/{tid}/delete", name="fund_transfer_delete")
async def fund_transfer_delete(
    pool_id: str,
    tid: int,
    current_user: User = Depends(require_module("fund_pools", edit=True)),
    db: Session = Depends(get_db),
):
    t = db.query(FundTransfer).filter(
        FundTransfer.id == tid,
        FundTransfer.company_id == current_user.company_id
    ).first()
    if t and t.fund_pool_id == pool_id:
        db.delete(t)
        db.commit()
    return RedirectResponse(url=f"/fund-pools/{pool_id}", status_code=status.HTTP_302_FOUND)


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------

@router.get("/{pool_id}/export", name="fund_pool_export")
async def fund_pool_export(
    pool_id: str,
    current_user: User = Depends(require_module("fund_pools")),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    pool = db.query(FundPool).filter(FundPool.id == pool_id, FundPool.company_id == cid).first()
    if not pool:
        raise HTTPException(status_code=404)

    stats = _pool_stats(pool)
    sym = CURRENCY_SYMBOLS.get(pool.currency, pool.currency)
    TL_FMT = f'"{sym}" #,##0.00'

    def _money_cell(cell, val, bold=False, color=None):
        cell.value = val
        cell.number_format = TL_FMT
        font_kwargs = {"bold": bold}
        if color:
            font_kwargs["color"] = color
        cell.font = Font(**font_kwargs)

    NAVY   = PatternFill("solid", fgColor="1A3A5C")
    BLUE   = PatternFill("solid", fgColor="1E5F8C")
    GREEN  = PatternFill("solid", fgColor="D6F4E0")
    YELLOW = PatternFill("solid", fgColor="FFF9C4")
    LBLUE  = PatternFill("solid", fgColor="DBEAFE")

    wb = openpyxl.Workbook()

    # ── Sayfa 1: Özet ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Özet"

    ws1.merge_cells("A1:D1")
    ws1["A1"] = f"FON HAVUZU — {pool.name}"
    ws1["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws1["A1"].fill = NAVY
    ws1["A1"].alignment = Alignment(horizontal="center")
    ws1.row_dimensions[1].height = 24

    ws1.append([])
    info = [
        ("Müşteri",      pool.customer.name if pool.customer else "—"),
        ("Para Birimi",  pool.currency),
        ("Yıl",          str(pool.year or "—")),
        ("Fatura No",    pool.invoice_no or "—"),
        ("Fatura Tarihi", pool.invoice_date.strftime("%d.%m.%Y") if pool.invoice_date else "—"),
        ("KDV %",        f"%{int(pool.vat_rate * 100)}"),
    ]
    for lbl, val in info:
        ws1.append([lbl, val])
        ws1.cell(ws1.max_row, 1).font = Font(bold=True)

    ws1.append([])
    ws1.append(["Başlangıç Tutarı (KDV dahil)", None])
    _money_cell(ws1.cell(ws1.max_row, 2), pool.initial_amount, bold=True)

    net_excl = pool.initial_amount / (1 + pool.vat_rate) if pool.vat_rate else pool.initial_amount
    ws1.append(["Başlangıç Tutarı (KDV hariç)", None])
    _money_cell(ws1.cell(ws1.max_row, 2), net_excl)

    ws1.append(["Toplam Dağıtılan", None])
    _money_cell(ws1.cell(ws1.max_row, 2), stats["total_out"], color="C00000")

    ws1.append(["Toplam İade", None])
    _money_cell(ws1.cell(ws1.max_row, 2), stats["total_in"], color="1A7A3C")

    ws1.append(["Kalan Bakiye", None])
    bal_color = "1A7A3C" if stats["balance"] >= 0 else "C00000"
    _money_cell(ws1.cell(ws1.max_row, 2), stats["balance"], bold=True, color=bal_color)

    ws1.append([])
    ws1.append(["Rapor Tarihi", datetime.now().strftime("%d.%m.%Y %H:%M")])
    ws1.cell(ws1.max_row, 1).font = Font(bold=True)

    for col, width in zip("ABCD", [28, 20, 12, 12]):
        ws1.column_dimensions[col].width = width

    # ── Sayfa 2: T Cetveli ────────────────────────────────────────────────
    ws2 = wb.create_sheet("T Cetveli")

    ws2.merge_cells("A1:G1")
    ws2["A1"] = f"Hesap Hareketleri — {pool.name}"
    ws2["A1"].font = Font(color="FFFFFF", bold=True, size=12)
    ws2["A1"].fill = NAVY
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 20

    headers = ["Tarih", "Tür", "İlişkili Referans", "Açıklama",
               f"Dağıtım ({sym})", f"İade ({sym})", f"Bakiye ({sym})"]
    ws2.append(headers)
    h_row = ws2.max_row
    for col_idx in range(1, len(headers) + 1):
        cell = ws2.cell(h_row, col_idx)
        cell.fill = BLUE
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Açılış satırı
    ws2.append([
        pool.invoice_date.strftime("%d.%m.%Y") if pool.invoice_date else "—",
        "Açılış", "—", "Fon açılış tutarı",
        None, None, None,
    ])
    r = ws2.max_row
    ws2.cell(r, 5).value = None
    _money_cell(ws2.cell(r, 7), pool.initial_amount, bold=True)
    for col_idx in range(1, 8):
        ws2.cell(r, col_idx).fill = GREEN

    running = pool.initial_amount
    for t in pool.transfers:
        excl = t.amount / (1 + t.vat_rate) if t.vat_rate else t.amount
        ref_no = t.reference.ref_no if t.reference else "—"
        if t.direction == "out":
            running -= t.amount
            row_data = [
                t.transfer_date.strftime("%d.%m.%Y"),
                "Dağıtım", ref_no, t.description or "—",
                None, None, None,
            ]
            ws2.append(row_data)
            r = ws2.max_row
            _money_cell(ws2.cell(r, 5), t.amount)
            _money_cell(ws2.cell(r, 7), running)
            fill = LBLUE
        else:
            running += t.amount
            row_data = [
                t.transfer_date.strftime("%d.%m.%Y"),
                "İade", ref_no, t.description or "—",
                None, None, None,
            ]
            ws2.append(row_data)
            r = ws2.max_row
            _money_cell(ws2.cell(r, 6), t.amount)
            _money_cell(ws2.cell(r, 7), running)
            fill = YELLOW
        for col_idx in range(1, 8):
            ws2.cell(r, col_idx).fill = fill

    for col, width in zip("ABCDEFG", [12, 10, 20, 30, 16, 16, 16]):
        ws2.column_dimensions[col].width = width

    # ── Sayfa 3: Referans Özeti ────────────────────────────────────────────
    ws3 = wb.create_sheet("Referans Özeti")

    ws3.merge_cells("A1:F1")
    ws3["A1"] = "Referans Bazlı Özet"
    ws3["A1"].font = Font(color="FFFFFF", bold=True, size=12)
    ws3["A1"].fill = NAVY
    ws3["A1"].alignment = Alignment(horizontal="center")

    headers3 = ["Referans No", "Etkinlik", f"Dağıtım ({sym})", f"İade ({sym})", f"Net ({sym})", "İşlem Sayısı"]
    ws3.append(headers3)
    h_row = ws3.max_row
    for col_idx in range(1, len(headers3) + 1):
        cell = ws3.cell(h_row, col_idx)
        cell.fill = BLUE
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    ref_data: dict = {}
    for t in pool.transfers:
        key = t.ref_id or 0
        if key not in ref_data:
            ref_data[key] = {"ref": t.reference, "out": 0.0, "in": 0.0, "count": 0}
        if t.direction == "out":
            ref_data[key]["out"] += t.amount
        else:
            ref_data[key]["in"] += t.amount
        ref_data[key]["count"] += 1

    for key, d in sorted(ref_data.items(), key=lambda x: -x[1]["out"]):
        ref = d["ref"]
        net = d["out"] - d["in"]
        ws3.append([
            ref.ref_no if ref else "—",
            ref.title if ref else "Referanssız",
            None, None, None,
            d["count"],
        ])
        r = ws3.max_row
        _money_cell(ws3.cell(r, 3), d["out"])
        _money_cell(ws3.cell(r, 4), d["in"])
        _money_cell(ws3.cell(r, 5), net, bold=True, color="C00000" if net > 0 else "1A7A3C")

    for col, width in zip("ABCDEF", [18, 30, 16, 16, 16, 12]):
        ws3.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"fon_{pool.name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
