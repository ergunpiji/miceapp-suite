from fastapi import APIRouter, Depends, Request, File, UploadFile, Form
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from templates_config import templates
from sqlalchemy.orm import Session
from datetime import date
from io import BytesIO
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config import url
from database import get_db
from models import Event, Participant, FlightRecord, AccommodationRecord, TransferRecord
from services.excel_parser import parse_participant_excel


# ---------------------------------------------------------------------------
# Şablon kolon tanımları
# ---------------------------------------------------------------------------
TEMPLATE_COLUMNS = [
    # (header_label, field_key, example_value)
    ("Ad",                    "first_name",       "Ahmet"),
    ("Soyad",                 "last_name",        "Yılmaz"),
    ("Şirket",                "company",          "ABC Şirketi"),
    ("Unvan",                 "title",            "Müdür"),
    ("E-posta",               "email",            "ahmet@example.com"),
    ("Telefon",               "phone",            "+90 532 000 00 00"),
    ("Yaka Kartı Adı",        "badge_name",       "Ahmet Y."),
    ("Beslenme Kısıtı",       "dietary",          "Vejetaryen"),
    ("Özel İhtiyaç",          "special_needs",    ""),
    ("Notlar",                "notes",            ""),
    # Geliş uçuşu
    ("Geliş Uçuş No",         "fi_number",        "TK 123"),
    ("Geliş Tarih",           "fi_date",          "2026-04-30"),
    ("Geliş Kalkış",          "fi_dep",           "IST"),
    ("Geliş Varış",           "fi_arr",           "ESB"),
    ("Geliş Kalkış Saati",    "fi_dep_time",      "08:00"),
    ("Geliş Varış Saati",     "fi_arr_time",      "09:15"),
    ("Geliş Koltuk",          "fi_seat",          "12A"),
    ("Geliş PNR",             "fi_pnr",           "ABC123"),
    # Dönüş uçuşu
    ("Dönüş Uçuş No",         "fo_number",        "TK 456"),
    ("Dönüş Tarih",           "fo_date",          "2026-05-02"),
    ("Dönüş Kalkış",          "fo_dep",           "ESB"),
    ("Dönüş Varış",           "fo_arr",           "IST"),
    ("Dönüş Kalkış Saati",    "fo_dep_time",      "18:00"),
    ("Dönüş Varış Saati",     "fo_arr_time",      "19:15"),
    ("Dönüş Koltuk",          "fo_seat",          "14B"),
    ("Dönüş PNR",             "fo_pnr",           "DEF456"),
    # Konaklama
    ("Otel",                  "acc_hotel",        "Hilton Ankara"),
    ("Oda No",                "acc_room_no",      ""),
    ("Oda Tipi",              "acc_room_type",    "DBL"),
    ("Konaklama Giriş",       "acc_check_in",     "2026-04-30"),
    ("Konaklama Çıkış",       "acc_check_out",    "2026-05-02"),
    ("Konaklama Notu",        "acc_notes",        ""),
]

# header_label → field_key haritası (büyük/küçük harf ve boşluk normalize edilmiş)
_HEADER_MAP = {col[0].lower().strip(): col[1] for col in TEMPLATE_COLUMNS}


def to_date(val) -> date | None:
    """String veya None → Python date objesi."""
    if not val:
        return None
    if isinstance(val, date):
        return val
    try:
        return date.fromisoformat(str(val)[:10])
    except (ValueError, TypeError):
        return None


def _cell_val(val) -> str:
    """Hücre değerini string'e çevir, None/boş → ''."""
    if val is None:
        return ""
    return str(val).strip()


router = APIRouter(prefix="/events/{event_id}/import", tags=["imports"])


# ---------------------------------------------------------------------------
# Sayfa
# ---------------------------------------------------------------------------
@router.get("/", response_class=HTMLResponse)
async def import_form(request: Request, event_id: str, db: Session = Depends(get_db)):
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        return RedirectResponse(url=url("/events"))
    return templates.TemplateResponse("imports/upload.html", {
        "request": request,
        "event": event,
        "active": "import"
    })


# ---------------------------------------------------------------------------
# Şablon İndir
# ---------------------------------------------------------------------------
@router.get("/template")
async def download_template(event_id: str):
    """Standart katılımcı Excel şablonunu indir."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Katılımcılar"

    # Renkler
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    example_fill = PatternFill("solid", fgColor="F0F4F8")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # Grup başlıkları (birleştirilmiş hücreler)
    groups = [
        ("Temel Bilgiler", 1, 10, "2563EB"),
        ("Geliş Uçuşu",   11, 18, "0369A1"),
        ("Dönüş Uçuşu",   19, 26, "0369A1"),
        ("Konaklama",      27, 32, "16A34A"),
    ]
    for label, start_col, end_col, color in groups:
        ws.merge_cells(
            start_row=1, start_column=start_col,
            end_row=1, end_column=end_col
        )
        cell = ws.cell(row=1, column=start_col, value=label)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Kolon başlıkları (2. satır)
    for col_idx, (label, _field, _example) in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Örnek satır (3. satır — gri, silinebilir)
    for col_idx, (_label, _field, example) in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=example if example else None)
        cell.fill = example_fill
        cell.font = Font(italic=True, color="94A3B8", size=9)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border

    # Boş veri satırları (4-53: 50 satır)
    for row_idx in range(4, 54):
        for col_idx in range(1, len(TEMPLATE_COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Kolon genişlikleri
    col_widths = [
        12, 12, 18, 14, 22, 16, 16, 16, 16, 16,  # temel (1-10)
        12, 13, 8,  8,  12, 12, 8,  10,            # geliş (11-18)
        12, 13, 8,  8,  12, 12, 8,  10,            # dönüş (19-26)
        20, 8,  10, 13, 13, 16,                     # konaklama (27-32)
    ]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"  # başlıklar donuk

    # Açıklama sayfası
    ws2 = wb.create_sheet(title="Açıklama")
    instructions = [
        ["Operasyon Ajanı — Katılımcı Listesi Şablonu"],
        [],
        ["NASIL KULLANILIR:"],
        ["1. 'Katılımcılar' sayfasında 3. satırı (gri örnek satır) silin"],
        ["2. Her katılımcı için bir satır doldurun"],
        ["3. Zorunlu alanlar: Ad ve Soyad"],
        ["4. Tarih formatı: YYYY-AA-GG (örn: 2026-04-30)"],
        ["5. Saat formatı: SS:DD (örn: 08:00)"],
        ["6. Uçuş bilgisi yoksa ilgili sütunları boş bırakın"],
        ["7. Konaklama bilgisi yoksa ilgili sütunları boş bırakın"],
        [],
        ["ODA TİPİ KISALTMAları:"],
        ["SGL = Tek kişilik"],
        ["DBL = Çift kişilik"],
        ["SUT = Suit"],
        ["TRP = Üç kişilik"],
        [],
        ["HAVALIMANLAR: IATA kodu kullanın (örn: IST, SAW, ESB, ADB, AYT)"],
    ]
    for row_data in instructions:
        ws2.append(row_data)
    ws2.column_dimensions["A"].width = 60
    ws2["A1"].font = Font(bold=True, size=13, color="1E3A5F")
    ws2["A3"].font = Font(bold=True, size=11)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=katilimci_sablonu.xlsx"},
    )


# ---------------------------------------------------------------------------
# Şablon ile Yükle (AI'sız direkt parse)
# ---------------------------------------------------------------------------
@router.post("/upload-template")
async def upload_template(
    request: Request,
    event_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        return RedirectResponse(url=url("/events"))

    content = await file.read()

    try:
        parsed = _parse_template(content)
    except Exception as e:
        return templates.TemplateResponse("imports/upload.html", {
            "request": request,
            "event": event,
            "error": f"Şablon okunamadı: {e}",
            "active": "import"
        })

    if not parsed:
        return templates.TemplateResponse("imports/upload.html", {
            "request": request,
            "event": event,
            "error": "Dosyada katılımcı bulunamadı. Lütfen şablonu doğru doldurun.",
            "active": "import"
        })

    return templates.TemplateResponse("imports/preview.html", {
        "request": request,
        "event": event,
        "parsed": parsed,
        "parsed_json": json.dumps(parsed, ensure_ascii=False, default=str),
        "filename": file.filename,
        "active": "import",
        "source": "template"
    })


def _parse_template(content: bytes) -> list[dict]:
    """Standart şablon formatını parse eder."""
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    # Başlık satırını bul — 1. veya 2. satır olabilir
    header_row_idx = None
    col_map: dict[int, str] = {}  # sütun indeksi → field_key

    for ri, row in enumerate(rows[:4]):  # ilk 4 satırda ara
        tmp: dict[int, str] = {}
        for ci, cell in enumerate(row):
            label = str(cell).lower().strip() if cell else ""
            if label in _HEADER_MAP:
                tmp[ci] = _HEADER_MAP[label]
        if len(tmp) >= 2:  # en az Ad + Soyad gibi 2 kolon bulunduysa başlık satırı
            header_row_idx = ri
            col_map = tmp
            break

    if header_row_idx is None or not col_map:
        raise ValueError("Başlık satırı bulunamadı. Lütfen orijinal şablonu kullanın.")

    # "Ad" ve "Soyad" kolonları zorunlu
    fn_col = next((ci for ci, fk in col_map.items() if fk == "first_name"), None)
    ln_col = next((ci for ci, fk in col_map.items() if fk == "last_name"), None)
    if fn_col is None or ln_col is None:
        raise ValueError("'Ad' veya 'Soyad' sütunu bulunamadı.")

    result = []
    for row in rows[header_row_idx + 1:]:
        # Tamamen boş satırı atla
        if not any(c for c in row if c is not None and str(c).strip()):
            continue

        def get(field_key: str) -> str:
            ci = next((c for c, fk in col_map.items() if fk == field_key), None)
            if ci is None or ci >= len(row):
                return ""
            return _cell_val(row[ci])

        first_name = get("first_name")
        last_name = get("last_name")
        if not first_name:
            continue  # zorunlu alan

        # Uçuş bilgisi var mı?
        fi_number = get("fi_number")
        flight_in = None
        if fi_number or get("fi_date") or get("fi_dep"):
            flight_in = {
                "flight_number":    fi_number or None,
                "airline":          None,
                "departure_airport": get("fi_dep") or None,
                "arrival_airport":  get("fi_arr") or None,
                "flight_date":      get("fi_date") or None,
                "departure_time":   get("fi_dep_time") or None,
                "arrival_time":     get("fi_arr_time") or None,
                "seat":             get("fi_seat") or None,
                "pnr":              get("fi_pnr") or None,
            }

        fo_number = get("fo_number")
        flight_out = None
        if fo_number or get("fo_date") or get("fo_dep"):
            flight_out = {
                "flight_number":    fo_number or None,
                "airline":          None,
                "departure_airport": get("fo_dep") or None,
                "arrival_airport":  get("fo_arr") or None,
                "flight_date":      get("fo_date") or None,
                "departure_time":   get("fo_dep_time") or None,
                "arrival_time":     get("fo_arr_time") or None,
                "seat":             get("fo_seat") or None,
                "pnr":              get("fo_pnr") or None,
            }

        # Konaklama bilgisi var mı?
        acc_hotel = get("acc_hotel")
        accommodation = None
        if acc_hotel or get("acc_check_in"):
            accommodation = {
                "hotel":       acc_hotel or None,
                "room_number": get("acc_room_no") or None,
                "room_type":   get("acc_room_type") or None,
                "check_in":    get("acc_check_in") or None,
                "check_out":   get("acc_check_out") or None,
                "notes":       get("acc_notes") or None,
            }

        result.append({
            "first_name":   first_name,
            "last_name":    last_name,
            "company":      get("company") or None,
            "title":        get("title") or None,
            "email":        get("email") or None,
            "phone":        get("phone") or None,
            "badge_name":   get("badge_name") or None,
            "dietary":      get("dietary") or None,
            "special_needs": get("special_needs") or None,
            "notes":        get("notes") or None,
            "flight_in":    flight_in,
            "flight_out":   flight_out,
            "accommodation": accommodation,
        })

    return result


# ---------------------------------------------------------------------------
# AI ile Yükle
# ---------------------------------------------------------------------------
@router.post("/upload")
async def upload_excel(
    request: Request,
    event_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        return RedirectResponse(url=url("/events"))

    content = await file.read()

    try:
        parsed = await parse_participant_excel(content, file.filename or "")
        return templates.TemplateResponse("imports/preview.html", {
            "request": request,
            "event": event,
            "parsed": parsed,
            "parsed_json": json.dumps(parsed, ensure_ascii=False, default=str),
            "filename": file.filename,
            "active": "import"
        })
    except Exception as e:
        return templates.TemplateResponse("imports/upload.html", {
            "request": request,
            "event": event,
            "error": str(e),
            "active": "import"
        })


# ---------------------------------------------------------------------------
# Onayla ve Kaydet
# ---------------------------------------------------------------------------
@router.post("/confirm")
async def confirm_import(
    request: Request,
    event_id: str,
    parsed_json: str = Form(...),
    db: Session = Depends(get_db)
):
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        return RedirectResponse(url=url("/events"))

    data = json.loads(parsed_json)
    imported = 0
    skipped = 0

    # İsim bazlı tekilleştir
    seen: set[str] = set()
    unique_data = []
    for row in data:
        key = f"{row.get('first_name','').strip().lower()}|{row.get('last_name','').strip().lower()}"
        if key in seen or not row.get("first_name"):
            skipped += 1
            continue
        seen.add(key)
        unique_data.append(row)
    data = unique_data

    # Mevcut katılımcıları atla
    existing = db.query(Participant).filter(Participant.event_id == event_id).all()
    existing_keys = {f"{p.first_name.strip().lower()}|{p.last_name.strip().lower()}" for p in existing}

    for row in data:
        key = f"{row.get('first_name','').strip().lower()}|{row.get('last_name','').strip().lower()}"
        if key in existing_keys:
            skipped += 1
            continue

        p = Participant(
            event_id=event_id,
            first_name=row.get("first_name", ""),
            last_name=row.get("last_name", ""),
            company=row.get("company"),
            title=row.get("title"),
            email=row.get("email"),
            phone=row.get("phone"),
            badge_name=row.get("badge_name"),
            dietary=row.get("dietary"),
            special_needs=row.get("special_needs"),
            notes=row.get("notes"),
        )
        db.add(p)
        db.flush()

        fi = row.get("flight_in")
        if fi:
            db.add(FlightRecord(
                participant_id=p.id,
                direction="in",
                flight_number=fi.get("flight_number"),
                airline=fi.get("airline"),
                departure_airport=fi.get("departure_airport"),
                arrival_airport=fi.get("arrival_airport"),
                flight_date=to_date(fi.get("flight_date")),
                departure_time=fi.get("departure_time"),
                arrival_time=fi.get("arrival_time"),
                seat=fi.get("seat"),
                pnr=fi.get("pnr"),
            ))

        fo = row.get("flight_out")
        if fo:
            db.add(FlightRecord(
                participant_id=p.id,
                direction="out",
                flight_number=fo.get("flight_number"),
                airline=fo.get("airline"),
                departure_airport=fo.get("departure_airport"),
                arrival_airport=fo.get("arrival_airport"),
                flight_date=to_date(fo.get("flight_date")),
                departure_time=fo.get("departure_time"),
                arrival_time=fo.get("arrival_time"),
                seat=fo.get("seat"),
                pnr=fo.get("pnr"),
            ))

        acc = row.get("accommodation")
        if acc:
            db.add(AccommodationRecord(
                participant_id=p.id,
                hotel=acc.get("hotel"),
                room_number=acc.get("room_number"),
                room_type=acc.get("room_type"),
                check_in=to_date(acc.get("check_in")),
                check_out=to_date(acc.get("check_out")),
                notes=acc.get("notes"),
            ))

        imported += 1

    db.commit()

    return RedirectResponse(
        url=url(f"/events/{event_id}/participants?imported={imported}&skipped={skipped}"),
        status_code=303
    )
