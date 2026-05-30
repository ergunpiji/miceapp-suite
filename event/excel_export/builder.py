"""
E-dem standart Excel teklif formatı.

- Sadece satış fiyatı gösterilir (maliyet/karlılık HİÇ YOK)
- KDV stratejisi: 'exclusive' (KDV hariç) | 'inclusive' (KDV dahil)
- Dinamik satırlar + bölüm başlıkları + ara toplamlar
- Para birimi: TRY / EUR / USD
"""
from __future__ import annotations

import io
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import column_index_from_string, get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ── Renk sabitleri ─────────────────────────────────────────────────────────────
C_HEADER_DARK = "1A3A5C"   # lacivert — sayfa başlığı / sütun başlığı
C_HEADER_MED  = "1E5F8C"   # orta mavi — bölüm başlığı
C_ROW_ALT     = "F0F4F8"   # açık gri  — alternatif satır
C_SUBTOTAL    = "DBEAFE"   # açık mavi — ara toplam / hizmet bedeli
C_FOOTER_BG   = "0F172A"   # koyu lacivert — genel toplam
C_WHITE       = "FFFFFF"
C_TEXT        = "1E293B"
C_LABEL_BG    = "EFF6FF"   # meta etiket arka plan

# ── Bölüm etiketleri ───────────────────────────────────────────────────────────
SECTION_LABELS: dict[str, str] = {
    "accommodation": "Konaklama",
    "meeting":       "Toplantı / Salon",
    "fb":            "F&B (Yiyecek & İçecek)",
    "teknik":        "Teknik Ekipman",
    "dekor":         "Dekor / Süsleme",
    "transfer":      "Transfer & Ulaşım",
    "tasarim":       "Tasarım & Basılı Malzeme",
    "other":         "Diğer Hizmetler",
}
SECTIONS_ORDER = [
    "accommodation", "meeting", "fb",
    "teknik", "dekor", "transfer", "tasarim", "other",
]

# ── Para birimi ────────────────────────────────────────────────────────────────
CURRENCY_INFO = {
    "TRY": {"symbol": "₺", "label": "Türk Lirası",  "fmt": '#,##0.00 "₺"'},
    "EUR": {"symbol": "€", "label": "Euro",          "fmt": '"€" #,##0.00'},
    "USD": {"symbol": "$", "label": "ABD Doları",    "fmt": '"$" #,##0.00'},
}


# ── Yardımcı stil fonksiyonları ────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border(color: str = "D1D5DB") -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _font(bold=False, color=C_TEXT, size=9, white=False) -> Font:
    return Font(bold=bold, color=C_WHITE if white else color, size=size)


def _align(h="left", v="center", wrap=False, indent=0) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)


def _write(ws, row, col_letter, value, *, bold=False, white=False, color=C_TEXT,
           fill_color=C_WHITE, size=9, h_align="left", indent=1, num_fmt=None, wrap=False):
    """Tek hücre yazar + biçimlendirme uygular."""
    c = ws.cell(row=row, column=column_index_from_string(col_letter), value=value)
    c.font      = _font(bold=bold, color=color, size=size, white=white)
    c.fill      = _fill(fill_color)
    c.border    = _border()
    c.alignment = _align(h=h_align, indent=indent if h_align == "left" else 0, wrap=wrap)
    if num_fmt:
        c.number_format = num_fmt
    return c


def _merge_write(ws, row, col_start, col_end, value, *, bold=False, white=False,
                 fill_color=C_WHITE, size=9, h_align="left", indent=1, num_fmt=None):
    """Merge eder + değer yazar."""
    ws.merge_cells(f"{col_start}{row}:{col_end}{row}")
    c = ws.cell(row=row, column=column_index_from_string(col_start), value=value)
    c.font      = _font(bold=bold, white=white, size=size)
    c.fill      = _fill(fill_color)
    c.border    = _border()
    c.alignment = _align(h=h_align, indent=indent if h_align == "left" else 0)
    if num_fmt:
        c.number_format = num_fmt
    return c


# ── Tek sheet yazar (workbook'a) ───────────────────────────────────────────────
def _write_sheet(
    ws,
    budget,
    request,
    customer,
    creator,
    vat_mode: str = "exclusive",
    custom_sections: list | None = None,
) -> None:
    """Verilen ws'e bütçe teklif sayfasını yazar."""

    currency = (budget.offer_currency or "TRY").upper()
    ci       = CURRENCY_INFO.get(currency, CURRENCY_INFO["TRY"])
    sym      = ci["symbol"]
    cur_lbl  = ci["label"]
    num_fmt  = ci["fmt"]
    rate     = budget.rate_to_try(currency)

    # ── Sütun tanımları (vat_mode'a göre) ─────────────────────────────────────
    # exclusive: A Hizmet | B Birim | C Miktar | D Gece | E BirimFiyat(hariç) | F KDV% | G Toplam(hariç) | H Not
    # inclusive: A Hizmet | B Birim | C Miktar | D Gece | E BirimFiyat(dahil) | F Toplam(dahil) | G Not
    # mixed:     A Hizmet | B İstek/Adet | C Gün | D BirimFiyat(EUR) | E KUR | F Toplam(TL) | G Not

    # mixed modu: birim fiyat yabancı para birimi, ayrı KUR kolonu, toplam TRY
    is_mixed = vat_mode == "mixed" and currency != "TRY" and rate and rate > 0
    try_ci    = CURRENCY_INFO["TRY"]
    try_fmt   = try_ci["fmt"]

    # mixed: ayrı sütun adları (COL_UNIT yok, COL_RATE ekstra)
    COL_RATE = None  # sadece mixed'da kullanılır

    if vat_mode == "inclusive":
        COL_SVC, COL_UNIT, COL_QTY, COL_NIGHT = "A", "B", "C", "D"
        COL_PRICE, COL_VAT, COL_TOTAL, COL_NOTE = "E", None, "F", "G"
        LAST_COL = "G"
        col_headers = {
            "A": ("Hizmet Adı",              30),
            "B": ("Birim",                   10),
            "C": ("Miktar",                   8),
            "D": ("Gece /\nGün",             10),
            "E": (f"Birim Fiyat\n(KDV dahil)", 18),
            "F": (f"Toplam\n(KDV dahil)",     18),
            "G": ("Not",                     32),
        }
    elif is_mixed:
        # A Hizmet | B İstek/Adet/Kişi | C Gün | D BirimFiyat(EUR) | E KUR | F Toplam(TL) | G Not
        COL_SVC   = "A"
        COL_UNIT  = None   # birim yok — qty başlığına gömülü
        COL_QTY   = "B"
        COL_NIGHT = "C"
        COL_PRICE = "D"
        COL_RATE  = "E"
        COL_VAT   = None
        COL_TOTAL = "F"
        COL_NOTE  = "G"
        LAST_COL  = "G"
        col_headers = {
            "A": ("Hizmet Adı",                    34),
            "B": ("İstek /\nAdet / Kişi",          12),
            "C": ("Gün",                            8),
            "D": (f"Birim Fiyat\n({sym})",         16),
            "E": ("KUR",                           10),
            "F": ("Toplam (TL)",                   18),
            "G": ("Not",                           28),
        }
    else:  # exclusive (varsayılan)
        COL_SVC, COL_UNIT, COL_QTY, COL_NIGHT = "A", "B", "C", "D"
        COL_PRICE, COL_VAT, COL_TOTAL, COL_NOTE = "E", "F", "G", "H"
        LAST_COL = "H"
        col_headers = {
            "A": ("Hizmet Adı",               30),
            "B": ("Birim",                    10),
            "C": ("Miktar",                    8),
            "D": ("Gece /\nGün",              10),
            "E": (f"Birim Fiyat\n(KDV hariç)", 18),
            "F": ("KDV %",                     8),
            "G": (f"Toplam\n(KDV hariç)",     18),
            "H": ("Not",                      32),
        }

    # Sütun genişlikleri
    for letter, (_, width) in col_headers.items():
        ws.column_dimensions[letter].width = width

    # ── Sayfa başlığı (satır 1) ────────────────────────────────────────────────
    ws.merge_cells(f"A1:{LAST_COL}1")
    ws["A1"].value     = "ETKİNLİK TEKLİFİ"
    ws["A1"].font      = _font(bold=True, size=14, white=True)
    ws["A1"].fill      = _fill(C_HEADER_DARK)
    ws["A1"].alignment = _align(h="center")
    ws.row_dimensions[1].height = 32

    # ── Meta bilgiler (satırlar 2–6) ───────────────────────────────────────────
    import datetime

    ref_no      = getattr(request, "request_no", "")   or ""
    ev_name     = getattr(request, "event_name",  "")  or budget.venue_name or ""
    cust_nm     = (getattr(customer, "name", None) or
                   getattr(request, "client_name", None) or "")
    venue_nm    = budget.venue_name or ""
    creator_nm  = ""
    if creator:
        creator_nm = f"{getattr(creator, 'name', '')} {getattr(creator, 'surname', '')}".strip()

    # Müşterinin birincil kontak kişisi
    contact_nm = ""
    if customer:
        contacts = getattr(customer, "contacts", None)
        if callable(contacts):
            contacts = contacts()
        if contacts and isinstance(contacts, list) and len(contacts) > 0:
            c0 = contacts[0]
            parts = [c0.get("name", ""), c0.get("title", ""), c0.get("phone", "")]
            contact_nm = "  |  ".join(p for p in parts if p)

    def _fmt_date(d) -> str:
        if not d:
            return ""
        if isinstance(d, str):
            try:
                from datetime import date
                return date.fromisoformat(d[:10]).strftime("%d.%m.%Y")
            except Exception:
                return d
        try:
            return d.strftime("%d.%m.%Y")
        except Exception:
            return str(d)

    date_str = ""
    if request:
        ci = getattr(request, "check_in",  None)
        co = getattr(request, "check_out", None)
        if ci:
            date_str = _fmt_date(ci)
            if co and co != ci:
                date_str += f" – {_fmt_date(co)}"

    if is_mixed:
        cur_display = f"{sym} birim fiyat / ₺ toplam"
    else:
        cur_display = f"{sym} {cur_lbl}"

    meta = [
        ("Referans No :", ref_no,      "Müşteri :",      cust_nm),
        ("Etkinlik :",   ev_name,      "Kontak Kişi :",  contact_nm),
        ("Mekan :",      venue_nm,     "Hazırlayan :",   creator_nm),
        ("Tarih :",      date_str,     "Para Birimi :",  cur_display),
    ]
    if currency != "TRY" and rate and rate != 1.0:
        meta.append((f"1 {sym} =", f"{rate:,.4f} ₺", "", ""))

    meta_end = 1
    for offset, (lbl_l, val_l, lbl_r, val_r) in enumerate(meta, start=2):
        r = offset
        # Sol: A-B etiket, C-D değer
        ws.merge_cells(f"A{r}:B{r}")
        cl = ws.cell(row=r, column=1, value=lbl_l)
        cl.font = _font(bold=True, size=9); cl.fill = _fill(C_LABEL_BG)
        cl.border = _border(); cl.alignment = _align(indent=1)

        ws.merge_cells(f"C{r}:D{r}")
        cv = ws.cell(row=r, column=3, value=val_l)
        cv.font = _font(size=9); cv.fill = _fill(C_WHITE)
        cv.border = _border(); cv.alignment = _align(indent=1)

        # Sağ: E-F etiket, G-H değer (exclusive=8 sütun, inclusive=7)
        if lbl_r:
            ws.merge_cells(f"E{r}:F{r}")
            cr2 = ws.cell(row=r, column=5, value=lbl_r)
            cr2.font = _font(bold=True, size=9); cr2.fill = _fill(C_LABEL_BG)
            cr2.border = _border(); cr2.alignment = _align(indent=1)

            end_c = "H" if LAST_COL == "H" else "G"
            ws.merge_cells(f"G{r}:{end_c}{r}")
            vr2 = ws.cell(row=r, column=7, value=val_r)
            vr2.font = _font(size=9); vr2.fill = _fill(C_WHITE)
            vr2.border = _border(); vr2.alignment = _align(indent=1)

        ws.row_dimensions[r].height = 16
        meta_end = r

    # Ince ayraç satırı
    sep_row = meta_end + 1
    ws.row_dimensions[sep_row].height = 6

    # ── Sütun başlık satırı ────────────────────────────────────────────────────
    hdr_row = sep_row + 1
    for letter, (label, _) in col_headers.items():
        c = ws.cell(row=hdr_row, column=column_index_from_string(letter), value=label)
        c.font      = _font(bold=True, size=9, white=True)
        c.fill      = _fill(C_HEADER_DARK)
        c.border    = _border()
        c.alignment = _align(h="center", wrap=True)
    ws.row_dimensions[hdr_row].height = 30

    current_row = hdr_row + 1

    # ── Bütçe satırlarını bölümlere ayır ──────────────────────────────────────
    rows_by_sec: dict[str, list] = defaultdict(list)
    service_fee_row       = None
    accommodation_tax_row = None
    for r in budget.rows:
        if r.get("is_service_fee"):
            service_fee_row = r
            continue
        if r.get("is_accommodation_tax"):
            accommodation_tax_row = r
            continue
        rows_by_sec[r.get("section", "other")].append(r)

    subtotal_refs: list[str] = []            # her bölümün ara toplam G adresi
    sf_base_refs:  list[str] = []            # servis bedeli tabanı: sadece data satırları (konaklama vergisi HARİÇ)
    data_rows_by_vat: dict[int, list] = defaultdict(list)  # vat% → data row G adresleri

    # Tüm bölüm sırası (standart + özel kategoriler)
    all_sections = list(SECTIONS_ORDER)
    if custom_sections:
        for cs in custom_sections:
            csid = cs.get("id") or cs.get("value", "")
            csname = cs.get("name") or cs.get("label", csid)
            if csid and csid not in all_sections:
                all_sections.append(csid)
                SECTION_LABELS[csid] = csname

    # ── Bölüm + satır döngüsü ─────────────────────────────────────────────────
    for sec in all_sections:
        sec_rows = rows_by_sec.get(sec, [])
        if not sec_rows:
            continue

        # Bölüm başlığı (merge tüm sütunlar)
        ws.merge_cells(f"A{current_row}:{LAST_COL}{current_row}")
        sh = ws.cell(row=current_row, column=1,
                     value=SECTION_LABELS.get(sec, sec))
        sh.font      = _font(bold=True, size=10, white=True)
        sh.fill      = _fill(C_HEADER_MED)
        sh.border    = _border()
        sh.alignment = _align(h="left", indent=1)
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        sec_data_start = current_row
        alt = False
        accom_data_refs: list[str] = []   # konaklama vergisi formülü için

        for row in sec_rows:
            qty    = float(row.get("qty",    1) or 1)
            nights = float(row.get("nights", 1) or 1)
            sale   = float(row.get("sale_price", 0) or 0)
            vat    = float(row.get("vat_rate",   0) or 0)

            # Para birimi dönüşümü (satır kendi para biriminde ise)
            row_cur = (row.get("currency") or "TRY").upper()
            if row_cur != currency:
                row_rate   = budget.rate_to_try(row_cur) or 1.0
                offer_rate = budget.rate_to_try(currency) or 1.0
                sale = sale * row_rate / offer_rate

            price_val = sale * (1 + vat / 100) if vat_mode == "inclusive" else sale

            bg = C_ROW_ALT if alt else C_WHITE
            alt = not alt
            r_idx = current_row  # closure bug önlemi — yerel kopya

            def _wc(letter, value, fmt=None, h="left", ind=1, _row=r_idx, _bg=bg):
                c2 = ws.cell(row=_row,
                             column=column_index_from_string(letter),
                             value=value)
                c2.font      = _font(size=9)
                c2.fill      = _fill(_bg)
                c2.border    = _border()
                c2.alignment = _align(h=h, indent=ind if h == "left" else 0)
                if fmt:
                    c2.number_format = fmt

            _wc(COL_SVC, row.get("service_name", ""))
            if COL_UNIT:
                _wc(COL_UNIT, row.get("unit", "Adet"), h="center")
            _wc(COL_QTY,   qty,    fmt="#,##0",  h="center")
            _wc(COL_NIGHT, nights, fmt="#,##0",  h="center")

            if is_mixed:
                # D: Birim Fiyat (EUR), E: KUR (kur değeri hücrede görünür), F: Toplam (TL)
                _wc(COL_PRICE, price_val, fmt=num_fmt, h="right", ind=0)
                _wc(COL_RATE,  round(rate, 4), fmt='#,##0.00', h="right", ind=0)
                # Toplam = Birim Fiyat × Miktar × Gün × KUR (hücre referansı)
                total_formula = (
                    f"={COL_PRICE}{r_idx}*{COL_QTY}{r_idx}"
                    f"*{COL_NIGHT}{r_idx}*{COL_RATE}{r_idx}"
                )
                _wc(COL_TOTAL, total_formula, fmt=try_fmt, h="right", ind=0)
            else:
                _wc(COL_PRICE, price_val, fmt=num_fmt, h="right", ind=0)
                if COL_VAT:
                    _wc(COL_VAT, vat / 100, fmt="0%", h="center")
                # Toplam sütunu: formüllü = Birim Fiyat × Miktar × Gece
                total_formula = f"={COL_PRICE}{r_idx}*{COL_QTY}{r_idx}*{COL_NIGHT}{r_idx}"
                _wc(COL_TOTAL, total_formula, fmt=num_fmt, h="right", ind=0)

            _wc(COL_NOTE,  row.get("notes", ""))
            # KDV grubuna kaydet (exclusive modda formül için)
            data_rows_by_vat[int(vat)].append(f"{COL_TOTAL}{r_idx}")
            # Servis bedeli tabanı (tüm gerçek data satırları)
            sf_base_refs.append(f"{COL_TOTAL}{r_idx}")
            # Konaklama bölümü için ayrıca takip et
            if sec == "accommodation":
                accom_data_refs.append(f"{COL_TOTAL}{r_idx}")

            ws.row_dimensions[r_idx].height = 15
            current_row += 1

        # ── Konaklama Vergisi (%2) — formüllü ─────────────────────────────────
        accom_tax_ref = None
        if sec == "accommodation" and accommodation_tax_row is not None and accom_data_refs:
            C_ACCOM_TAX = "FFF7ED"  # açık turuncu arka plan
            tax_formula = f"=SUM({','.join(accom_data_refs)})*0.02"

            merge_end = get_column_letter(column_index_from_string(COL_TOTAL) - 1)
            ws.merge_cells(f"A{current_row}:{merge_end}{current_row}")
            lc = ws.cell(row=current_row, column=1, value="Konaklama Vergisi (%2)")
            lc.font = _font(bold=True, size=9, color="92400E")
            lc.fill = _fill(C_ACCOM_TAX); lc.border = _border()
            lc.alignment = _align(h="right", indent=1)

            vc = ws.cell(row=current_row,
                         column=column_index_from_string(COL_TOTAL),
                         value=tax_formula)
            vc.font = _font(bold=True, size=9, color="92400E")
            vc.fill = _fill(C_ACCOM_TAX); vc.border = _border()
            vc.number_format = try_fmt if is_mixed else num_fmt
            vc.alignment = _align(h="right")
            accom_tax_ref = f"{COL_TOTAL}{current_row}"
            # KDV grubuna ekle (verginin KDV'si yok ama toplamda yer almalı)
            data_rows_by_vat[0].append(accom_tax_ref)

            if COL_NOTE:
                nc = ws.cell(row=current_row, column=column_index_from_string(COL_NOTE))
                nc.fill = _fill(C_ACCOM_TAX); nc.border = _border()

            ws.row_dimensions[current_row].height = 15
            current_row += 1

        # Ara toplam satırı
        sec_data_end = current_row - 1
        merge_end = get_column_letter(column_index_from_string(COL_TOTAL) - 1)
        ws.merge_cells(f"A{current_row}:{merge_end}{current_row}")
        lc = ws.cell(row=current_row, column=1,
                     value=f"{SECTION_LABELS.get(sec, sec)} Ara Toplam")
        lc.font      = _font(bold=True, size=9)
        lc.fill      = _fill(C_SUBTOTAL)
        lc.border    = _border()
        lc.alignment = _align(h="right", indent=1)

        stc = ws.cell(row=current_row,
                      column=column_index_from_string(COL_TOTAL),
                      value=f"=SUM({COL_TOTAL}{sec_data_start}:{COL_TOTAL}{sec_data_end})")
        stc.font         = _font(bold=True, size=9)
        stc.fill         = _fill(C_SUBTOTAL)
        stc.border       = _border()
        stc.number_format = try_fmt if is_mixed else num_fmt
        stc.alignment    = _align(h="right")
        subtotal_refs.append(f"{COL_TOTAL}{current_row}")

        if COL_NOTE:
            ws.cell(row=current_row,
                    column=column_index_from_string(COL_NOTE)).fill = _fill(C_SUBTOTAL)
            ws.cell(row=current_row,
                    column=column_index_from_string(COL_NOTE)).border = _border()

        ws.row_dimensions[current_row].height = 16
        current_row += 1

        # Bölümler arası ince boşluk
        ws.row_dimensions[current_row].height = 5
        current_row += 1

    # ── Hizmet Bedeli (%X) — formüllü ────────────────────────────────────────
    sf_vat      = 20.0
    sf_cell_ref = None

    if service_fee_row:
        sf_vat = float(service_fee_row.get("vat_rate", 20) or 20)

        # Servis bedeli yüzdesi
        sf_pct = float(getattr(budget, "service_fee_pct", 0) or 0)
        sf_label = f"Hizmet Bedeli (%{sf_pct:g})" if sf_pct else "Hizmet Bedeli"

        # Formül: SUM(tüm gerçek data satırları, konaklama vergisi HARİÇ) × yüzde
        # JS calcServiceFee ile birebir aynı mantık
        base_refs = sf_base_refs  # konaklama vergisi eklenmez
        if base_refs and sf_pct:
            sf_formula = f"=SUM({','.join(base_refs)})*{sf_pct/100}"
        elif base_refs:
            # Yüzde bilinmiyorsa sabit değer fallback
            sf_sale = float(service_fee_row.get("sale_price", 0) or 0)
            sf_formula = str(round(sf_sale, 2))
        else:
            sf_formula = "=0"

        if vat_mode == "inclusive":
            sf_formula = f"=({sf_formula.lstrip('=')})*{1 + sf_vat/100}" if sf_formula.startswith("=") \
                         else f"={sf_formula}*{1 + sf_vat/100}"

        merge_end = get_column_letter(column_index_from_string(COL_TOTAL) - 1)
        ws.merge_cells(f"A{current_row}:{merge_end}{current_row}")
        lc = ws.cell(row=current_row, column=1, value=sf_label)
        lc.font = _font(bold=True, size=9); lc.fill = _fill(C_SUBTOTAL)
        lc.border = _border(); lc.alignment = _align(h="right", indent=1)

        vc = ws.cell(row=current_row,
                     column=column_index_from_string(COL_TOTAL),
                     value=sf_formula)
        vc.font = _font(bold=True, size=9); vc.fill = _fill(C_SUBTOTAL)
        vc.border = _border(); vc.number_format = try_fmt if is_mixed else num_fmt; vc.alignment = _align(h="right")
        sf_cell_ref = f"{COL_TOTAL}{current_row}"
        # Servis bedeli de KDV grubuna eklenir
        data_rows_by_vat[int(sf_vat)].append(sf_cell_ref)

        if COL_NOTE:
            nc = ws.cell(row=current_row, column=column_index_from_string(COL_NOTE))
            nc.fill = _fill(C_SUBTOTAL); nc.border = _border()

        ws.row_dimensions[current_row].height = 16
        current_row += 1

    # ── GENEL TOPLAM (KDV Hariç) — formüllü ─────────────────────────────────
    ws.row_dimensions[current_row].height = 6
    current_row += 1

    all_excl_refs = subtotal_refs + ([sf_cell_ref] if sf_cell_ref else [])
    gt_excl_formula = ("=SUM(" + ",".join(all_excl_refs) + ")") if all_excl_refs else "=0"

    merge_end = get_column_letter(column_index_from_string(COL_TOTAL) - 1)
    ws.merge_cells(f"A{current_row}:{merge_end}{current_row}")
    lc = ws.cell(row=current_row, column=1,
                 value="GENEL TOPLAM (KDV Hariç)" if vat_mode == "exclusive" else "GENEL TOPLAM")
    lc.font = _font(bold=True, size=10, white=True)
    lc.fill = _fill(C_FOOTER_BG); lc.border = _border()
    lc.alignment = _align(h="right", indent=1)

    gt_excl_cell = ws.cell(row=current_row,
                           column=column_index_from_string(COL_TOTAL),
                           value=gt_excl_formula)
    gt_excl_cell.font = _font(bold=True, size=10, white=True)
    gt_excl_cell.fill = _fill(C_FOOTER_BG); gt_excl_cell.border = _border()
    gt_excl_cell.number_format = try_fmt if is_mixed else num_fmt; gt_excl_cell.alignment = _align(h="right")
    gt_excl_ref = f"{COL_TOTAL}{current_row}"

    if COL_NOTE:
        nc = ws.cell(row=current_row, column=column_index_from_string(COL_NOTE))
        nc.fill = _fill(C_FOOTER_BG); nc.border = _border()

    ws.row_dimensions[current_row].height = 20
    current_row += 1

    # ── KDV Dökümü — her grup için formüllü ──────────────────────────────────
    kdv_cell_refs: list[str] = []

    if vat_mode == "exclusive":
        ws.row_dimensions[current_row].height = 5
        current_row += 1

        for vat_rate_key in sorted(data_rows_by_vat.keys()):
            row_refs = data_rows_by_vat[vat_rate_key]
            if not row_refs or vat_rate_key == 0:
                continue

            # Formül: (G_row1 + G_row2 + ...) × KDV_oranı
            vat_decimal = vat_rate_key / 100
            kdv_formula = f"=SUM({','.join(row_refs)})*{vat_decimal}"

            merge_end = get_column_letter(column_index_from_string(COL_TOTAL) - 1)
            ws.merge_cells(f"A{current_row}:{merge_end}{current_row}")
            lc = ws.cell(row=current_row, column=1, value=f"KDV %{vat_rate_key}")
            lc.font = _font(size=9); lc.fill = _fill(C_ROW_ALT)
            lc.border = _border(); lc.alignment = _align(h="right", indent=1)

            vc = ws.cell(row=current_row,
                         column=column_index_from_string(COL_TOTAL),
                         value=kdv_formula)
            vc.font = _font(size=9); vc.fill = _fill(C_ROW_ALT)
            vc.border = _border(); vc.number_format = try_fmt if is_mixed else num_fmt
            vc.alignment = _align(h="right")
            kdv_cell_refs.append(f"{COL_TOTAL}{current_row}")

            if COL_NOTE:
                nc = ws.cell(row=current_row, column=column_index_from_string(COL_NOTE))
                nc.fill = _fill(C_ROW_ALT); nc.border = _border()

            ws.row_dimensions[current_row].height = 15
            current_row += 1

    # ── GENEL TOPLAM (KDV Dahil) — formüllü ──────────────────────────────────
    if vat_mode == "exclusive" and kdv_cell_refs:
        ws.row_dimensions[current_row].height = 5
        current_row += 1

        kdv_inc_formula = f"={gt_excl_ref}+SUM({','.join(kdv_cell_refs)})"

        merge_end = get_column_letter(column_index_from_string(COL_TOTAL) - 1)
        ws.merge_cells(f"A{current_row}:{merge_end}{current_row}")
        lc = ws.cell(row=current_row, column=1, value="GENEL TOPLAM (KDV Dahil)")
        lc.font = _font(bold=True, size=11, white=True)
        lc.fill = _fill(C_FOOTER_BG); lc.border = _border()
        lc.alignment = _align(h="right", indent=1)

        vc = ws.cell(row=current_row,
                     column=column_index_from_string(COL_TOTAL),
                     value=kdv_inc_formula)
        vc.font = _font(bold=True, size=11, white=True)
        vc.fill = _fill(C_FOOTER_BG); vc.border = _border()
        vc.number_format = try_fmt if is_mixed else num_fmt; vc.alignment = _align(h="right")

        if COL_NOTE:
            nc = ws.cell(row=current_row, column=column_index_from_string(COL_NOTE))
            nc.fill = _fill(C_FOOTER_BG); nc.border = _border()

        ws.row_dimensions[current_row].height = 24

    # Yazdırma ayarları
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.freeze_panes = f"A{hdr_row + 1}"


# ── Tek bütçe → BytesIO ────────────────────────────────────────────────────────
def build_standard(
    budget,
    request,
    customer,
    creator,
    vat_mode: str = "exclusive",
    custom_sections: list | None = None,
) -> io.BytesIO:
    """Tek bütçe için standart teklif Excel'i döndürür."""
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl kurulu değil. pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Teklif"
    _write_sheet(ws, budget, request, customer, creator, vat_mode, custom_sections)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── Bütçe toplamlarını Python'da hesapla (özet sayfası için) ─────────────────
def _calc_totals(budget, vat_mode: str = "exclusive") -> dict:
    """
    Bir bütçenin KDV hariç ve dahil toplamlarını Python'da hesaplar.
    Döner: {"excl": float, "incl": float, "currency": str}
    """
    currency = (budget.offer_currency or "TRY").upper()
    rate     = budget.rate_to_try(currency) or 1.0
    is_mixed = vat_mode == "mixed" and currency != "TRY"

    excl_total = 0.0
    vat_total  = 0.0
    sf_pct     = float(getattr(budget, "service_fee_pct", 0) or 0)
    sf_row     = None

    for r in budget.rows:
        if r.get("is_service_fee"):
            sf_row = r
            continue
        if r.get("is_accommodation_tax"):
            continue

        sale  = float(r.get("sale_price", 0) or 0)
        qty   = float(r.get("qty", 1) or 1)
        night = float(r.get("nights", 1) or 1) or 1.0
        vat   = float(r.get("vat_rate", 20) or 20)
        line  = sale * qty * night
        if is_mixed:
            line = line * rate
        excl_total += line
        vat_total  += line * vat / 100

    # Hizmet bedeli
    if sf_pct:
        sf_val = excl_total * sf_pct / 100
    elif sf_row:
        sf_val = float(sf_row.get("sale_price", 0) or 0)
        if is_mixed:
            sf_val *= rate
    else:
        sf_val = 0.0

    sf_vat = float((sf_row or {}).get("vat_rate", 20) or 20) if sf_row else 20.0
    vat_total += sf_val * sf_vat / 100
    excl_total += sf_val

    incl_total = excl_total + vat_total
    return {"excl": round(excl_total, 2), "incl": round(incl_total, 2), "currency": currency}


def _write_summary_sheet(wb, entries: list, vat_mode: str, sheet_titles: list[str]) -> None:
    """
    wb'ye 'Özet' sheet'i ekler (ilk konuma).
    entries: [{"budget": ..., "request": ..., "venue": ..., "sheet_title": str}, ...]
    """
    ws = wb.create_sheet(title="Özet", index=0)

    # Sütun genişlikleri
    ws.column_dimensions["A"].width = 30   # Otel Adı
    ws.column_dimensions["B"].width = 18   # KDV Hariç
    ws.column_dimensions["C"].width = 18   # KDV Dahil
    ws.column_dimensions["D"].width = 28   # Web Sitesi
    ws.column_dimensions["E"].width = 40   # Notlar

    # ── Başlık satırı ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    c = ws.cell(row=1, column=1, value="OTEL / MEKAN KARŞILAŞTIRMA ÖZETİ")
    c.font      = _font(bold=True, size=13, white=True)
    c.fill      = _fill(C_HEADER_DARK)
    c.alignment = _align(h="center")
    c.border    = _border()
    ws.row_dimensions[1].height = 30

    # ── Sütun başlıkları ───────────────────────────────────────────────────────
    headers = ["Otel / Mekan Adı", "KDV Hariç Teklif", "KDV Dahil Teklif", "Web Sitesi", "Notlar"]
    for ci, hdr in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=hdr)
        c.font      = _font(bold=True, size=9, white=True)
        c.fill      = _fill(C_HEADER_MED)
        c.alignment = _align(h="center")
        c.border    = _border()
    ws.row_dimensions[2].height = 20

    # ── Veri satırları ─────────────────────────────────────────────────────────
    for row_i, entry in enumerate(entries, 3):
        b        = entry["budget"]
        venue    = entry.get("venue")
        stitle   = entry.get("sheet_title", b.venue_name or "")
        totals   = _calc_totals(b, vat_mode)
        currency = totals["currency"]
        ci_info  = CURRENCY_INFO.get(currency, CURRENCY_INFO["TRY"])
        sym      = ci_info["symbol"]
        num_fmt  = ci_info["fmt"]

        website  = (getattr(venue, "website", None) or "").strip()
        notes    = (b.manager_notes or "").strip()

        fill_bg  = C_WHITE if row_i % 2 == 1 else C_ROW_ALT

        # A: Otel adı — tıklanabilir (sheet'e hyperlink)
        name_cell = ws.cell(row=row_i, column=1, value=b.venue_name or "—")
        name_cell.font      = _font(bold=True, size=9, color=C_HEADER_DARK)
        name_cell.fill      = _fill(fill_bg)
        name_cell.border    = _border()
        name_cell.alignment = _align(h="left", indent=1)
        if stitle:
            try:
                name_cell.hyperlink = f"#{stitle}!A1"
                name_cell.style = "Hyperlink"
                name_cell.font = _font(bold=True, size=9, color="1D4ED8")
            except Exception:
                pass

        # B: KDV Hariç
        ec = ws.cell(row=row_i, column=2, value=totals["excl"])
        ec.font          = _font(size=9)
        ec.fill          = _fill(fill_bg)
        ec.border        = _border()
        ec.number_format = num_fmt
        ec.alignment     = _align(h="right")

        # C: KDV Dahil
        ic = ws.cell(row=row_i, column=3, value=totals["incl"])
        ic.font          = _font(bold=True, size=9)
        ic.fill          = _fill(fill_bg)
        ic.border        = _border()
        ic.number_format = num_fmt
        ic.alignment     = _align(h="right")

        # D: Web Sitesi
        wc = ws.cell(row=row_i, column=4, value=website or "—")
        wc.font      = _font(size=9, color="1D4ED8" if website else C_TEXT)
        wc.fill      = _fill(fill_bg)
        wc.border    = _border()
        wc.alignment = _align(h="left", indent=1)
        if website:
            try:
                wc.hyperlink  = website if website.startswith("http") else f"https://{website}"
                wc.style = "Hyperlink"
            except Exception:
                pass

        # E: Notlar
        nc = ws.cell(row=row_i, column=5, value=notes or "—")
        nc.font      = _font(size=9)
        nc.fill      = _fill(fill_bg)
        nc.border    = _border()
        nc.alignment = _align(h="left", indent=1, wrap=True)

        ws.row_dimensions[row_i].height = 18

    ws.freeze_panes = "A3"


# ── Çoklu bütçe → tek Excel, her bütçe ayrı sheet ────────────────────────────
def build_multi_sheet(
    budgets: list,
    vat_mode: str = "exclusive",
    custom_sections: list | None = None,
) -> io.BytesIO:
    """
    Birden fazla bütçeyi tek Excel dosyasında birleştirir.
    - 2+ bütçe varsa ilk sheet 'Özet' olur (otel adı, KDV hariç/dahil, web, notlar)
    - Her bütçe ayrı bir sheet olur, sheet adı venue_name'den gelir.

    Args:
        budgets: list of dicts:
            [{"budget": ..., "request": ..., "customer": ..., "creator": ..., "venue": ...}, ...]
        vat_mode: 'exclusive' | 'inclusive'
        custom_sections: Admin özel kategorileri

    Returns:
        BytesIO — xlsx dosyası
    """
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl kurulu değil. pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # boş default sheet'i kaldır

    used_titles: list[str] = []
    summary_entries: list[dict] = []

    for i, entry in enumerate(budgets, 1):
        b   = entry["budget"]
        req = entry.get("request")
        cus = entry.get("customer")
        cre = entry.get("creator")

        # Sheet adı: venue_name, max 31 karakter (Excel sınırı), benzersiz
        raw_title = (b.venue_name or f"Bütçe {i}")[:28].strip()
        title = raw_title
        suffix = 2
        while title in used_titles:
            title = f"{raw_title[:25]} {suffix}"
            suffix += 1
        used_titles.append(title)

        ws = wb.create_sheet(title=title)
        _write_sheet(ws, b, req, cus, cre, vat_mode, custom_sections)

        summary_entries.append({
            "budget":      b,
            "request":     req,
            "venue":       entry.get("venue"),
            "sheet_title": title,
        })

    # Özet sayfası: sadece 2+ bütçe varsa ekle
    if len(summary_entries) >= 2:
        _write_summary_sheet(wb, summary_entries, vat_mode, used_titles)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
