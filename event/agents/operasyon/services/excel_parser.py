"""
Claude API ile Excel Katılımcı Listesi Parser

Müşterinin gönderdiği (her seferinde farklı formatlı) Excel dosyasını okur,
Claude ile analiz eder ve standart katılımcı formatına dönüştürür.
"""

import json
import os
import time
import openpyxl
from io import BytesIO
from pathlib import Path
from dotenv import load_dotenv
import anthropic

# .env dosyasını yükle (varsa)
load_dotenv(Path(__file__).parent.parent / ".env")

client = anthropic.Anthropic()  # ANTHROPIC_API_KEY env'den okunur

SYSTEM_PROMPT = """Sen bir etkinlik yönetim şirketinin veri analisti asistanısın.
Sana bir Excel dosyasının içeriği verilecek. Bu dosya bir müşterinin gönderdiği
katılımcı listesidir. Formatı standart değil — her müşteri kendi formatını kullanıyor.

Görevin: İçeriği analiz edip her katılımcı için aşağıdaki standart JSON yapısını üretmek.

ÇIKTI FORMATI — SADECE JSON ARRAY, BAŞKA HİÇBİR ŞEY YAZMA:
[
  {
    "first_name": "string (zorunlu)",
    "last_name": "string (zorunlu)",
    "company": "string veya null",
    "title": "string veya null",
    "email": "string veya null",
    "phone": "string veya null",
    "badge_name": "string veya null (yaka kartı adı, yoksa null)",
    "dietary": "string veya null (beslenme kısıtı)",
    "special_needs": "string veya null",
    "notes": "string veya null",
    "flight_in": {
      "flight_number": "string veya null (örn: TK 2341)",
      "airline": "string veya null",
      "departure_airport": "string veya null (IATA kodu tercihen)",
      "arrival_airport": "string veya null",
      "flight_date": "YYYY-MM-DD veya null",
      "departure_time": "HH:MM veya null",
      "arrival_time": "HH:MM veya null",
      "seat": "string veya null",
      "pnr": "string veya null"
    },
    "flight_out": {
      "flight_number": "string veya null",
      "airline": "string veya null",
      "departure_airport": "string veya null",
      "arrival_airport": "string veya null",
      "flight_date": "YYYY-MM-DD veya null",
      "departure_time": "HH:MM veya null",
      "arrival_time": "HH:MM veya null",
      "seat": "string veya null",
      "pnr": "string veya null"
    },
    "accommodation": {
      "hotel": "string veya null",
      "room_number": "string veya null",
      "room_type": "string veya null (SGL/DBL/SUT vb.)",
      "check_in": "YYYY-MM-DD veya null",
      "check_out": "YYYY-MM-DD veya null",
      "notes": "string veya null"
    }
  }
]

KESİN KURALLAR:
- İlk karakter [ olmalı, son karakter ] olmalı. Açıklama, yorum, başlık YAZMA.
- flight_in, flight_out, accommodation: bilgi yoksa null yaz (boş obje {} değil)
- Tarih formatı MUTLAKA YYYY-MM-DD
- Saat formatı MUTLAKA HH:MM
- Bilinmeyen/boş alanlar için null kullan, boş string "" kullanma
- İsim ayrıştırma: "Ad Soyad" formatındaysa first_name/last_name'e böl
- Yanıtın tamamı valid JSON olmalı — markdown, açıklama, kod bloğu EKLEME
"""

RETRY_PROMPT = """Önceki yanıtın geçerli JSON içermiyor veya beklenmedik formatta.
Lütfen SADECE ve SADECE aşağıdaki formatta bir JSON array döndür — başka hiçbir metin olmadan:
[{"first_name": "...", "last_name": "...", ...}, ...]
İlk karakter [ son karakter ] olmalı. Hiçbir açıklama ekleme."""


def _read_excel_as_text(content: bytes, filename: str) -> str:
    """Excel dosyasını okunabilir metin formatına çevirir."""
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        sheets_text = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                # Tamamen boş satırları atla
                if not any(cell is not None and str(cell).strip() for cell in row):
                    continue
                cells = [str(cell) if cell is not None else "" for cell in row]
                rows.append(" | ".join(cells))

            if rows:
                sheets_text.append(f"=== Sayfa: {sheet_name} ===\n" + "\n".join(rows))

        return "\n\n".join(sheets_text)
    except Exception as e:
        raise ValueError(f"Excel dosyası okunamadı: {e}")


def _call_claude(messages: list[dict]) -> str:
    """Claude API'yi çağırır, overload durumunda 3 kez yeniden dener. Ham metin döner."""
    last_error = None
    for attempt in range(3):
        try:
            message = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=8096,
                system=SYSTEM_PROMPT,
                messages=messages,
            )
            return message.content[0].text.strip()
        except anthropic.APIStatusError as e:
            last_error = e
            if e.status_code in (529, 503):
                time.sleep(5 * (attempt + 1))
                continue
            raise ValueError(f"Claude API hatası ({e.status_code}): {e.message}")
        except anthropic.AuthenticationError:
            raise ValueError(
                "ANTHROPIC_API_KEY geçersiz veya eksik. "
                "Railway ortam değişkenlerini kontrol edin."
            )
    raise ValueError(
        f"Claude API şu an yoğun (overloaded). Birkaç dakika sonra tekrar deneyin. "
        f"Hata: {last_error}"
    )


async def parse_participant_excel(content: bytes, filename: str) -> list[dict]:
    """
    Excel içeriğini Claude ile analiz eder, standart katılımcı listesi döner.
    JSON parse başarısız olursa Claude'a bir kez daha sorar.
    """
    excel_text = _read_excel_as_text(content, filename)

    if not excel_text.strip():
        raise ValueError("Excel dosyası boş veya okunamadı.")

    # Çok büyükse kırp (Claude'un bağlam sınırı için)
    MAX_CHARS = 60_000
    if len(excel_text) > MAX_CHARS:
        excel_text = excel_text[:MAX_CHARS] + "\n\n[... dosya kırpıldı, ilk kısım işleniyor ...]"

    user_message = f"Dosya adı: {filename}\n\nİçerik:\n{excel_text}"
    messages = [{"role": "user", "content": user_message}]

    # İlk deneme
    response_text = _call_claude(messages)

    try:
        return _extract_json_array(response_text)
    except ValueError:
        pass  # JSON parse başarısız — tekrar sor

    # İkinci deneme: Claude'a düzeltmesini söyle
    messages += [
        {"role": "assistant", "content": response_text},
        {"role": "user", "content": RETRY_PROMPT},
    ]
    response_text2 = _call_claude(messages)

    try:
        return _extract_json_array(response_text2)
    except ValueError:
        # Son çare: boş liste değil, hata fırlat
        raise ValueError(
            "Claude geçerli bir JSON array döndürmedi. "
            "Lütfen tekrar deneyin veya farklı bir Excel formatı kullanın."
        )


def _extract_json_array(text: str) -> list[dict]:
    """
    Claude'un çıktısından JSON array'i çıkarır.
    Farklı formatları dener: düz JSON, ```json blok, metin içine gömülü [...].
    """
    import re

    # BOM ve görünmez unicode boşluklarını temizle
    text = text.lstrip("\ufeff\u200b\u200c\u200d\u00a0").strip()

    # 1. Düz JSON array
    if text.startswith("["):
        try:
            result = json.loads(text)
            if isinstance(result, list):
                return result
        except json.JSONDecodeError:
            pass

    # 2. ```json ... ``` veya ``` ... ``` bloğu
    code_block = re.search(r"```(?:json)?\s*(\[[\s\S]*?\])\s*```", text)
    if code_block:
        try:
            result = json.loads(code_block.group(1))
            if isinstance(result, list):
                return result
        except json.JSONDecodeError:
            pass

    # 3. Metin içinde ilk [ ... ] bloğunu bul (iç içe bracket sayarak)
    start = text.find("[")
    if start != -1:
        depth = 0
        in_str = False
        escape = False
        for i, ch in enumerate(text[start:], start):
            if escape:
                escape = False
                continue
            if ch == "\\" and in_str:
                escape = True
                continue
            if ch == '"':
                in_str = not in_str
                continue
            if in_str:
                continue
            if ch == "[":
                depth += 1
            elif ch == "]":
                depth -= 1
                if depth == 0:
                    candidate = text[start:i + 1]
                    try:
                        result = json.loads(candidate)
                        if isinstance(result, list):
                            return result
                    except json.JSONDecodeError:
                        break

    # 4. Son çare: tüm metni JSON olarak parse etmeyi dene
    try:
        result = json.loads(text)
        if isinstance(result, list):
            return result
    except json.JSONDecodeError:
        pass

    raise ValueError(
        "Claude geçerli bir JSON array döndürmedi. "
        "Lütfen tekrar deneyin."
    )
