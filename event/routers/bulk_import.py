"""
Satın Alma — Toplu İçe Aktarma (Bulk Import) router'ı
Admin only.

GET  /bulk-import                     → Upload sayfası
GET  /bulk-import/template/{type}     → Örnek Excel şablonu indir
POST /bulk-import/preview             → Dosyayı parse et, önizleme JSON döndür
POST /bulk-import/import              → Onaylanan satırları kaydet
"""

from __future__ import annotations

import io
import json
import re

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from sqlalchemy.orm import Session

from auth import require_admin
from database import get_db
from models import Customer, Vendor, User, _uuid, _now, SUPPLIER_TYPES

router = APIRouter(prefix="/bulk-import", tags=["bulk_import"])
from templates_config import templates

# ── Openpyxl (zaten requirements'ta mevcut) ──────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


# ── Sütun tanımları ──────────────────────────────────────────────────────────

CUSTOMER_COLS = [
    ("Firma Adı *",        "name",         True),
    ("Kod (3 harf) *",     "code",         True),
    ("Sektör",             "sector",       False),
    ("Adres",              "address",      False),
    ("Vergi Dairesi",      "tax_office",   False),
    ("Vergi No",           "tax_no",   False),
    ("E-posta",            "email",        False),
    ("Telefon",            "phone",        False),
    ("Notlar",             "notes",        False),
    ("Yetkili Adı",        "c_name",       False),
    ("Yetkili Ünvanı",     "c_title",      False),
    ("Yetkili E-posta",    "c_email",      False),
    ("Yetkili Telefon",    "c_phone",      False),
    ("Ödeme Vadesi",       "payment_term", False),
]

VENUE_COLS = [
    ("Firma Adı *",             "name",          True),
    ("Şehir *",                 "city",          True),
    ("Diğer Şehirler",          "cities_extra",  False),
    ("Tedarikçi Tipi *",        "supplier_type", True),
    ("Adres",                   "address",       False),
    ("Yıldız (1-5)",            "stars",         False),
    ("Toplam Oda",              "total_rooms",   False),
    ("Website",                 "website",       False),
    ("Ödeme Vadesi",            "payment_term",  False),
    ("Notlar",                  "notes",         False),
    ("Yetkili Adı",             "c_name",        False),
    ("Yetkili Ünvanı",          "c_title",       False),
    ("Yetkili E-posta",         "c_email",       False),
    ("Yetkili Telefon",         "c_phone",       False),
]

VALID_STYPES = {s["value"] for s in SUPPLIER_TYPES}


# ── Yardımcı: slug benzeri kod üret ─────────────────────────────────────────

def _auto_code(name: str, existing_codes: set) -> str:
    base = re.sub(r"[^a-zA-Z]", "", name.lower())[:3].ljust(3, "x")
    code = base
    i = 1
    while code in existing_codes:
        suffix = str(i)
        code = base[: 3 - len(suffix)] + suffix
        i += 1
    return code


# ── Yardımcı: Excel oluştur ──────────────────────────────────────────────────

def _make_template_wb(cols: list, example_rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Veri"

    header_fill = PatternFill("solid", fgColor="1E293B")
    req_fill    = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for ci, (label, _, required) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font   = header_font
        cell.fill   = req_fill if required else header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = max(18, len(label) + 4)

    ws.row_dimensions[1].height = 30

    ex_fill = PatternFill("solid", fgColor="F8FAFC")
    for ri, row_vals in enumerate(example_rows, 2):
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = ex_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Parse: Excel → list[dict] ────────────────────────────────────────────────

def _parse_excel(file_bytes: bytes, cols: list) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    field_names = [c[1] for c in cols]
    rows = []
    for r_idx in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=r_idx, column=ci).value for ci in range(1, len(cols) + 1)]
        if all(v is None or str(v).strip() == "" for v in row_vals):
            continue
        row_dict = {}
        for fi, fname in enumerate(field_names):
            raw = row_vals[fi]
            row_dict[fname] = str(raw).strip() if raw is not None else ""
        rows.append(row_dict)
    return rows


# ── Validate & enrich ─────────────────────────────────────────────────────────

def _validate_customers(rows: list[dict], db: Session) -> list[dict]:
    existing_codes = {c.code for c in db.query(Customer.code).all()}
    existing_names = {c.name.lower() for c in db.query(Customer.name).all()}
    out = []
    seen_codes: set[str] = set()
    for i, r in enumerate(rows):
        errors = []
        if not r.get("name"):
            errors.append("Firma Adı boş")
        if r.get("name", "").lower() in existing_names:
            errors.append(f"'{r['name']}' zaten kayıtlı")

        code = r.get("code", "").lower().strip()
        if not code:
            code = _auto_code(r.get("name", "row"), existing_codes | seen_codes)
            r["code"] = code
            r["_code_auto"] = True
        else:
            if len(code) > 10:
                errors.append("Kod en fazla 10 karakter")
            if code in existing_codes or code in seen_codes:
                errors.append(f"Kod '{code}' zaten kullanılıyor")

        seen_codes.add(code)
        out.append({**r, "_row": i + 2, "_errors": errors, "_valid": len(errors) == 0})
    return out


def _validate_venues(rows: list[dict], db: Session) -> list[dict]:
    # Mevcut isimler — sadece bilgi amaçlı (hata değil, güncelleme yapılır)
    existing_names = {v.name.lower() for v in db.query(Vendor.name).all()}
    out = []
    for i, r in enumerate(rows):
        errors = []
        if not r.get("name"):
            errors.append("Firma Adı boş")
        if not r.get("city"):
            errors.append("Şehir boş")
        stype = r.get("supplier_type", "").strip().lower()
        if not stype:
            errors.append("Tedarikçi Tipi boş")
        elif stype not in VALID_STYPES:
            errors.append(f"Geçersiz tip: {stype} (otel/etkinlik/teknik/...)")
        r["supplier_type"] = stype
        if not r.get("c_name"):
            r["c_name"] = "Yetkili"
        if not r.get("c_email"):
            r["c_email"] = "eposta yok"
        # Mevcut kayıt varsa güncelleme yapılacağını işaretle (uyarı, hata değil)
        r["_will_update"] = r.get("name", "").lower() in existing_names
        out.append({**r, "_row": i + 2, "_errors": errors, "_valid": len(errors) == 0})
    return out


# ── DB kayıt ─────────────────────────────────────────────────────────────────

def _save_customers(rows: list[dict], db: Session) -> tuple[int, int]:
    saved = skipped = 0
    for r in rows:
        if not r.get("_valid"):
            skipped += 1
            continue
        contact = {}
        if r.get("c_name"):
            contact = {
                "name":  r["c_name"],
                "title": r.get("c_title", ""),
                "email": r.get("c_email", ""),
                "phone": r.get("c_phone", ""),
            }
        db.add(Customer(
            id           = _uuid(),
            name         = r["name"],
            code         = r["code"].lower(),
            sector       = r.get("sector", ""),
            address      = r.get("address", ""),
            tax_office   = r.get("tax_office", ""),
            tax_number   = r.get("tax_no", ""),
            email        = r.get("email", ""),
            phone        = r.get("phone", ""),
            notes        = r.get("notes", ""),
            payment_term = r.get("payment_term", ""),
            contacts_json = json.dumps([contact] if contact else []),
            created_at   = _now(),
        ))
        saved += 1
    db.commit()
    return saved, skipped


def _save_venues(rows: list[dict], db: Session) -> tuple[int, int]:
    saved = skipped = 0
    for r in rows:
        if not r.get("_valid"):
            skipped += 1
            continue
        cities_extra = [c.strip() for c in r.get("cities_extra", "").split(",") if c.strip()]
        all_cities = list({r["city"]} | set(cities_extra))
        contact = {
            "name":  r.get("c_name", ""),
            "title": r.get("c_title", ""),
            "email": r.get("c_email", ""),
            "phone": r.get("c_phone", ""),
        }
        stars = None
        try:
            stars = int(float(r.get("stars") or 0)) or None
        except (ValueError, TypeError):
            pass
        total_rooms = 0
        try:
            total_rooms = int(float(r.get("total_rooms") or 0))
        except (ValueError, TypeError):
            pass

        # Upsert: aynı isimli kayıt varsa (aktif/pasif) güncelle, yoksa ekle
        existing = db.query(Vendor).filter(Vendor.name == r["name"]).first()
        if existing:
            existing.city          = r["city"]
            existing.cities_json   = json.dumps(all_cities)
            existing.supplier_type = r["supplier_type"]
            existing.address       = r.get("address", "")
            existing.stars         = stars
            existing.total_rooms   = total_rooms
            existing.website       = r.get("website", "")
            existing.notes         = r.get("notes", "")
            _pt_val = r.get("payment_term", "")
            existing.payment_term  = int(_pt_val) if str(_pt_val).strip().isdigit() else 30
            existing.contacts_json = json.dumps([contact])
            existing.halls_json    = "[]"
            existing.active        = True
        else:
            db.add(Vendor(
                id            = _uuid(),
                name          = r["name"],
                city          = r["city"],
                cities_json   = json.dumps(all_cities),
                supplier_type = r["supplier_type"],
                address       = r.get("address", ""),
                stars         = stars,
                total_rooms   = total_rooms,
                website       = r.get("website", ""),
                notes         = r.get("notes", ""),
                payment_term  = int(r["payment_term"]) if str(r.get("payment_term","")).strip().isdigit() else 30,
                contacts_json = json.dumps([contact]),
                halls_json    = "[]",
                active        = True,
                created_at    = _now(),
            ))
        saved += 1
    db.commit()
    return saved, skipped


# ── Endpoint'ler ──────────────────────────────────────────────────────────────

@router.get("", response_class=HTMLResponse, name="bulk_import_index")
async def bulk_import_index(
    request: Request,
    current_user: User = Depends(require_admin),
):
    return templates.TemplateResponse(
        "bulk_import/index.html",
        {
            "request":      request,
            "current_user": current_user,
            "page_title":   "Toplu İçe Aktarma",
            "supplier_types": SUPPLIER_TYPES,
        },
    )


@router.get("/template/{import_type}")
async def bulk_import_template(
    import_type: str,
    current_user: User = Depends(require_admin),
):
    """Örnek Excel şablonu indir."""
    if import_type == "customers":
        example = [
            ["Örnek Firma A.Ş.", "orf", "İlaç", "İstanbul", "Kadıköy VD", "1234567890",
             "info@ornek.com", "+90 212 000 0000", "", "Ahmet Yılmaz", "Satın Alma Müdürü",
             "ahmet@ornek.com", "+90 532 000 0000", "30 gün"],
        ]
        content = _make_template_wb(CUSTOMER_COLS, example)
        filename = "musteri_sablonu.xlsx"
    elif import_type == "venues":
        example = [
            ["Örnek Otel", "İstanbul", "Ankara, İzmir", "otel", "Örnek Mah. No:1",
             "5", "200", "https://ornekotel.com", "30 gün", "",
             "Mehmet Demir", "Satış Müdürü", "mehmet@ornekotel.com", "+90 532 111 2233"],
        ]
        content = _make_template_wb(VENUE_COLS, example)
        filename = "tedarikci_sablonu.xlsx"
    else:
        return JSONResponse({"error": "Geçersiz tip"}, status_code=400)

    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/preview")
async def bulk_import_preview(
    import_type: str = Form(...),
    file: UploadFile = File(...),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Excel dosyasını parse edip önizleme JSON'u döndür."""
    file_bytes = await file.read()
    try:
        if import_type == "customers":
            raw_rows = _parse_excel(file_bytes, CUSTOMER_COLS)
            rows = _validate_customers(raw_rows, db)
            cols = CUSTOMER_COLS
        elif import_type == "venues":
            raw_rows = _parse_excel(file_bytes, VENUE_COLS)
            rows = _validate_venues(raw_rows, db)
            cols = VENUE_COLS
        else:
            return JSONResponse({"error": "Geçersiz tip"}, status_code=400)
    except Exception as e:
        return JSONResponse({"error": f"Dosya okunamadı: {e}"}, status_code=400)

    valid_count   = sum(1 for r in rows if r["_valid"])
    invalid_count = len(rows) - valid_count

    return JSONResponse({
        "rows":          rows,
        "col_labels":    [c[0] for c in cols],
        "field_names":   [c[1] for c in cols],
        "valid_count":   valid_count,
        "invalid_count": invalid_count,
        "total":         len(rows),
    })


@router.post("/import")
async def bulk_import_do(
    import_type: str = Form(...),
    rows_json:   str = Form(...),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Önizleme sonrası onaylanan satırları kaydet."""
    try:
        rows = json.loads(rows_json)
    except Exception:
        return JSONResponse({"error": "Geçersiz veri"}, status_code=400)

    if import_type == "customers":
        saved, skipped = _save_customers(rows, db)
    elif import_type == "venues":
        saved, skipped = _save_venues(rows, db)
    else:
        return JSONResponse({"error": "Geçersiz tip"}, status_code=400)

    return JSONResponse({"saved": saved, "skipped": skipped})
