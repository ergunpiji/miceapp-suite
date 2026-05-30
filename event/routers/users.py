"""
E-dem — Kullanıcı yönetimi router'ı (Admin only)
GET    /users                    → Kullanıcı listesi
GET    /users/new                → Yeni kullanıcı formu
POST   /users/new                → Yeni kullanıcı oluştur
GET    /users/{id}/edit          → Düzenleme formu
POST   /users/{id}/edit          → Kullanıcı güncelle
POST   /users/{id}/delete        → Kullanıcı sil (soft delete)
GET    /users/org-titles         → Organizasyon unvanları yönetimi
POST   /users/org-titles/{id}    → Unvan bütçe limiti güncelle
"""

import io

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile, status
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session

from auth import hash_password, require_admin
from database import get_db
from models import OrgTitle, Team, User, USER_ROLES, _uuid, _now

router = APIRouter(prefix="/users", tags=["users"])
from templates_config import templates


def _get_org_titles(db: Session) -> list:
    return db.query(OrgTitle).order_by(OrgTitle.sort_order).all()


def _get_teams(db: Session) -> list:
    return db.query(Team).filter(Team.active == True).order_by(Team.name).all()


def _get_pm_users(db: Session, exclude_id: str | None = None) -> list:
    """Yönetici olabilecek tüm aktif kullanıcılar (admin + PM tarafı + muhasebe müdürü)."""
    q = db.query(User).filter(
        User.active == True,
        User.role.in_(["admin", "mudur", "yonetici", "asistan", "muhasebe_muduru"]),
    )
    if exclude_id:
        q = q.filter(User.id != exclude_id)
    return q.order_by(User.name).all()


@router.get("", response_class=HTMLResponse, name="users_list")
async def users_list(
    request: Request,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
    welcome_for: str = "",
):
    users = db.query(User).order_by(User.created_at.desc()).all()
    welcome_user = db.query(User).filter(User.id == welcome_for).first() if welcome_for else None
    return templates.TemplateResponse(
        "users/list.html",
        {
            "request":      request,
            "current_user": current_user,
            "users":        users,
            "page_title":   "Kullanıcı Yönetimi",
            "user_roles":   USER_ROLES,
            "welcome_user": welcome_user,
        },
    )


@router.get("/org-titles", response_class=HTMLResponse, name="users_org_titles")
async def org_titles_page(
    request: Request,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
    saved: str = "",
):
    titles = _get_org_titles(db)
    # Tüm kullanıcı sayısını unvan başına hesapla
    user_counts = {}
    for u in db.query(User).filter(User.active == True).all():
        if u.org_title_id:
            user_counts[u.org_title_id] = user_counts.get(u.org_title_id, 0) + 1

    return templates.TemplateResponse(
        "users/org_titles.html",
        {
            "request":      request,
            "current_user": current_user,
            "titles":       titles,
            "user_counts":  user_counts,
            "page_title":   "Organizasyon Yapısı",
            "saved":        saved == "1",
        },
    )


@router.post("/org-titles/{title_id}", name="users_org_title_update")
async def org_title_update(
    title_id: str,
    budget_limit: str = Form(""),
    pm_permission_level: str = Form(""),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    title = db.query(OrgTitle).filter(OrgTitle.id == title_id).first()
    if title:
        val = budget_limit.strip().replace(".", "").replace(",", "")
        title.budget_limit = float(val) if val else None
        lvl = pm_permission_level.strip()
        title.pm_permission_level = lvl if lvl in ("mudur", "yonetici", "asistan") else None
        db.commit()
    return RedirectResponse(url="/users/org-titles?saved=1", status_code=status.HTTP_302_FOUND)


@router.get("/new", response_class=HTMLResponse, name="users_new")
async def users_new(
    request: Request,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    return templates.TemplateResponse(
        "users/form.html",
        {
            "request":      request,
            "current_user": current_user,
            "user":         None,
            "page_title":   "Yeni Kullanıcı",
            "user_roles":   USER_ROLES,
            "org_titles":   _get_org_titles(db),
            "teams":        _get_teams(db),
            "pm_users":     _get_pm_users(db),
            "error":        None,
        },
    )


@router.post("/new", name="users_create")
async def users_create(
    request: Request,
    email:         str = Form(...),
    password:      str = Form(...),
    role:          str = Form(...),
    name:          str = Form(...),
    surname:       str = Form(...),
    title:         str = Form(""),
    phone:         str = Form(""),
    org_title_id:  str = Form(""),
    team_id:       str = Form(""),
    manager_id:    str = Form(""),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    existing = db.query(User).filter(User.email == email.lower().strip()).first()
    if existing:
        return templates.TemplateResponse(
            "users/form.html",
            {
                "request":      request,
                "current_user": current_user,
                "user":         None,
                "page_title":   "Yeni Kullanıcı",
                "user_roles":   USER_ROLES,
                "org_titles":   _get_org_titles(db),
                "teams":        _get_teams(db),
                "pm_users":     _get_pm_users(db),
                "error":        "Bu e-posta adresi zaten kayıtlı.",
                "form_data":    {"email": email, "role": role, "name": name, "surname": surname,
                                 "title": title, "phone": phone, "org_title_id": org_title_id},
            },
            status_code=400,
        )

    tid = team_id.strip() or None
    user = User(
        id=_uuid(),
        email=email.lower().strip(),
        password_hash=hash_password(password),
        role=role,
        name=name.strip(),
        surname=surname.strip(),
        title=title.strip(),
        phone=phone.strip(),
        org_title_id=org_title_id.strip() or None,
        team_id=tid,
        manager_id=manager_id.strip() or None,
        active=True,
        created_at=_now(),
    )
    db.add(user)
    db.commit()
    return RedirectResponse(
        url=f"/users?welcome_for={user.id}",
        status_code=status.HTTP_302_FOUND,
    )


@router.get("/{user_id}/edit", response_class=HTMLResponse, name="users_edit")
async def users_edit(
    user_id: str,
    request: Request,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        return RedirectResponse(url="/users", status_code=status.HTTP_302_FOUND)
    return templates.TemplateResponse(
        "users/form.html",
        {
            "request":      request,
            "current_user": current_user,
            "user":         user,
            "page_title":   f"{user.full_name} — Düzenle",
            "user_roles":   USER_ROLES,
            "org_titles":   _get_org_titles(db),
            "teams":        _get_teams(db),
            "pm_users":     _get_pm_users(db, exclude_id=user_id),
            "error":        None,
        },
    )


@router.post("/{user_id}/edit", name="users_update")
async def users_update(
    user_id:  str,
    request: Request,
    email:         str = Form(...),
    role:          str = Form(...),
    name:          str = Form(...),
    surname:       str = Form(...),
    title:         str = Form(""),
    phone:         str = Form(""),
    password:      str = Form(""),
    active:        str = Form("on"),
    org_title_id:  str = Form(""),
    team_id:       str = Form(""),
    manager_id:    str = Form(""),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        return RedirectResponse(url="/users", status_code=status.HTTP_302_FOUND)

    conflict = db.query(User).filter(
        User.email == email.lower().strip(),
        User.id != user_id,
    ).first()
    if conflict:
        return templates.TemplateResponse(
            "users/form.html",
            {
                "request":      request,
                "current_user": current_user,
                "user":         user,
                "page_title":   f"{user.full_name} — Düzenle",
                "user_roles":   USER_ROLES,
                "org_titles":   _get_org_titles(db),
                "teams":        _get_teams(db),
                "pm_users":     _get_pm_users(db, exclude_id=user_id),
                "error":        "Bu e-posta adresi başka bir kullanıcıya ait.",
            },
            status_code=400,
        )

    user.email        = email.lower().strip()
    user.role         = role
    user.name         = name.strip()
    user.surname      = surname.strip()
    user.title        = title.strip()
    user.phone        = phone.strip()
    user.active       = (active == "on")
    user.org_title_id = org_title_id.strip() or None
    user.team_id      = team_id.strip() or None
    user.manager_id   = manager_id.strip() or None

    if password.strip():
        user.password_hash = hash_password(password.strip())

    db.commit()
    return RedirectResponse(url="/users", status_code=status.HTTP_302_FOUND)


# ---------------------------------------------------------------------------
# Excel şablon indir + toplu içe aktar
# ---------------------------------------------------------------------------

@router.get("/template", name="users_template")
async def users_template(current_user: User = Depends(require_admin)):
    """Kullanıcı listesi Excel şablonu indir."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kullanıcılar"

    headers = ["Ad", "Soyad", "E-posta", "Şifre", "Rol", "Telefon", "Aktif (E/H)"]
    header_fill = PatternFill("solid", fgColor="1e3a5f")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    widths = [15, 15, 30, 20, 20, 15, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Rol dropdown
    valid_roles = ",".join([r["value"] for r in USER_ROLES])
    dv_role = DataValidation(type="list", formula1=f'"{valid_roles}"', allow_blank=False)
    ws.add_data_validation(dv_role)
    dv_role.sqref = "E2:E1000"

    # Aktif dropdown
    dv_active = DataValidation(type="list", formula1='"E,H"', allow_blank=False)
    ws.add_data_validation(dv_active)
    dv_active.sqref = "G2:G1000"

    # Örnek satır
    ws.cell(row=2, column=1, value="Ahmet")
    ws.cell(row=2, column=2, value="Yılmaz")
    ws.cell(row=2, column=3, value="ahmet@sirket.com")
    ws.cell(row=2, column=4, value="Sifre123!")
    ws.cell(row=2, column=5, value="asistan")
    ws.cell(row=2, column=6, value="0532 000 0000")
    ws.cell(row=2, column=7, value="E")

    # Rol referans sekmesi
    ws2 = wb.create_sheet("Roller")
    ws2.cell(row=1, column=1, value="Rol ID").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Açıklama").font = Font(bold=True)
    for r, role in enumerate(USER_ROLES, 2):
        ws2.cell(row=r, column=1, value=role["value"])
        ws2.cell(row=r, column=2, value=role["label"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=kullanici_sablonu.xlsx"},
    )


@router.post("/import", name="users_import")
async def users_import(
    file: UploadFile = File(...),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Excel'den kullanıcı toplu içe aktar."""
    import openpyxl

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    valid_role_ids = {r["value"] for r in USER_ROLES}
    added = skipped = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        name, surname, email, password, role, phone, active_str = (row + (None,) * 7)[:7]
        name     = str(name or "").strip()
        surname  = str(surname or "").strip()
        email    = str(email or "").strip().lower()
        password = str(password or "").strip()
        role     = str(role or "").strip()
        phone    = str(phone or "").strip()
        active   = str(active_str or "E").strip().upper() != "H"

        if not email:
            errors.append(f"Satır {row_idx}: E-posta boş")
            skipped += 1
            continue
        if not password:
            errors.append(f"Satır {row_idx}: Şifre boş")
            skipped += 1
            continue
        if role not in valid_role_ids:
            errors.append(f"Satır {row_idx}: Geçersiz rol '{role}'")
            skipped += 1
            continue

        # Aynı e-posta varsa atla
        if db.query(User).filter(User.email == email).first():
            skipped += 1
            continue

        db.add(User(
            id=_uuid(), email=email,
            password_hash=hash_password(password),
            name=name, surname=surname,
            role=role, phone=phone,
            active=active, created_at=_now(),
        ))
        added += 1

    db.commit()

    msg = f"{added} kullanıcı eklendi"
    if skipped:
        msg += f", {skipped} atlandı"
    if errors:
        msg += f". Hatalar: " + " | ".join(errors[:5])

    return RedirectResponse(url=f"/users?import_msg={msg}", status_code=status.HTTP_302_FOUND)


@router.post("/{user_id}/delete", name="users_delete")
async def users_delete(
    user_id: str,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    user = db.query(User).filter(User.id == user_id).first()
    if user and user.id != current_user.id:
        user.active = False
        db.commit()
    return RedirectResponse(url="/users", status_code=status.HTTP_302_FOUND)
