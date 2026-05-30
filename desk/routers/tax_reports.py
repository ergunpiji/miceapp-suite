"""
Vergi Raporları — KDV özet, BA/BS, geçici vergi tahmini, yıllık projeksiyon.
Mali mühür gerekmez; webapp içi hesaplamalar. Mali müşavir bu raporları
alır ve kendi sisteminden GİB'e gönderir.
"""
import io
from datetime import date, datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi import APIRouter, Depends, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy import extract
from sqlalchemy.orm import Session

from auth import require_admin, require_mudur, get_company_id
from database import get_db
from models import User, Invoice, GeneralExpense, SystemSetting
from templates_config import templates


router = APIRouter(prefix="/tax-reports", tags=["tax_reports"])


def _require_module_active(db: Session):
    s = db.query(SystemSetting).filter(SystemSetting.key == "module_tax_reports_enabled").first()
    if not (s and s.value == "1"):
        from fastapi import HTTPException
        raise HTTPException(404, "Vergi Raporları modülü aktif değil")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _kdv_summary(db: Session, year: int, month: int, cid: int) -> dict:
    """KDV1 (giden, hesaplanan) + KDV2 (gelen, indirilen) + net devreden."""
    invoices = db.query(Invoice).filter(
        Invoice.company_id == cid,
        Invoice.status.in_(["approved", "partial", "paid"]),
        extract("year", Invoice.invoice_date) == year,
        extract("month", Invoice.invoice_date) == month,
    ).all()

    # Çıkış (kesilen + komisyon) hesaplanan KDV
    out_base = sum(i.amount for i in invoices if i.invoice_type in ("kesilen", "komisyon"))
    out_vat = sum((i.amount * i.vat_rate) for i in invoices if i.invoice_type in ("kesilen", "komisyon"))
    out_iade_base = sum(i.amount for i in invoices if i.invoice_type == "iade_kesilen")
    out_iade_vat = sum((i.amount * i.vat_rate) for i in invoices if i.invoice_type == "iade_kesilen")
    net_out_vat = out_vat - out_iade_vat

    # Giriş (gelen) indirilen KDV
    in_base = sum(i.amount for i in invoices if i.invoice_type == "gelen")
    in_vat = sum((i.amount * i.vat_rate) for i in invoices if i.invoice_type == "gelen")
    in_iade_base = sum(i.amount for i in invoices if i.invoice_type == "iade_gelen")
    in_iade_vat = sum((i.amount * i.vat_rate) for i in invoices if i.invoice_type == "iade_gelen")
    net_in_vat = in_vat - in_iade_vat

    net_payable = net_out_vat - net_in_vat  # >0 → ödenecek, <0 → devreden

    return {
        "year": year, "month": month,
        "out_base": round(out_base, 2),
        "out_vat": round(net_out_vat, 2),
        "in_base": round(in_base, 2),
        "in_vat": round(net_in_vat, 2),
        "net_payable": round(net_payable, 2),
        "iade_kesilen": round(out_iade_base, 2),
        "iade_kesilen_vat": round(out_iade_vat, 2),
        "iade_gelen": round(in_iade_base, 2),
        "iade_gelen_vat": round(in_iade_vat, 2),
    }


def _ba_bs_lines(db: Session, year: int, month: int, cid: int, threshold: float = 5000.0) -> dict:
    """BA (alımlar/giderler) ve BS (satışlar) — eşik üzeri kayıtlar.
    Karşı tarafa göre toplulanır."""
    invoices = db.query(Invoice).filter(
        Invoice.company_id == cid,
        Invoice.status.in_(["approved", "partial", "paid"]),
        extract("year", Invoice.invoice_date) == year,
        extract("month", Invoice.invoice_date) == month,
    ).all()

    # BA: gelen (alımlar) — vergi no bazında topla
    ba_groups = defaultdict(lambda: {"name": "", "tax_no": "", "total": 0.0, "count": 0})
    for inv in invoices:
        if inv.invoice_type != "gelen":
            continue
        v = inv.vendor
        if not v:
            continue
        key = v.tax_no or f"V{v.id}"
        ba_groups[key]["name"] = v.name
        ba_groups[key]["tax_no"] = v.tax_no or ""
        ba_groups[key]["total"] += inv.total_with_vat
        ba_groups[key]["count"] += 1
    ba_rows = [g for g in ba_groups.values() if g["total"] >= threshold]
    ba_rows.sort(key=lambda x: x["total"], reverse=True)

    # BS: kesilen (satışlar) — müşteri vergi no bazında topla (Reference üzerinden)
    bs_groups = defaultdict(lambda: {"name": "", "tax_no": "", "total": 0.0, "count": 0})
    for inv in invoices:
        if inv.invoice_type not in ("kesilen", "komisyon"):
            continue
        c = inv.reference.customer if inv.reference else None
        if not c:
            continue
        key = c.tax_no or f"C{c.id}"
        bs_groups[key]["name"] = c.name
        bs_groups[key]["tax_no"] = c.tax_no or ""
        bs_groups[key]["total"] += inv.total_with_vat
        bs_groups[key]["count"] += 1
    bs_rows = [g for g in bs_groups.values() if g["total"] >= threshold]
    bs_rows.sort(key=lambda x: x["total"], reverse=True)

    return {
        "ba": ba_rows, "bs": bs_rows,
        "threshold": threshold,
        "ba_total": round(sum(r["total"] for r in ba_rows), 2),
        "bs_total": round(sum(r["total"] for r in bs_rows), 2),
    }


def _quarterly_corp_tax(db: Session, year: int, cid: int) -> list:
    """Çeyreklik kâr × kurumlar vergisi tahmini (%25)."""
    rate = 0.25
    invoices = db.query(Invoice).filter(
        Invoice.company_id == cid,
        Invoice.status.in_(["approved", "partial", "paid"]),
        extract("year", Invoice.invoice_date) == year,
    ).all()
    expenses = db.query(GeneralExpense).filter(
        GeneralExpense.company_id == cid,
        extract("year", GeneralExpense.expense_date) == year,
    ).all()

    quarters = []
    for q in range(1, 5):
        months = [(q - 1) * 3 + 1, (q - 1) * 3 + 2, (q - 1) * 3 + 3]
        out_base = sum(i.amount for i in invoices if i.invoice_type in ("kesilen", "komisyon")
                       and i.invoice_date and i.invoice_date.month in months)
        out_iade = sum(i.amount for i in invoices if i.invoice_type == "iade_kesilen"
                       and i.invoice_date and i.invoice_date.month in months)
        in_base = sum(i.amount for i in invoices if i.invoice_type == "gelen"
                      and i.invoice_date and i.invoice_date.month in months)
        in_iade = sum(i.amount for i in invoices if i.invoice_type == "iade_gelen"
                      and i.invoice_date and i.invoice_date.month in months)
        exp = sum(e.amount for e in expenses if e.expense_date and e.expense_date.month in months)
        gross = (out_base - out_iade) - (in_base - in_iade)
        net = gross - exp
        tax = max(0.0, net * rate)
        quarters.append({
            "q": q,
            "income": round(out_base - out_iade, 2),
            "cost": round(in_base - in_iade, 2),
            "expense": round(exp, 2),
            "gross_profit": round(gross, 2),
            "net_profit": round(net, 2),
            "tax_estimate": round(tax, 2),
        })
    return quarters


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@router.get("", response_class=HTMLResponse, name="tax_reports_index")
async def tax_reports_index(
    request: Request,
    year: int = None,
    month: int = None,
    current_user: User = Depends(require_mudur),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    _require_module_active(db)
    today = date.today()
    if not year:
        year = today.year
    if not month:
        month = today.month

    kdv = _kdv_summary(db, year, month, cid)
    babs = _ba_bs_lines(db, year, month, cid)
    quarters = _quarterly_corp_tax(db, year, cid)

    return templates.TemplateResponse(
        "tax_reports/index.html",
        {
            "request": request,
            "current_user": current_user,
            "page_title": "Vergi Raporları",
            "year": year, "month": month,
            "kdv": kdv,
            "babs": babs,
            "quarters": quarters,
            "month_names": ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
                            "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"],
        },
    )


@router.get("/export", name="tax_reports_export")
async def tax_reports_export(
    year: int = None,
    month: int = None,
    current_user: User = Depends(require_mudur),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    _require_module_active(db)
    today = date.today()
    if not year:
        year = today.year
    if not month:
        month = today.month

    kdv = _kdv_summary(db, year, month, cid)
    babs = _ba_bs_lines(db, year, month, cid)
    quarters = _quarterly_corp_tax(db, year, cid)

    NAVY = "1A3A5C"
    NAVY_2 = "1E5F8C"
    GRAY = "F1F5F9"
    money_fmt = '#,##0.00 [$₺-tr-TR]'
    thin = Side(border_style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()

    # Sheet 1 — KDV Özet
    ws = wb.active
    ws.title = "KDV"
    ws.sheet_view.showGridLines = False

    month_names = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
                   "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]

    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    from templates_config import company as _company
    _short = _company('short_name') or _company('name') or 'Prizma Finans'
    c.value = f"{_short} · KDV Özet — {month_names[month-1]} {year}"
    c.font = Font(size=14, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    rows = [
        ["", "Kalem", "Matrah", "KDV", "Açıklama"],
        ["GİDEN", "Hesaplanan KDV (Kesilen+Komisyon)", kdv["out_base"], kdv["out_vat"], "Müşterilerden tahsil edilen"],
        ["GİDEN", "İade Kesilen", kdv["iade_kesilen"], kdv["iade_kesilen_vat"], "Yukarıdan düşülür"],
        ["GELEN", "İndirilen KDV (Gelen)", kdv["in_base"], kdv["in_vat"], "Tedarikçilere ödenen"],
        ["GELEN", "İade Gelen", kdv["iade_gelen"], kdv["iade_gelen_vat"], "Yukarıdan düşülür"],
        ["", "NET KDV", "", kdv["net_payable"], "Pozitif → Ödenecek; Negatif → Devreden"],
    ]
    for r_idx, row in enumerate(rows, start=3):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            if r_idx == 3:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=NAVY_2)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif r_idx == len(rows) + 2:  # NET satırı
                cell.font = Font(bold=True, color=NAVY, size=11)
                cell.fill = PatternFill("solid", fgColor=GRAY)
            if c_idx in (3, 4) and isinstance(val, (int, float)):
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal="right")

    for letter, w in zip("ABCDE", [12, 38, 16, 16, 30]):
        ws.column_dimensions[letter].width = w

    # Sheet 2 — BA Formu
    ws2 = wb.create_sheet("BA (Alımlar)")
    ws2.sheet_view.showGridLines = False
    ws2.row_dimensions[1].height = 28
    ws2.merge_cells("A1:D1")
    c = ws2["A1"]
    c.value = f"BA — Alımlar Listesi (≥ {babs['threshold']:,.0f} TL)"
    c.font = Font(size=13, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ba_headers = ["Tedarikçi", "Vergi No", "Fatura Adedi", "Toplam (KDV dahil)"]
    for i, h in enumerate(ba_headers, 1):
        cell = ws2.cell(row=3, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY_2)
        cell.alignment = Alignment(horizontal="center")
    for r_idx, row in enumerate(babs["ba"], start=4):
        ws2.cell(row=r_idx, column=1, value=row["name"])
        ws2.cell(row=r_idx, column=2, value=row["tax_no"])
        ws2.cell(row=r_idx, column=3, value=row["count"]).alignment = Alignment(horizontal="center")
        cell = ws2.cell(row=r_idx, column=4, value=row["total"])
        cell.number_format = money_fmt
        cell.alignment = Alignment(horizontal="right")
    # Toplam
    total_row = 4 + len(babs["ba"]) + 1
    ws2.cell(row=total_row, column=3, value="TOPLAM").font = Font(bold=True)
    ws2.cell(row=total_row, column=3).alignment = Alignment(horizontal="right")
    cell = ws2.cell(row=total_row, column=4, value=babs["ba_total"])
    cell.font = Font(bold=True, color=NAVY)
    cell.number_format = money_fmt
    cell.fill = PatternFill("solid", fgColor=GRAY)
    for letter, w in zip("ABCD", [38, 16, 14, 22]):
        ws2.column_dimensions[letter].width = w

    # Sheet 3 — BS Formu
    ws3 = wb.create_sheet("BS (Satışlar)")
    ws3.sheet_view.showGridLines = False
    ws3.row_dimensions[1].height = 28
    ws3.merge_cells("A1:D1")
    c = ws3["A1"]
    c.value = f"BS — Satışlar Listesi (≥ {babs['threshold']:,.0f} TL)"
    c.font = Font(size=13, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    for i, h in enumerate(["Müşteri", "Vergi No", "Fatura Adedi", "Toplam (KDV dahil)"], 1):
        cell = ws3.cell(row=3, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY_2)
        cell.alignment = Alignment(horizontal="center")
    for r_idx, row in enumerate(babs["bs"], start=4):
        ws3.cell(row=r_idx, column=1, value=row["name"])
        ws3.cell(row=r_idx, column=2, value=row["tax_no"])
        ws3.cell(row=r_idx, column=3, value=row["count"]).alignment = Alignment(horizontal="center")
        cell = ws3.cell(row=r_idx, column=4, value=row["total"])
        cell.number_format = money_fmt
        cell.alignment = Alignment(horizontal="right")
    total_row = 4 + len(babs["bs"]) + 1
    ws3.cell(row=total_row, column=3, value="TOPLAM").font = Font(bold=True)
    ws3.cell(row=total_row, column=3).alignment = Alignment(horizontal="right")
    cell = ws3.cell(row=total_row, column=4, value=babs["bs_total"])
    cell.font = Font(bold=True, color=NAVY)
    cell.number_format = money_fmt
    cell.fill = PatternFill("solid", fgColor=GRAY)
    for letter, w in zip("ABCD", [38, 16, 14, 22]):
        ws3.column_dimensions[letter].width = w

    # Sheet 4 — Geçici Vergi Tahmini
    ws4 = wb.create_sheet("Geçici Vergi")
    ws4.sheet_view.showGridLines = False
    ws4.row_dimensions[1].height = 28
    ws4.merge_cells("A1:F1")
    c = ws4["A1"]
    c.value = f"{year} — Geçici Vergi Tahmini (Çeyreklik, %25 oranı)"
    c.font = Font(size=13, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    for i, h in enumerate(["Çeyrek", "Net Gelir", "Net Maliyet", "Genel Gider", "Net Kâr", "Geçici Vergi"], 1):
        cell = ws4.cell(row=3, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY_2)
        cell.alignment = Alignment(horizontal="center")
    for r_idx, q in enumerate(quarters, start=4):
        ws4.cell(row=r_idx, column=1, value=f"Q{q['q']}").alignment = Alignment(horizontal="center")
        ws4.cell(row=r_idx, column=2, value=q["income"]).number_format = money_fmt
        ws4.cell(row=r_idx, column=3, value=q["cost"]).number_format = money_fmt
        ws4.cell(row=r_idx, column=4, value=q["expense"]).number_format = money_fmt
        cell = ws4.cell(row=r_idx, column=5, value=q["net_profit"])
        cell.number_format = money_fmt
        cell.font = Font(bold=True, color="16A34A" if q["net_profit"] >= 0 else "DC2626")
        cell = ws4.cell(row=r_idx, column=6, value=q["tax_estimate"])
        cell.number_format = money_fmt
        cell.font = Font(bold=True, color=NAVY)
    for letter, w in zip("ABCDEF", [10, 18, 18, 18, 18, 18]):
        ws4.column_dimensions[letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"vergi-raporu-{year}-{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
