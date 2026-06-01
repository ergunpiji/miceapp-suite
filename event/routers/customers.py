"""
Satın Alma — Müşteri yönetimi router'ı (Admin only, PM autocomplete)
"""

import json
import os
import shutil

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile, status
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse

from storage import save_upload, delete_upload
from sqlalchemy.orm import Session

from auth import get_current_user, require_admin
from database import get_db
from models import Customer, Team, User, _uuid, _now

router = APIRouter(prefix="/customers", tags=["customers"])
from templates_config import templates


@router.get("", response_class=HTMLResponse, name="customers_list")
async def customers_list(
    request: Request,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    query = db.query(Customer)
    # Tenant izolasyonu — sadece kendi şirketinin müşterileri
    if current_user.company_id:
        query = query.filter(Customer.company_id == current_user.company_id)
    if not current_user.is_gm and current_user.role == "mudur" and current_user.team_id:
        _team = db.query(Team).filter(Team.id == current_user.team_id).first()
        if not (_team and _team.is_support_team):
            query = query.filter(Customer.team_id == current_user.team_id)
    customers = query.order_by(Customer.name).all()
    return templates.TemplateResponse(
        "customers/list.html",
        {
            "request":      request,
            "current_user": current_user,
            "customers":    customers,
            "page_title":   "Müşteri Yönetimi",
        },
    )


@router.get("/autocomplete")
async def customers_autocomplete(
    q: str = "",
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """PM için müşteri autocomplete endpoint'i"""
    query = db.query(Customer)
    # Tenant izolasyonu
    if current_user.company_id:
        query = query.filter(Customer.company_id == current_user.company_id)
    if q:
        query = query.filter(Customer.name.ilike(f"%{q}%"))
    customers = query.order_by(Customer.name).limit(20).all()
    return JSONResponse([{"id": c.id, "name": c.name, "code": c.code} for c in customers])


@router.get("/new", response_class=HTMLResponse, name="customers_new")
async def customers_new(
    request: Request,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    teams = db.query(Team).filter(Team.active == True).order_by(Team.name).all()
    return templates.TemplateResponse(
        "customers/form.html",
        {
            "request":      request,
            "current_user": current_user,
            "customer":     None,
            "page_title":   "Yeni Müşteri",
            "error":        None,
            "teams":        teams,
        },
    )


@router.post("/new", name="customers_create")
async def customers_create(
    name:          str = Form(...),
    code:          str = Form(...),
    sector:        str = Form(""),
    address:       str = Form(""),
    tax_office:    str = Form(""),
    tax_no:    str = Form(""),
    email:         str = Form(""),
    phone:         str = Form(""),
    notes:         str = Form(""),
    contacts_json: str = Form("[]"),
    payment_term:  str = Form(""),
    team_id:       str = Form(""),
    request: Request = None,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    code_clean = code.lower().strip()[:10]

    existing = db.query(Customer).filter(Customer.code == code_clean).first()
    if existing:
        teams = db.query(Team).filter(Team.active == True).order_by(Team.name).all()
        return templates.TemplateResponse(
            "customers/form.html",
            {
                "request":      request,
                "current_user": current_user,
                "customer":     None,
                "page_title":   "Yeni Müşteri",
                "error":        f"'{code_clean}' kodu zaten kullanılıyor.",
                "teams":        teams,
            },
            status_code=400,
        )

    customer = Customer(
        id=_uuid(),
        name=name.strip(),
        code=code_clean,
        sector=sector.strip(),
        address=address.strip(),
        tax_office=tax_office.strip(),
        tax_no=tax_no.strip(),
        email=email.strip(),
        phone=phone.strip(),
        notes=notes.strip(),
        contacts_json=contacts_json,
        payment_term=payment_term.strip(),
        team_id=team_id.strip() or None,
        created_at=_now(),
    )
    db.add(customer)
    db.commit()
    return RedirectResponse(url="/customers", status_code=status.HTTP_302_FOUND)


@router.get("/{customer_id}/edit", response_class=HTMLResponse, name="customers_edit")
async def customers_edit(
    customer_id: str,
    request: Request,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    from fastapi import HTTPException
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        return RedirectResponse(url="/customers", status_code=status.HTTP_302_FOUND)
    if not current_user.is_gm and current_user.role == "mudur":
        _team = db.query(Team).filter(Team.id == current_user.team_id).first() if current_user.team_id else None
        _is_support = _team and _team.is_support_team
        if not _is_support and customer.team_id and customer.team_id != current_user.team_id:
            raise HTTPException(403, "Bu müşteri takımınıza ait değil.")
    teams = db.query(Team).filter(Team.active == True).order_by(Team.name).all()
    return templates.TemplateResponse(
        "customers/form.html",
        {
            "request":      request,
            "current_user": current_user,
            "customer":     customer,
            "page_title":   f"{customer.name} — Düzenle",
            "error":        None,
            "teams":        teams,
        },
    )


@router.post("/{customer_id}/edit", name="customers_update")
async def customers_update(
    customer_id: str,
    request: Request,
    name:          str = Form(...),
    code:          str = Form(...),
    sector:        str = Form(""),
    address:       str = Form(""),
    tax_office:    str = Form(""),
    tax_no:    str = Form(""),
    email:         str = Form(""),
    phone:         str = Form(""),
    notes:         str = Form(""),
    contacts_json: str = Form("[]"),
    payment_term:  str = Form(""),
    team_id:       str = Form(""),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        return RedirectResponse(url="/customers", status_code=status.HTTP_302_FOUND)

    code_clean = code.lower().strip()[:10]
    conflict = db.query(Customer).filter(
        Customer.code == code_clean, Customer.id != customer_id
    ).first()
    if conflict:
        teams = db.query(Team).filter(Team.active == True).order_by(Team.name).all()
        return templates.TemplateResponse(
            "customers/form.html",
            {
                "request":      request,
                "current_user": current_user,
                "customer":     customer,
                "page_title":   f"{customer.name} — Düzenle",
                "error":        f"'{code_clean}' kodu başka müşteriye ait.",
                "teams":        teams,
            },
            status_code=400,
        )

    customer.name          = name.strip()
    customer.code          = code_clean
    customer.sector        = sector.strip()
    customer.address       = address.strip()
    customer.tax_office    = tax_office.strip()
    customer.tax_no    = tax_no.strip()
    customer.email         = email.strip()
    customer.phone         = phone.strip()
    customer.notes         = notes.strip()
    customer.contacts_json = contacts_json
    customer.payment_term  = payment_term.strip()
    customer.team_id       = team_id.strip() or None
    db.commit()
    return RedirectResponse(url="/customers", status_code=status.HTTP_302_FOUND)


@router.post("/{customer_id}/upload-template", name="customers_upload_template")
async def customers_upload_template(
    customer_id:   str,
    template_file: UploadFile = File(...),
    current_user:  User = Depends(require_admin),
    db:            Session = Depends(get_db),
):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        return RedirectResponse(url="/customers", status_code=status.HTTP_302_FOUND)

    ext = os.path.splitext(template_file.filename or "")[1].lower()
    if ext not in (".xlsx", ".xls"):
        return RedirectResponse(
            url=f"/customers/{customer_id}/edit?error=Sadece+.xlsx+dosyası+yüklenebilir",
            status_code=status.HTTP_302_FOUND,
        )

    if customer.excel_template_path:
        delete_upload(customer.excel_template_path)

    contents = await template_file.read()
    key = save_upload(contents, "customer_templates", f"{customer_id}{ext}")

    import base64
    customer.excel_template_path = key
    customer.excel_template_b64  = base64.b64encode(contents).decode("ascii")
    db.commit()
    return RedirectResponse(
        url=f"/customers/{customer_id}/edit?saved=template",
        status_code=status.HTTP_302_FOUND,
    )


@router.post("/{customer_id}/excel-config", name="customers_excel_config")
async def customers_excel_config(
    customer_id: str,
    vat_mode:    str = Form("exclusive"),   # exclusive | inclusive
    cell_map:    str = Form("{}"),          # JSON string
    current_user: User = Depends(require_admin),
    db:           Session = Depends(get_db),
):
    """
    Müşteriye ait Excel export ayarlarını kaydeder.
    cell_map: AI analiz sonucu veya manuel düzenleme (JSON string)
    vat_mode: 'exclusive' → KDV hariç | 'inclusive' → KDV dahil
    """
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        return RedirectResponse(url="/customers", status_code=status.HTTP_302_FOUND)

    try:
        parsed_map = json.loads(cell_map or "{}")
    except json.JSONDecodeError:
        parsed_map = {}

    # Mevcut config'i al, sadece ilgili alanları güncelle
    existing = customer.excel_config
    existing["vat_mode"] = vat_mode if vat_mode in ("exclusive", "inclusive") else "exclusive"
    if parsed_map:
        existing["cell_map"] = parsed_map

    customer.excel_config_json = json.dumps(existing, ensure_ascii=False)
    db.commit()
    return RedirectResponse(
        url=f"/customers/{customer_id}/edit?saved=config",
        status_code=status.HTTP_302_FOUND,
    )


@router.post("/{customer_id}/analyze-template", name="customers_analyze_template")
async def customers_analyze_template(
    customer_id:  str,
    current_user: User = Depends(require_admin),
    db:           Session = Depends(get_db),
):
    """
    Yüklü Excel template'ini Claude API ile analiz eder.
    Oluşan cell_map'i customer.excel_config_json'a kaydeder.
    Sonucu JSON olarak döndürür (UI önizleme için).
    """
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        return JSONResponse({"error": "Müşteri bulunamadı"}, status_code=404)

    b64_data = (getattr(customer, "excel_template_b64", None) or "")

    # Analiz için temp local dosya gerekli — b64'ten restore et
    import base64 as _b64, tempfile
    if not b64_data:
        return JSONResponse({"error": "Template dosyası yüklenmemiş veya bulunamadı"}, status_code=400)

    _tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    _tmp.write(_b64.b64decode(b64_data))
    _tmp.close()
    template_path = _tmp.name

    try:
        from excel_export import analyze_template
        result = await analyze_template(template_path=template_path)
    except Exception as exc:
        print(f"[TEMPLATE ANALYZE] error: {exc}", flush=True)
        return JSONResponse({"error": "Template analizi başarısız."}, status_code=500)
    finally:
        try:
            os.unlink(template_path)
        except Exception:
            pass

    if result.get("error"):
        return JSONResponse(result, status_code=422)

    existing = customer.excel_config
    existing["cell_map"] = result["cell_map"]
    if "vat_mode" in result["cell_map"]:
        existing["vat_mode"] = result["cell_map"].pop("vat_mode")
    customer.excel_config_json = json.dumps(existing, ensure_ascii=False)
    db.commit()

    return JSONResponse({
        "cell_map":     result["cell_map"],
        "vat_mode":     existing.get("vat_mode", "exclusive"),
        "raw_response": result.get("raw_response", ""),
        "error":        None,
    })


@router.get("/{customer_id}/template-status", name="customers_template_status")
async def customers_template_status(
    customer_id:  str,
    current_user: User = Depends(require_admin),
    db:           Session = Depends(get_db),
):
    """Müşterinin template durumunu JSON olarak döndürür (debug için)."""
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        return JSONResponse({"error": "Müşteri bulunamadı"}, status_code=404)

    b64 = getattr(customer, "excel_template_b64", "") or ""
    cfg = customer.excel_config
    cell_map = cfg.get("cell_map", {})
    template_path = customer.excel_template_path or ""
    file_exists = os.path.exists(template_path) if template_path else False

    return JSONResponse({
        "customer_name":    customer.name,
        "template_path":    template_path,
        "file_exists_disk": file_exists,
        "b64_stored":       bool(b64),
        "b64_length":       len(b64),
        "vat_mode":         cfg.get("vat_mode", "exclusive"),
        "cell_map_keys":    list(cell_map.keys()),
        "cell_map_header":  cell_map.get("header", {}),
        "cell_map_data_block": cell_map.get("data_block", {}),
        "will_use_template": bool(template_path and (file_exists or b64) and cell_map),
    })


@router.post("/{customer_id}/upload-doc", name="customers_upload_doc")
async def customers_upload_doc(
    customer_id: str,
    doc_file:    UploadFile = File(...),
    current_user: User = Depends(require_admin),
    db:           Session = Depends(get_db),
):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        return RedirectResponse(url="/customers", status_code=status.HTTP_302_FOUND)

    filename = os.path.basename(doc_file.filename or "dosya")
    key = save_upload(doc_file.file.read(), f"customer_docs/{customer_id}", filename)

    try:
        doc_list = json.loads(customer.docs_json or "[]")
    except Exception:
        doc_list = []
    doc_list.append({"name": filename, "path": key})
    customer.docs_json = json.dumps(doc_list, ensure_ascii=False)
    db.commit()
    return RedirectResponse(url=f"/customers/{customer_id}/edit", status_code=status.HTTP_302_FOUND)


@router.post("/{customer_id}/delete-doc", name="customers_delete_doc")
async def customers_delete_doc(
    customer_id: str,
    filename:    str = Form(...),
    current_user: User = Depends(require_admin),
    db:           Session = Depends(get_db),
):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        return RedirectResponse(url="/customers", status_code=status.HTTP_302_FOUND)

    try:
        doc_list = json.loads(customer.docs_json or "[]")
    except Exception:
        doc_list = []

    remaining = []
    for d in doc_list:
        if d["name"] == filename:
            delete_upload(d.get("path", ""))
        else:
            remaining.append(d)

    customer.docs_json = json.dumps(remaining, ensure_ascii=False)
    db.commit()
    return RedirectResponse(url=f"/customers/{customer_id}/edit", status_code=status.HTTP_302_FOUND)


def _read_excel_for_editor(path: str, max_rows: int = 50, max_cols: int = 20,
                           sheet_name: str | None = None) -> dict:
    """Excel dosyasını template editörü için hücre matrisi olarak okur."""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        return {"rows": [], "col_letters": [], "sheet_names": [], "error": "openpyxl kurulu değil"}

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        sheet_names = wb.sheetnames
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Merged cell bölgeleri
        merged_skip = set()   # (row, col) — ana hücre değil, atla
        merged_spans = {}     # (row, col) → {rowspan, colspan}
        for region in ws.merged_cells.ranges:
            for r in range(region.min_row, region.max_row + 1):
                for c in range(region.min_col, region.max_col + 1):
                    if r == region.min_row and c == region.min_col:
                        rs = region.max_row - region.min_row + 1
                        cs = region.max_col - region.min_col + 1
                        if rs > 1 or cs > 1:
                            merged_spans[(r, c)] = (rs, cs)
                    else:
                        merged_skip.add((r, c))

        real_max_row = min(max_rows, ws.max_row or max_rows)
        real_max_col = min(max_cols, ws.max_column or max_cols)
        col_letters = [get_column_letter(i) for i in range(1, real_max_col + 1)]

        rows = []
        for ri in range(1, real_max_row + 1):
            row_cells = []
            for ci in range(1, real_max_col + 1):
                if (ri, ci) in merged_skip:
                    continue
                cell = ws.cell(row=ri, column=ci)
                val = cell.value
                if val is None:
                    val_str = ""
                elif hasattr(val, "isoformat"):
                    val_str = str(val)
                else:
                    val_str = str(val)[:60]

                col_letter = get_column_letter(ci)
                rs, cs = merged_spans.get((ri, ci), (1, 1))
                row_cells.append({
                    "coord":   f"{col_letter}{ri}",
                    "col":     col_letter,
                    "row":     ri,
                    "value":   val_str,
                    "rowspan": rs,
                    "colspan": cs,
                })
            rows.append(row_cells)

        # Formül sütunlarını tespit et (data_only=False ile yeniden yükle)
        detected_formulas = {}
        try:
            import re as _re
            wb2 = openpyxl.load_workbook(path, data_only=False)
            ws2 = (wb2[sheet_name]
                   if sheet_name and sheet_name in wb2.sheetnames
                   else wb2.active)
            for scan_row in range(1, min(real_max_row + 1, ws2.max_row + 1)):
                row_has_formula = False
                for ci in range(1, real_max_col + 1):
                    cell2 = ws2.cell(row=scan_row, column=ci)
                    if isinstance(cell2.value, str) and cell2.value.startswith("="):
                        col_letter = get_column_letter(ci)
                        # Satır numarasını {row} ile değiştir
                        tpl = _re.sub(
                            r'(?<=[A-Za-z\$])(' + str(scan_row) + r')(?=[^0-9]|$)',
                            '{row}', cell2.value
                        )
                        detected_formulas[col_letter] = tpl
                        row_has_formula = True
                if row_has_formula:
                    break   # ilk formül satırından al
        except Exception:
            pass

        return {"rows": rows, "col_letters": col_letters,
                "detected_formulas": detected_formulas,
                "sheet_names": sheet_names,
                "active_sheet": ws.title,
                "error": None}
    except Exception as exc:
        return {"rows": [], "col_letters": [], "detected_formulas": {},
                "sheet_names": [], "error": str(exc)}


@router.get("/{customer_id}/template-editor", response_class=HTMLResponse, name="customers_template_editor")
async def customers_template_editor(
    customer_id: str,
    request: Request,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        return RedirectResponse(url="/customers", status_code=status.HTTP_302_FOUND)

    b64 = getattr(customer, "excel_template_b64", "") or ""
    if not b64:
        return RedirectResponse(
            url=f"/customers/{customer_id}/edit?error=Template+dosyası+yüklenmemiş",
            status_code=status.HTTP_302_FOUND,
        )

    import base64 as _b64, tempfile
    _tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    _tmp.write(_b64.b64decode(b64))
    _tmp.close()
    tpl_path = _tmp.name

    try:
        excel_data = _read_excel_for_editor(tpl_path)
    finally:
        try:
            os.unlink(tpl_path)
        except Exception:
            pass
    cfg = customer.excel_config

    return templates.TemplateResponse("customers/template_editor.html", {
        "request":           request,
        "current_user":      current_user,
        "customer":          customer,
        "excel_data":        excel_data,
        "existing_config":   json.dumps(cfg, ensure_ascii=False),
        "detected_formulas": json.dumps(excel_data.get("detected_formulas", {}), ensure_ascii=False),
        "sheet_names":       json.dumps(excel_data.get("sheet_names", []), ensure_ascii=False),
        "page_title":        f"{customer.name} — Şablon Eşleştirme",
    })


@router.get("/{customer_id}/template-editor/sheet-data", name="customers_template_sheet_data")
async def customers_template_sheet_data(
    customer_id: str,
    sheet: str = "",
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Belirtilen sheet'in hücre verisini JSON olarak döndürür (AJAX)."""
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        return JSONResponse({"error": "Müşteri bulunamadı"}, status_code=404)

    b64 = getattr(customer, "excel_template_b64", "") or ""
    if not b64:
        return JSONResponse({"error": "Template dosyası bulunamadı"}, status_code=404)

    import base64 as _b64, tempfile
    _tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    _tmp.write(_b64.b64decode(b64))
    _tmp.close()
    try:
        data = _read_excel_for_editor(_tmp.name, sheet_name=sheet or None)
    finally:
        try:
            os.unlink(_tmp.name)
        except Exception:
            pass
    return JSONResponse(data)


@router.post("/{customer_id}/template-editor", name="customers_template_editor_save")
async def customers_template_editor_save(
    customer_id:  str,
    config_json:  str = Form("{}"),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        return JSONResponse({"error": "Müşteri bulunamadı"}, status_code=404)
    try:
        cfg = json.loads(config_json)
    except Exception:
        return JSONResponse({"error": "Geçersiz JSON"}, status_code=400)
    customer.excel_config_json = json.dumps(cfg, ensure_ascii=False)
    db.commit()
    return JSONResponse({"ok": True})


@router.post("/{customer_id}/delete", name="customers_delete")
async def customers_delete(
    customer_id: str,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if customer:
        db.delete(customer)
        db.commit()
    return RedirectResponse(url="/customers", status_code=status.HTTP_302_FOUND)
