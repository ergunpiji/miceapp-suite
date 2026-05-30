"""
Nakit kasa yönetimi
"""

from collections import defaultdict
from datetime import date, datetime
import io
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from fastapi import APIRouter, Depends, Form, HTTPException, Query, Request, status
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from auth import get_current_user, require_admin, require_module, get_company_id
from database import get_db
from models import CashBook, CashEntry, CashDayClose, User
from templates_config import templates

router = APIRouter(prefix="/cash", tags=["cash"])


def _balance(db, book_id: str) -> float:
    ins = db.query(func.sum(CashEntry.amount)).filter(
        CashEntry.book_id == book_id, CashEntry.entry_type == "giris"
    ).scalar() or 0
    outs = db.query(func.sum(CashEntry.amount)).filter(
        CashEntry.book_id == book_id, CashEntry.entry_type == "cikis"
    ).scalar() or 0
    return ins - outs


def _balance_before(db, book_id: str, d: date) -> float:
    ins = db.query(func.sum(CashEntry.amount)).filter(
        CashEntry.book_id == book_id,
        CashEntry.entry_type == "giris",
        CashEntry.entry_date < d,
    ).scalar() or 0
    outs = db.query(func.sum(CashEntry.amount)).filter(
        CashEntry.book_id == book_id,
        CashEntry.entry_type == "cikis",
        CashEntry.entry_date < d,
    ).scalar() or 0
    return ins - outs


def _daily_summary(db, book_id: str):
    """Tüm girişleri gün bazında grupla, açılış/kapanış bakiyelerini hesapla."""
    all_entries = (
        db.query(CashEntry)
        .filter(CashEntry.book_id == book_id)
        .order_by(CashEntry.entry_date)
        .all()
    )
    closed_map = {
        dc.close_date: dc
        for dc in db.query(CashDayClose).filter(CashDayClose.book_id == book_id).all()
    }

    daily = defaultdict(lambda: {"giris": 0.0, "cikis": 0.0, "count": 0})
    for e in all_entries:
        daily[e.entry_date]["count"] += 1
        if e.entry_type == "giris":
            daily[e.entry_date]["giris"] += e.amount
        else:
            daily[e.entry_date]["cikis"] += e.amount

    rows = []
    running = 0.0
    for d in sorted(daily.keys()):
        opening = running
        g = daily[d]["giris"]
        c = daily[d]["cikis"]
        closing = opening + g - c
        close_rec = closed_map.get(d)
        rows.append({
            "date": d,
            "opening": opening,
            "giris": g,
            "cikis": c,
            "closing": closing,
            "count": daily[d]["count"],
            "close_rec": close_rec,
        })
        running = closing

    return rows, closed_map


@router.get("", response_class=HTMLResponse, name="cash_list")
async def cash_list(
    request: Request,
    current_user: User = Depends(require_module("cash")),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    books = db.query(CashBook).filter(CashBook.company_id == cid).all()
    books_with_balance = [{"book": b, "balance": _balance(db, b.id)} for b in books]
    return templates.TemplateResponse(
        "cash/list.html",
        {"request": request, "current_user": current_user,
         "books_with_balance": books_with_balance, "page_title": "Nakit Kasalar"},
    )


@router.get("/new", response_class=HTMLResponse, name="cash_new_get")
async def cash_new_get(
    request: Request,
    current_user: User = Depends(require_module("cash", edit=True)),
):
    return templates.TemplateResponse(
        "cash/book_form.html",
        {"request": request, "current_user": current_user,
         "book": None, "page_title": "Yeni Kasa"},
    )


@router.post("/new", name="cash_new_post")
async def cash_new_post(
    name: str = Form(...),
    currency: str = Form("TRY"),
    current_user: User = Depends(require_module("cash", edit=True)),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    b = CashBook(name=name.strip(), currency=currency, company_id=cid)
    db.add(b)
    db.commit()
    return RedirectResponse(url="/cash", status_code=status.HTTP_302_FOUND)


@router.get("/{book_id}", response_class=HTMLResponse, name="cash_detail")
async def cash_detail(
    book_id: str,
    request: Request,
    tab: str = "hareketler",
    type_filter: str = "",
    category_filter: str = "",
    date_from: str = "",
    date_to: str = "",
    current_user: User = Depends(require_module("cash")),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    book = db.query(CashBook).filter(CashBook.id == book_id, CashBook.company_id == cid).first()
    if not book:
        raise HTTPException(status_code=404)

    balance = _balance(db, book_id)

    # Kapalı günler seti
    closed_map = {
        dc.close_date: dc
        for dc in db.query(CashDayClose).filter(CashDayClose.book_id == book_id).all()
    }
    closed_dates = set(closed_map.keys())

    # Tüm kategoriler (filtre dropdown için)
    cats_raw = db.query(CashEntry.category).filter(
        CashEntry.book_id == book_id,
        CashEntry.category.isnot(None),
    ).distinct().all()
    all_categories = sorted([c[0] for c in cats_raw if c[0]])

    # Filtrelenmiş hareket sorgusu
    q = db.query(CashEntry).filter(CashEntry.book_id == book_id)
    if type_filter in ("giris", "cikis"):
        q = q.filter(CashEntry.entry_type == type_filter)
    if category_filter:
        q = q.filter(CashEntry.category == category_filter)
    if date_from:
        try:
            q = q.filter(CashEntry.entry_date >= date.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            q = q.filter(CashEntry.entry_date <= date.fromisoformat(date_to))
        except ValueError:
            pass

    entries = q.order_by(CashEntry.entry_date.desc(), CashEntry.id.desc()).all()

    toplam_giris = sum(e.amount for e in entries if e.entry_type == "giris")
    toplam_cikis = sum(e.amount for e in entries if e.entry_type == "cikis")

    # Günlük özet
    daily_rows, _ = _daily_summary(db, book_id)

    # Tüm hareketleri tarih bazlı grupla (popup için JSON)
    all_entries_raw = (
        db.query(CashEntry)
        .filter(CashEntry.book_id == book_id)
        .order_by(CashEntry.entry_date, CashEntry.id)
        .all()
    )
    ebd: dict = defaultdict(list)
    for e in all_entries_raw:
        ebd[e.entry_date.isoformat()].append({
            "type": e.entry_type,
            "amount": e.amount,
            "desc": e.description or "",
            "cat": e.category or "",
            "party": e.related_party or "",
        })
    entries_json = json.dumps(ebd)

    today = date.today()
    today_closed = today in closed_dates
    today_row = next((r for r in daily_rows if r["date"] == today), None)

    return templates.TemplateResponse(
        "cash/detail.html",
        {
            "request": request, "current_user": current_user,
            "book": book, "entries": entries, "balance": balance,
            "toplam_giris": toplam_giris, "toplam_cikis": toplam_cikis,
            "all_categories": all_categories,
            "type_filter": type_filter,
            "category_filter": category_filter,
            "date_from": date_from, "date_to": date_to,
            "tab": tab,
            "daily_rows": daily_rows,
            "closed_dates": closed_dates,
            "entries_json": entries_json,
            "today": today,
            "today_closed": today_closed,
            "today_str": today.isoformat(),
            "today_row": today_row,
            "page_title": f"Kasa — {book.name}",
        },
    )


@router.post("/{book_id}/entry", name="cash_entry_add")
async def cash_entry_add(
    book_id: str,
    entry_date: str = Form(...),
    entry_type: str = Form(...),
    amount: float = Form(...),
    description: str = Form(""),
    category: str = Form(""),
    related_party: str = Form(""),
    current_user: User = Depends(require_module("cash")),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    book = db.query(CashBook).filter(CashBook.id == book_id, CashBook.company_id == cid).first()
    if not book:
        raise HTTPException(status_code=404)
    d = date.fromisoformat(entry_date)
    closed = db.query(CashDayClose).filter(
        CashDayClose.book_id == book_id,
        CashDayClose.close_date == d,
    ).first()
    if closed:
        raise HTTPException(status_code=400, detail=f"{d.strftime('%d.%m.%Y')} günü kapalı — işlem yapılamaz.")
    db.add(CashEntry(
        book_id=book_id,
        entry_date=d,
        entry_type=entry_type,
        amount=amount,
        description=description.strip() or None,
        category=category.strip() or None,
        related_party=related_party.strip() or None,
        company_id=cid,
    ))
    db.commit()
    return RedirectResponse(url=f"/cash/{book_id}", status_code=status.HTTP_302_FOUND)


@router.post("/{book_id}/entry/{entry_id}/edit", name="cash_entry_edit")
async def cash_entry_edit(
    book_id: str,
    entry_id: str,
    entry_date: str = Form(...),
    entry_type: str = Form(...),
    amount: float = Form(...),
    description: str = Form(""),
    category: str = Form(""),
    related_party: str = Form(""),
    current_user: User = Depends(require_module("cash", edit=True)),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    e = db.query(CashEntry).filter(CashEntry.id == entry_id, CashEntry.company_id == cid).first()
    if not e or e.book_id != book_id:
        raise HTTPException(status_code=404)
    # Orijinal gün kapalıysa düzenleme engellendi
    closed = db.query(CashDayClose).filter(
        CashDayClose.book_id == book_id,
        CashDayClose.close_date == e.entry_date,
    ).first()
    if closed:
        raise HTTPException(status_code=400, detail="Kapalı güne ait hareket düzenlenemez.")
    e.entry_date = date.fromisoformat(entry_date)
    e.entry_type = entry_type
    e.amount = amount
    e.description = description.strip() or None
    e.category = category.strip() or None
    e.related_party = related_party.strip() or None
    db.commit()
    return RedirectResponse(url=f"/cash/{book_id}", status_code=status.HTTP_302_FOUND)


@router.post("/{book_id}/entry/{entry_id}/delete", name="cash_entry_delete")
async def cash_entry_delete(
    book_id: str,
    entry_id: str,
    current_user: User = Depends(require_module("cash", edit=True)),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    e = db.query(CashEntry).filter(CashEntry.id == entry_id, CashEntry.company_id == cid).first()
    if not e or e.book_id != book_id:
        raise HTTPException(status_code=404)
    closed = db.query(CashDayClose).filter(
        CashDayClose.book_id == book_id,
        CashDayClose.close_date == e.entry_date,
    ).first()
    if closed:
        raise HTTPException(status_code=400, detail="Kapalı güne ait hareket silinemez.")
    db.delete(e)
    db.commit()
    return RedirectResponse(url=f"/cash/{book_id}", status_code=status.HTTP_302_FOUND)


@router.get("/{book_id}/daily-export", name="cash_daily_export")
async def cash_daily_export(
    book_id: str,
    export_date: str = Query(...),
    current_user: User = Depends(require_module("cash")),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    book = db.query(CashBook).filter(CashBook.id == book_id, CashBook.company_id == cid).first()
    if not book:
        raise HTTPException(status_code=404)

    d = date.fromisoformat(export_date)
    entries = (
        db.query(CashEntry)
        .filter(CashEntry.book_id == book_id, CashEntry.entry_date == d)
        .order_by(CashEntry.id)
        .all()
    )
    opening = _balance_before(db, book_id, d)
    close_rec = db.query(CashDayClose).filter(
        CashDayClose.book_id == book_id,
        CashDayClose.close_date == d,
    ).first()

    TL_FMT = '₺ #,##0.00'

    def _tl(cell, value, bold=False):
        cell.value = value
        cell.number_format = TL_FMT
        if bold:
            cell.font = Font(bold=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = d.strftime("%d.%m.%Y")

    # Başlık satırı
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Kasa: {book.name}  —  {d.strftime('%d.%m.%Y')}"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    # Açılış bakiyesi
    ws.append([])
    ws.append(["Açılış Bakiyesi", "", "", "", "", None])
    r = ws.max_row
    ws.cell(r, 1).font = Font(bold=True)
    _tl(ws.cell(r, 6), opening, bold=True)

    # Sütun başlıkları
    ws.append([])
    col_headers = ["Tür", "Kategori", "İlgili Taraf", "Açıklama", "Giriş", "Çıkış"]
    ws.append(col_headers)
    ch_row = ws.max_row
    ch_fill = PatternFill("solid", fgColor="1E5F8C")
    for col_idx, _ in enumerate(col_headers, 1):
        cell = ws.cell(ch_row, col_idx)
        cell.fill = ch_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    total_giris = 0.0
    total_cikis = 0.0
    for e in entries:
        row_data = [
            "Giriş" if e.entry_type == "giris" else "Çıkış",
            e.category or "",
            e.related_party or "",
            e.description or "",
            None, None,
        ]
        ws.append(row_data)
        r = ws.max_row
        if e.entry_type == "giris":
            _tl(ws.cell(r, 5), e.amount)
            ws.cell(r, 5).font = Font(color="1A7A3C")
            total_giris += e.amount
        else:
            _tl(ws.cell(r, 6), e.amount)
            ws.cell(r, 6).font = Font(color="C00000")
            total_cikis += e.amount

    ws.append([])
    closing_sys = opening + total_giris - total_cikis
    summary_data = [
        ("Toplam Giriş", total_giris, "1A7A3C"),
        ("Toplam Çıkış", total_cikis, "C00000"),
        ("Kapanış Bakiyesi (Sistem)", closing_sys, None),
    ]
    if close_rec:
        summary_data += [
            ("Fiziksel Sayım", close_rec.physical_count, None),
            ("Fark", close_rec.difference, "C00000" if close_rec.difference != 0 else "1A7A3C"),
        ]
    for label, val, color in summary_data:
        ws.append([label, "", "", "", "", None])
        r = ws.max_row
        ws.cell(r, 1).font = Font(bold=True)
        ws.merge_cells(f"A{r}:E{r}")
        _tl(ws.cell(r, 6), val, bold=True)
        if color:
            ws.cell(r, 6).font = Font(bold=True, color=color)

    # Sütun genişlikleri
    for col, width in zip("ABCDEF", [10, 20, 22, 30, 16, 16]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"kasa_{book.name}_{d.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/{book_id}/close", name="cash_day_close")
async def cash_day_close(
    book_id: str,
    close_date: str = Form(...),
    physical_count: float = Form(...),
    notes: str = Form(""),
    current_user: User = Depends(require_module("cash", edit=True)),
    db: Session = Depends(get_db),
    cid: int = Depends(get_company_id),
):
    book = db.query(CashBook).filter(CashBook.id == book_id, CashBook.company_id == cid).first()
    if not book:
        raise HTTPException(status_code=404)
    d = date.fromisoformat(close_date)

    existing = db.query(CashDayClose).filter(
        CashDayClose.book_id == book_id,
        CashDayClose.close_date == d,
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Bu gün zaten kapatılmış.")

    # Kapanış bakiyesi = o güne kadar tüm hareketlerin neti
    ins = db.query(func.sum(CashEntry.amount)).filter(
        CashEntry.book_id == book_id,
        CashEntry.entry_type == "giris",
        CashEntry.entry_date <= d,
    ).scalar() or 0
    outs = db.query(func.sum(CashEntry.amount)).filter(
        CashEntry.book_id == book_id,
        CashEntry.entry_type == "cikis",
        CashEntry.entry_date <= d,
    ).scalar() or 0
    closing_balance = ins - outs
    opening_balance = _balance_before(db, book_id, d)

    dc = CashDayClose(
        book_id=book_id,
        close_date=d,
        opening_balance=opening_balance,
        closing_balance=closing_balance,
        physical_count=physical_count,
        difference=physical_count - closing_balance,
        notes=notes.strip() or None,
        closed_by=current_user.id,
        closed_at=datetime.utcnow(),
        company_id=cid,
    )
    db.add(dc)
    db.commit()
    return RedirectResponse(url=f"/cash/{book_id}?tab=gunluk", status_code=status.HTTP_302_FOUND)
