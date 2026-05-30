"""
E-dem — Hizmet kataloğu router'ı (Admin only)
"""

import io
import os

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile, status
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session

from auth import get_current_user, require_admin
from database import get_db
from models import CustomCategory, Service, SERVICE_CATEGORIES, User, _uuid

router = APIRouter(prefix="/services", tags=["services"])
from templates_config import templates

# Geçerli kategori id'leri
_CAT_IDS = [c["id"] for c in SERVICE_CATEGORIES]


@router.get("", response_class=HTMLResponse, name="services_list")
async def services_list(
    request: Request,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    services    = db.query(Service).order_by(Service.category, Service.sort_order, Service.name).all()
    custom_cats = db.query(CustomCategory).all()

    # Servisleri kategoriye göre grupla
    grouped: dict = {}
    for svc in services:
        grouped.setdefault(svc.category, []).append(svc)

    return templates.TemplateResponse(
        "services/list.html",
        {
            "request":           request,
            "current_user":      current_user,
            "grouped_services":  grouped,
            "service_categories": SERVICE_CATEGORIES,
            "custom_categories": custom_cats,
            "page_title":        "Hizmet Kataloğu",
        },
    )


@router.get("/api", name="services_api")
async def services_api(
    category: str = "",
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Katalogdan ekle için JSON endpoint"""
    query = db.query(Service).filter(Service.active == True)
    if category:
        query = query.filter(Service.category == category)
    services = query.order_by(Service.name).all()
    return JSONResponse([s.to_dict() for s in services])


@router.post("/new", name="services_create")
async def services_create(
    category:    str = Form(...),
    name:        str = Form(...),
    unit:        str = Form("Adet"),
    description: str = Form(""),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    svc = Service(id=_uuid(), category=category, name=name.strip(), unit=unit.strip(), description=description.strip(), active=True)
    db.add(svc)
    db.commit()
    return RedirectResponse(url="/services", status_code=status.HTTP_302_FOUND)


@router.post("/{svc_id}/toggle", name="services_toggle")
async def services_toggle(
    svc_id: str,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    svc = db.query(Service).filter(Service.id == svc_id).first()
    if svc:
        svc.active = not svc.active
        db.commit()
    return RedirectResponse(url="/services", status_code=status.HTTP_302_FOUND)


@router.post("/{svc_id}/delete", name="services_delete")
async def services_delete(
    svc_id: str,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    svc = db.query(Service).filter(Service.id == svc_id).first()
    if svc:
        db.delete(svc)
        db.commit()
    return RedirectResponse(url="/services", status_code=status.HTTP_302_FOUND)


# ---------------------------------------------------------------------------
# Özel Kategoriler
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Excel şablon indir + toplu içe aktar
# ---------------------------------------------------------------------------

@router.post("/reorder", name="services_reorder")
async def services_reorder(
    request: Request,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Sürükle-bırak sıra kaydı. Body: {"ids": ["id1","id2",...]}"""
    from fastapi.responses import JSONResponse as _JSON
    body = await request.json()
    ids: list[str] = body.get("ids", [])
    for order, svc_id in enumerate(ids):
        svc = db.query(Service).filter(Service.id == svc_id).first()
        if svc:
            svc.sort_order = order
    db.commit()
    return _JSON({"ok": True})


@router.get("/template", name="services_template")
async def services_template(current_user: User = Depends(require_admin)):
    """Hizmet kataloğu Excel şablonu indir."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hizmetler"

    # Başlık satırı
    headers = ["Kategori", "Hizmet Adı", "Birim", "Aktif (E/H)"]
    header_fill = PatternFill("solid", fgColor="1e3a5f")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    # Kategori dropdown doğrulama
    cat_list = ",".join(_CAT_IDS)
    dv_cat = DataValidation(type="list", formula1=f'"{cat_list}"', allow_blank=False)
    ws.add_data_validation(dv_cat)
    dv_cat.sqref = "A2:A1000"

    # Aktif dropdown
    dv_active = DataValidation(type="list", formula1='"E,H"', allow_blank=False)
    ws.add_data_validation(dv_active)
    dv_active.sqref = "D2:D1000"

    # Örnek satırlar
    examples = [
        ("accommodation", "Standart Oda DBL", "Gece", "E"),
        ("meeting",       "Salon Kirası",     "Gün",  "E"),
        ("fb",            "Kahvaltı",         "Kişi", "E"),
    ]
    for r, row in enumerate(examples, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    # Kategori referans sekmesi
    ws2 = wb.create_sheet("Kategoriler")
    ws2.cell(row=1, column=1, value="Kategori ID").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Kategori Adı").font = Font(bold=True)
    for r, cat in enumerate(SERVICE_CATEGORIES, 2):
        ws2.cell(row=r, column=1, value=cat["id"])
        ws2.cell(row=r, column=2, value=cat["label"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=hizmet_sablonu.xlsx"},
    )


@router.post("/import", name="services_import")
async def services_import(
    file: UploadFile = File(...),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Excel'den hizmet toplu içe aktar."""
    import openpyxl

    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in {".xlsx", ".xls"}:
        return RedirectResponse(url="/services?import_msg=Hata:+Sadece+.xlsx+dosyaları+yüklenebilir", status_code=303)

    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        return RedirectResponse(url="/services?import_msg=Hata:+Dosya+5+MB+limitini+aşıyor", status_code=303)
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    added = skipped = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        cat, name, unit, active_str = (row + (None, None, None, None))[:4]
        cat   = str(cat or "").strip()
        name  = str(name or "").strip()
        unit  = str(unit or "Adet").strip()
        active = str(active_str or "E").strip().upper() != "H"

        if not name:
            errors.append(f"Satır {row_idx}: Hizmet adı boş")
            skipped += 1
            continue
        if cat not in _CAT_IDS:
            errors.append(f"Satır {row_idx}: Geçersiz kategori '{cat}'")
            skipped += 1
            continue

        # Aynı kategori + isim varsa atla
        exists = db.query(Service).filter(
            Service.category == cat, Service.name == name
        ).first()
        if exists:
            skipped += 1
            continue

        db.add(Service(id=_uuid(), category=cat, name=name, unit=unit, active=active))
        added += 1

    db.commit()

    msg = f"{added} hizmet eklendi"
    if skipped:
        msg += f", {skipped} atlandı"
    if errors:
        msg += f". Hatalar: " + " | ".join(errors[:5])

    return RedirectResponse(url=f"/services?import_msg={msg}", status_code=status.HTTP_302_FOUND)


@router.post("/categories/new", name="custom_cat_create")
async def custom_cat_create(
    name:      str = Form(...),
    icon:      str = Form("📋"),
    bg_color:  str = Form("#e0f2fe"),
    txt_color: str = Form("#0c4a6e"),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    cat = CustomCategory(id=_uuid(), name=name.strip(), icon=icon, bg_color=bg_color, txt_color=txt_color)
    db.add(cat)
    db.commit()
    return RedirectResponse(url="/services", status_code=status.HTTP_302_FOUND)


@router.post("/categories/{cat_id}/delete", name="custom_cat_delete")
async def custom_cat_delete(
    cat_id: str,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    cat = db.query(CustomCategory).filter(CustomCategory.id == cat_id).first()
    if cat:
        db.delete(cat)
        db.commit()
    return RedirectResponse(url="/services", status_code=status.HTTP_302_FOUND)
