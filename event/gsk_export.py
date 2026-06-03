"""
gsk_export.py — GSK şablon doldurma modülü (event app)

Yüksek seviyeli kullanım:
    from gsk_export import gsk_doldur
    xlsx_bytes = gsk_doldur(data=teklif, hekim=31, staff=4,
                             yetkili="Ad Soyad", sablon_yolu=SABLON_YOLU)

Düşük seviyeli kullanım:
    from gsk_export import fill_gsk_template, LineItem
    result = fill_gsk_template(template_path, output_path,
                               items_by_section=..., header=...,
                               commission_rate=0.055)
"""
from __future__ import annotations

import io
import os
from dataclasses import dataclass
from typing import Optional

from openpyxl import load_workbook
from openpyxl.cell import MergedCell


# ---------------------------------------------------------------------------
# Şablon yapısı
# ---------------------------------------------------------------------------

GSK_SECTIONS: dict[str, dict] = {
    "hekim_yiyecek":       {"label": "Hekim Yiyecek",       "rows": (13, 14, 15),                "vat_cell": "J12"},
    "hekim_icecek":        {"label": "Hekim İçecek",         "rows": (17,),                       "vat_cell": "J16"},
    "staff_yiyecek":       {"label": "Staff Yiyecek",        "rows": (20, 21),                    "vat_cell": "J19"},
    "staff_icecek":        {"label": "Staff İçecek",          "rows": (23,),                       "vat_cell": "J22"},
    "konusmaci_konaklama": {"label": "Konuşmacı Konaklama",  "rows": (26, 27),                    "vat_cell": "J25"},
    "konusmaci_ulasim":    {"label": "Konuşmacı Ulaşım",     "rows": (30,),                       "vat_cell": "J29"},
    "diger_hizmetler":     {"label": "Diğer Hizmetler",      "rows": (33, 34, 35, 36, 37, 38, 39), "vat_cell": "J32"},
}

SECTION_LABELS = {k: v["label"] for k, v in GSK_SECTIONS.items()}

COMMISSION_CELL = "B10"

DEFAULT_VAT_RATES: dict[str, float] = {
    "J12": 0.10,
    "J16": 0.20,
    "J19": 0.10,
    "J22": 0.20,
    "J25": 0.10,
    "J29": 0.20,
    "J32": 0.20,
}

HEADER_CELLS: dict[str, str] = {
    "toplanti_adi":   "B2",
    "tarih":          "B3",
    "opsiyon_tarihi": "B4",
    "saat":           "B5",
    "mekan":          "B6",
    "acente":         "B7",
    "gsk_grup":       "B8",
    "yetkili":        "B9",
}

DATA_COLS = "ABCDEFGHIJ"

# F&B algılama anahtar kelimeleri
_FB_WORDS    = ("yemek", "yiyecek", "kahvaltı", "kahvalti", "öğle", "ogle",
                "akşam", "aksam", "gala", "meze", "kokteyl", "coffee",
                "içecek", "icecek", "drink", "brunch", "tabldot")
_DRINK_WORDS = ("içecek", "icecek", "drink", "coffee", "su ikramı", "su ikami")


@dataclass
class LineItem:
    description: str
    unit_price:  float
    quantity:    float = 1.0
    days:        float = 1.0
    rate:        float = 1.0


class GSKOverflowError(Exception):
    def __init__(self, overflow: list[tuple[str, int, int]]):
        self.overflow = overflow
        detail = "; ".join(f"{name}: {n} kalem > {cap} satır" for name, n, cap in overflow)
        super().__init__(f"Bölüm satır kapasitesi aşıldı → {detail}")


# ---------------------------------------------------------------------------
# Yardımcı
# ---------------------------------------------------------------------------

def _abs(cell: str) -> str:
    col = "".join(c for c in cell if c.isalpha())
    row = "".join(c for c in cell if c.isdigit())
    return f"${col}${row}"


_COMM_ABS = _abs(COMMISSION_CELL)


def _safe_set(ws, cell_ref: str, value) -> None:
    cell = ws[cell_ref]
    if not isinstance(cell, MergedCell):
        cell.value = value


def _write_item(ws, r: int, item: LineItem, vat_abs: str) -> None:
    _safe_set(ws, f"A{r}", item.description)
    _safe_set(ws, f"B{r}", item.unit_price)
    _safe_set(ws, f"C{r}", item.rate)
    _safe_set(ws, f"D{r}", f"=B{r}*C{r}")
    _safe_set(ws, f"E{r}", f"=B{r}*(1+{_COMM_ABS})")
    _safe_set(ws, f"F{r}", item.quantity)
    _safe_set(ws, f"G{r}", item.days)
    _safe_set(ws, f"H{r}", f"=B{r}*F{r}*G{r}")
    _safe_set(ws, f"I{r}", f"=E{r}*F{r}*G{r}")
    _safe_set(ws, f"J{r}", f"=H{r}*(1+{vat_abs})+(I{r}-H{r})*1.2")


def _clear_row(ws, r: int) -> None:
    for col in DATA_COLS:
        _safe_set(ws, f"{col}{r}", None)


# ---------------------------------------------------------------------------
# Otomatik bölümleme
# ---------------------------------------------------------------------------

def rows_to_items(
    rows: list[dict],
    hekim: float,
    staff: float,
    price_overrides: dict[str, float] | None = None,
) -> tuple[dict[str, list[LineItem]], list[str]]:
    """
    Budget satırlarını GSK bölümlerine otomatik dağıtır.

    rows: [{"id", "section", "description", "sale_price", "qty", "nights"}, ...]
    Döner: (items_by_section, uyarılar)
    """
    price_overrides = price_overrides or {}
    raw: dict[str, list[LineItem]] = {k: [] for k in GSK_SECTIONS}
    warnings: list[str] = []

    for r in rows:
        row_id   = r.get("id", "")
        sec      = (r.get("section") or "").lower()
        desc_low = (r.get("description") or "").lower()
        desc     = r.get("description", "")
        price    = price_overrides.get(row_id, float(r.get("sale_price") or 0))
        qty      = float(r.get("qty") or 1)
        nights   = float(r.get("nights") or 1)

        is_accom    = sec == "accommodation" or any(w in desc_low for w in ("konaklama", "otel", "oda"))
        is_transfer = sec == "transfer"      or any(w in desc_low for w in ("transfer", "ulaşım", "ulasim", "araç", "arac"))
        is_drink    = any(w in desc_low for w in _DRINK_WORDS)
        is_fb       = sec in ("fb", "f&b")   or any(w in desc_low for w in _FB_WORDS)

        if is_accom:
            raw["konusmaci_konaklama"].append(LineItem(desc, price, qty, nights))
        elif is_transfer:
            raw["konusmaci_ulasim"].append(LineItem(desc, price, qty, nights))
        elif is_fb:
            explicit_hekim = "hekim" in desc_low
            explicit_staff = "staff" in desc_low
            if explicit_hekim:
                gsk_sec = "hekim_icecek" if is_drink else "hekim_yiyecek"
                raw[gsk_sec].append(LineItem(desc, price, qty, nights))
            elif explicit_staff:
                gsk_sec = "staff_icecek" if is_drink else "staff_yiyecek"
                raw[gsk_sec].append(LineItem(desc, price, qty, nights))
            else:
                if hekim > 0:
                    gsk_sec = "hekim_icecek" if is_drink else "hekim_yiyecek"
                    raw[gsk_sec].append(LineItem(desc, price, hekim, nights))
                if staff > 0:
                    gsk_sec = "staff_icecek" if is_drink else "staff_yiyecek"
                    raw[gsk_sec].append(LineItem(desc, price, staff, nights))
                if hekim == 0 and staff == 0:
                    gsk_sec = "hekim_icecek" if is_drink else "hekim_yiyecek"
                    raw[gsk_sec].append(LineItem(desc, price, qty, nights))
        else:
            raw["diger_hizmetler"].append(LineItem(desc, price, qty, nights))

    # Taşma: kapasiteyi aşan kalemler diger_hizmetler'e
    overflow: list[LineItem] = []
    items: dict[str, list[LineItem]] = {}
    for key, sec_def in GSK_SECTIONS.items():
        if key == "diger_hizmetler":
            continue
        cap = len(sec_def["rows"])
        sec_items = raw[key]
        if len(sec_items) > cap:
            warnings.append(
                f"{sec_def['label']}: {len(sec_items)} kalem, max {cap} → "
                f"{len(sec_items) - cap} kalem 'Diğer Hizmetler'e taşındı"
            )
            overflow.extend(sec_items[cap:])
            items[key] = sec_items[:cap]
        else:
            items[key] = sec_items

    diger = raw["diger_hizmetler"] + overflow
    diger_cap = len(GSK_SECTIONS["diger_hizmetler"]["rows"])
    if len(diger) > diger_cap:
        warnings.append(f"Diğer Hizmetler kapasitesi ({diger_cap}) aşıldı")
        diger = diger[:diger_cap]
    items["diger_hizmetler"] = diger

    return items, warnings


# ---------------------------------------------------------------------------
# Yüksek seviyeli API
# ---------------------------------------------------------------------------

def gsk_doldur(
    data: dict,
    hekim: int,
    staff: int,
    yetkili: str,
    sablon_yolu: str,
    commission_rate: float = 0.055,
    price_overrides: dict[str, float] | None = None,
) -> bytes:
    """
    GSK şablonunu doldurur ve xlsx bytes döner.

    data sözlüğü:
        toplanti_adi, tarih, opsiyon_tarihi, saat, mekan,
        gsk_grup, acente  → header alanları
        rows              → list[dict]  bütçe kalemleri

    Örnek:
        xlsx = gsk_doldur(
            data={"toplanti_adi": "...", "mekan": "...", "rows": [...]},
            hekim=31, staff=4,
            yetkili="Ad Soyad",
            sablon_yolu=SABLON_YOLU,
        )
        with open("teklif.xlsx", "wb") as f:
            f.write(xlsx)
    """
    rows = data.get("rows") or []
    items, _ = rows_to_items(rows, float(hekim), float(staff), price_overrides)

    header = {
        "toplanti_adi":   data.get("toplanti_adi", ""),
        "tarih":          data.get("tarih", ""),
        "opsiyon_tarihi": data.get("opsiyon_tarihi", ""),
        "saat":           data.get("saat", ""),
        "mekan":          data.get("mekan", ""),
        "acente":         data.get("acente", "STOK MICE"),
        "gsk_grup":       data.get("gsk_grup", ""),
        "yetkili":        yetkili,
    }

    # Belleğe yaz (disk I/O yok)
    wb = _fill_workbook(sablon_yolu, items, header, commission_rate)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Düşük seviyeli: dosyaya yaz
# ---------------------------------------------------------------------------

def fill_gsk_template(
    template_path: str,
    output_path: str,
    *,
    items_by_section: dict[str, list[LineItem]],
    header: Optional[dict[str, str]] = None,
    commission_rate: float = 0.055,
    vat_rates: Optional[dict[str, float]] = None,
    gsk_sheet_name: Optional[str] = None,
) -> dict:
    """Şablonu doldurur, output_path'e kaydeder. Dict döner."""
    vat_rates = {**DEFAULT_VAT_RATES, **(vat_rates or {})}

    unknown = set(items_by_section) - set(GSK_SECTIONS)
    if unknown:
        raise KeyError(f"Bilinmeyen bölüm: {sorted(unknown)}. Geçerli: {sorted(GSK_SECTIONS)}")

    overflow = [
        (GSK_SECTIONS[k]["label"], len(v), len(GSK_SECTIONS[k]["rows"]))
        for k, v in items_by_section.items()
        if len(v) > len(GSK_SECTIONS[k]["rows"])
    ]
    if overflow:
        raise GSKOverflowError(overflow)

    wb = _fill_workbook(template_path, items_by_section, header, commission_rate,
                        vat_rates=vat_rates, sheet_name=gsk_sheet_name)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)

    filled_rows = [
        (GSK_SECTIONS[k]["label"], r)
        for k, items in items_by_section.items()
        for idx, r in enumerate(GSK_SECTIONS[k]["rows"])
        if idx < len(items)
    ]
    empty_sections = [
        GSK_SECTIONS[k]["label"]
        for k, items in items_by_section.items()
        if not items
    ]
    return {
        "filled_rows":    filled_rows,
        "empty_sections": empty_sections,
        "totals":         compute_totals(items_by_section, commission_rate, vat_rates),
        "output_path":    output_path,
    }


def _fill_workbook(
    template_path: str,
    items_by_section: dict[str, list[LineItem]],
    header: Optional[dict[str, str]],
    commission_rate: float,
    vat_rates: Optional[dict[str, float]] = None,
    sheet_name: Optional[str] = None,
):
    vat_rates = {**DEFAULT_VAT_RATES, **(vat_rates or {})}
    wb = load_workbook(template_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    if header:
        for field, cell in HEADER_CELLS.items():
            if field in header and header[field] is not None:
                _safe_set(ws, cell, header[field])

    _safe_set(ws, COMMISSION_CELL, commission_rate)
    try:
        if not isinstance(ws[COMMISSION_CELL], MergedCell):
            ws[COMMISSION_CELL].number_format = "0.0%"
    except Exception:
        pass

    for cell_ref, rate in vat_rates.items():
        _safe_set(ws, cell_ref, rate)
        try:
            if not isinstance(ws[cell_ref], MergedCell):
                ws[cell_ref].number_format = "0%"
        except Exception:
            pass

    for key, sec in GSK_SECTIONS.items():
        items = items_by_section.get(key) or []
        vat_abs = _abs(sec["vat_cell"])
        for idx, r in enumerate(sec["rows"]):
            if idx < len(items):
                _write_item(ws, r, items[idx], vat_abs)
            else:
                _clear_row(ws, r)

    return wb


def compute_totals(
    items_by_section: dict[str, list[LineItem]],
    commission_rate: float,
    vat_rates: Optional[dict[str, float]] = None,
) -> dict:
    vat_rates = {**DEFAULT_VAT_RATES, **(vat_rates or {})}
    h = i = j = 0.0
    per_section: dict[str, dict] = {}
    for key, sec in GSK_SECTIONS.items():
        items = items_by_section.get(key) or []
        vat = vat_rates.get(sec["vat_cell"], 0.0)
        sh = si = sj = 0.0
        for it in items:
            _h = it.unit_price * it.quantity * it.days
            _e = it.unit_price * (1 + commission_rate)
            _i = _e * it.quantity * it.days
            _j = _h * (1 + vat) + (_i - _h) * 1.2
            sh += _h; si += _i; sj += _j
        per_section[key] = {"H": round(sh, 2), "I": round(si, 2), "J": round(sj, 2)}
        h += sh; i += si; j += sj
    return {
        "H41_servis_haric_kdv_haric": round(h, 2),
        "I41_servis_dahil_kdv_haric": round(i, 2),
        "J41_servis_dahil_kdv_dahil": round(j, 2),
        "per_section": per_section,
    }
